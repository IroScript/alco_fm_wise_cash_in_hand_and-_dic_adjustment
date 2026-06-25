import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import re

# ── 50K AD Design System ──────────────────────────────
FONT_FAMILY   = 'Aptos'

C_VOID        = '060816'
C_DEEP_NAVY   = '0D1425'
C_MIDNIGHT    = '1E293B'
C_DARK_SURF   = '0F172A'
C_ZEBRA_A     = 'FFFFFF'
C_ZEBRA_B     = 'F8FAFC'
C_TOTAL_DATA  = 'ECFDF5'
C_TOTAL_HEAD  = '065F46'

T_NEON        = '00F2FE'
T_WHITE       = 'FFFFFF'
T_SLATE       = 'CBD5E1'
T_INK         = '0F172A'
T_MINT        = 'A7F3D0'
T_TOTAL_DARK  = '064E3B'
T_DATE        = '475569'

B_NEON        = '0E7490'
B_SLATE       = '334155'
B_LIGHT       = 'E2E8F0'
B_TOTAL       = '6EE7B7'

# Fonts
font_title    = Font(name=FONT_FAMILY, size=18, bold=True,  color=T_NEON)
font_hdr      = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
font_sub      = Font(name=FONT_FAMILY, size=9,  bold=True,  color=T_SLATE)
font_name     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
font_date     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_DATE)
font_body     = Font(name=FONT_FAMILY, size=10, bold=False, color=T_INK)
font_total_h  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_MINT)
font_total_d  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_TOTAL_DARK)

# Fills
fill_void     = PatternFill('solid', start_color=C_VOID,       end_color=C_VOID)
fill_navy     = PatternFill('solid', start_color=C_DEEP_NAVY,  end_color=C_DEEP_NAVY)
fill_mid      = PatternFill('solid', start_color=C_MIDNIGHT,   end_color=C_MIDNIGHT)
fill_dark     = PatternFill('solid', start_color=C_DARK_SURF,  end_color=C_DARK_SURF)
fill_za       = PatternFill('solid', start_color=C_ZEBRA_A,    end_color=C_ZEBRA_A)
fill_zb       = PatternFill('solid', start_color=C_ZEBRA_B,    end_color=C_ZEBRA_B)
fill_tot_data = PatternFill('solid', start_color=C_TOTAL_DATA, end_color=C_TOTAL_DATA)
fill_tot_head = PatternFill('solid', start_color=C_TOTAL_HEAD, end_color=C_TOTAL_HEAD)

# Border helpers
def side(style='thin', color=B_LIGHT):
    return Side(style=style, color=color)

bd_data  = Border(left=side(), right=side(), top=side(), bottom=side())
bd_hdr   = Border(left=side('thin', B_SLATE), right=side('thin', B_SLATE),
                  top=side('thin', B_SLATE),  bottom=side('thin', B_SLATE))
bd_total = Border(left=side('thin', B_TOTAL), right=side('thin', B_TOTAL),
                  top=side('thin', B_TOTAL),  bottom=side('thin', B_TOTAL))
bd_date  = Border(left=side('medium', B_SLATE), right=side('thin', B_LIGHT),
                  top=side('thin', B_LIGHT),   bottom=side('thin', B_LIGHT))

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_right  = Alignment(horizontal='right', vertical='center')

# 31 JUN'26 down to 1 JUN'26 (matching sample sheet)
dates = [f"{d} JUN'26" for d in range(31, 0, -1)]

def clean_person_name(name):
    if not name:
        return ""
    name = str(name).strip()
    name = re.sub(r'^(MR|MD|MRS|MST|DR)\.?\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(MR|MD|MRS|MST|DR)\.?\s+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', ' ', name)
    return name.strip().upper()

def build_fm_block_sheet1(ws, start_col, fm_name, fm_data, zone):
    markets = fm_data['markets']
    
    # Get unique DAs associated with this FM
    das = []
    for m in markets:
        if m['da_name'] and str(m['da_name']).strip().upper() != 'VACANT':
            da_str = str(m['da_name']).strip().upper()
            if da_str not in das:
                das.append(da_str)
                
    num_mpos = len(markets)
    num_das = len(das)
    
    block_cols = 1 + 1 + num_mpos + num_das + 1
    total_col = start_col + 2 + num_mpos + num_das
    
    # Merged title
    ws.merge_cells(start_row=2, start_column=start_col + 1, end_row=3, end_column=start_col + block_cols)
    title_cell = ws.cell(row=2, column=start_col + 1, value="CASH IN HAND")
    title_cell.font = font_title
    title_cell.fill = fill_void
    title_cell.alignment = align_center
    
    # Apply border/fill to title block
    for r in range(2, 4):
        for c in range(start_col + 1, start_col + block_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_void
            left = side('medium', B_NEON) if c == start_col + 1 else None
            right = side('medium', B_NEON) if c == start_col + block_cols else None
            top = side('medium', B_NEON) if r == 2 else None
            bottom = side('medium', B_NEON) if r == 3 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            
    mpo_start = start_col + 2
    mpo_end = start_col + 1 + num_mpos
    da_start = start_col + 2 + num_mpos
    da_end = start_col + 1 + num_mpos + num_das
    
    ws.merge_cells(start_row=5, start_column=mpo_start, end_row=5, end_column=mpo_end)
    c_mpo = ws.cell(row=5, column=mpo_start, value="MPO")
    c_mpo.font = font_hdr
    c_mpo.fill = fill_mid
    c_mpo.alignment = align_center
    
    if num_das > 0:
        ws.merge_cells(start_row=5, start_column=da_start, end_row=5, end_column=da_end)
        c_da = ws.cell(row=5, column=da_start, value="DA")
        c_da.font = font_hdr
        c_da.fill = fill_mid
        c_da.alignment = align_center
        
    ws.merge_cells(start_row=5, start_column=total_col, end_row=11, end_column=total_col)
    c_tot_header = ws.cell(row=5, column=total_col, value="TOTAL CASH IN HAND")
    c_tot_header.font = font_total_h
    c_tot_header.fill = fill_tot_head
    c_tot_header.alignment = align_center
    
    # Fill remaining cells in headers
    for r in range(5, 12):
        ws.cell(row=r, column=total_col).fill = fill_tot_head
        ws.cell(row=r, column=total_col).border = bd_total
    for c in range(mpo_start, mpo_end + 1):
        ws.cell(row=5, column=c).fill = fill_mid
        ws.cell(row=5, column=c).border = bd_hdr
    if num_das > 0:
        for c in range(da_start, da_end + 1):
            ws.cell(row=5, column=c).fill = fill_mid
            ws.cell(row=5, column=c).border = bd_hdr
            
    # Populate metadata rows (Row 6 to 10)
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        ws.cell(row=6, column=c_idx, value=zone).font = font_sub
        ws.cell(row=7, column=c_idx, value=fm_name).font = font_sub
        ws.cell(row=8, column=c_idx, value=m['market_name']).font = font_sub
        ws.cell(row=9, column=c_idx, value=m['mpo_code']).font = font_sub
        ws.cell(row=10, column=c_idx, value=m['fm_code']).font = font_sub
        
        for r in range(6, 11):
            ws.cell(row=r, column=c_idx).fill = fill_navy
            ws.cell(row=r, column=c_idx).border = bd_hdr
            ws.cell(row=r, column=c_idx).alignment = align_center
            
    # Row 11 Name Row
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        ws.cell(row=11, column=c_idx, value=m['mpo_name']).font = font_name
        ws.cell(row=11, column=c_idx).fill = fill_dark
        ws.cell(row=11, column=c_idx).border = bd_hdr
        ws.cell(row=11, column=c_idx).alignment = align_center
        
    for d_idx, da_name in enumerate(das):
        c_idx = da_start + d_idx
        ws.cell(row=11, column=c_idx, value=da_name).font = font_name
        ws.cell(row=11, column=c_idx).fill = fill_dark
        ws.cell(row=11, column=c_idx).border = bd_hdr
        ws.cell(row=11, column=c_idx).alignment = align_center
        
    # DATE
    ws.merge_cells(start_row=11, start_column=start_col, end_row=14, end_column=start_col)
    c_date_h = ws.cell(row=11, column=start_col, value="DATE")
    c_date_h.font = font_hdr
    c_date_h.fill = fill_mid
    c_date_h.alignment = align_center
    
    # FM SELF
    ws.merge_cells(start_row=11, start_column=start_col + 1, end_row=14, end_column=start_col + 1)
    c_fm_h = ws.cell(row=11, column=start_col + 1, value="FM SELF")
    c_fm_h.font = font_hdr
    c_fm_h.fill = fill_mid
    c_fm_h.alignment = align_center
    
    for r in range(11, 15):
        ws.cell(row=r, column=start_col).fill = fill_mid
        ws.cell(row=r, column=start_col).border = bd_hdr
        ws.cell(row=r, column=start_col + 1).fill = fill_mid
        ws.cell(row=r, column=start_col + 1).border = bd_hdr
        
    # Metadata rows 12 to 14
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        ws.cell(row=12, column=c_idx, value=m['desig']).font = font_sub
        ws.cell(row=13, column=c_idx, value=m['is_vacant']).font = font_sub
        ws.cell(row=14, column=c_idx, value=m['da_name']).font = font_sub
        for r in range(12, 15):
            ws.cell(row=r, column=c_idx).fill = fill_navy
            ws.cell(row=r, column=c_idx).border = bd_hdr
            ws.cell(row=r, column=c_idx).alignment = align_center
            
    # Rows 15 to 17 formulas
    for r in range(15, 18):
        c_tot = ws.cell(row=r, column=total_col, value=f"=SUM({get_column_letter(start_col + 1)}{r}:{get_column_letter(total_col-1)}{r})")
        c_tot.font = font_total_d
        c_tot.alignment = align_center
        c_tot.fill = fill_tot_data
        c_tot.border = bd_total
        for c in range(start_col, total_col):
            ws.cell(row=r, column=c).border = bd_data
            
    # Rows 18 to 48 data
    for r_idx, date_val in enumerate(dates):
        row_num = 18 + r_idx
        row_fill = fill_za if (r_idx % 2 == 0) else fill_zb
        
        c_val = ws.cell(row=row_num, column=start_col, value=date_val)
        c_val.font = font_date
        c_val.alignment = align_center
        c_val.border = bd_date
        c_val.fill = row_fill
        
        for c in range(start_col + 1, total_col):
            cell = ws.cell(row=row_num, column=c)
            cell.border = bd_data
            cell.alignment = align_right
            cell.number_format = '#,##0'
            cell.font = font_body
            cell.fill = row_fill
            
        c_tot = ws.cell(row=row_num, column=total_col, value=f"=SUM({get_column_letter(start_col + 1)}{row_num}:{get_column_letter(total_col-1)}{row_num})")
        c_tot.font = font_total_d
        c_tot.alignment = align_right
        c_tot.border = bd_total
        c_tot.fill = fill_tot_data
        c_tot.number_format = '#,##0'
        
    # Hide vacant columns
    for m_idx, m in enumerate(markets):
        if m['is_vacant'] == 'Y':
            col_letter = get_column_letter(mpo_start + m_idx)
            ws.column_dimensions[col_letter].hidden = True
            
    # Widths
    ws.column_dimensions[get_column_letter(start_col)].width = 14  # DATE
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 13  # FM SELF
    for m_idx, m in enumerate(markets):
        c_let = get_column_letter(mpo_start + m_idx)
        name_len = len(m['mpo_name'] or '')
        ws.column_dimensions[c_let].width = max(name_len * 1.1, 13)
    for d_idx, da_name in enumerate(das):
        c_let = get_column_letter(da_start + d_idx)
        name_len = len(da_name or '')
        ws.column_dimensions[c_let].width = max(name_len * 1.1, 13)
    ws.column_dimensions[get_column_letter(total_col)].width = 18  # TOTAL
    
    return block_cols

def generate_zonal_file(zone, zone_fms, output_path):
    wb = openpyxl.Workbook()
    
    # ── SHEET 1: Sheet1 (FMs side-by-side) ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.views.sheetView[0].showGridLines = True
    ws1.sheet_view.zoomScale = 90
    
    ws1.row_dimensions[1].height = 10
    ws1.row_dimensions[2].height = 28
    ws1.row_dimensions[3].height = 28
    ws1.row_dimensions[4].height = 8
    ws1.row_dimensions[5].height = 22
    ws1.row_dimensions[11].height = 26
    for r in range(18, 49):
        ws1.row_dimensions[r].height = 20
        
    ws1.freeze_panes = 'D18'  # Freeze D18 (frozen DATE + headers)
    
    ws1.column_dimensions['A'].width = 3
    
    curr_col = 2
    for fm_name, fm_data in zone_fms.items():
        fm_clean_name = fm_name.split(',')[0].strip()
        cols_used = build_fm_block_sheet1(ws1, curr_col, fm_clean_name, fm_data, zone)
        
        # Spacer column after this FM
        spacer_col = curr_col + cols_used
        ws1.column_dimensions[get_column_letter(spacer_col)].width = 3
        curr_col = spacer_col + 1
        
    # ── SHEET 2: JUN'2026 (Zonal Summary) ──────────────────────────────
    ws2 = wb.create_sheet("JUN'2026")
    ws2.views.sheetView[0].showGridLines = True
    ws2.sheet_view.zoomScale = 90
    
    ws2.row_dimensions[1].height = 10
    ws2.row_dimensions[2].height = 28
    ws2.row_dimensions[3].height = 28
    ws2.row_dimensions[4].height = 8
    ws2.row_dimensions[5].height = 22
    ws2.row_dimensions[11].height = 26
    for r in range(18, 49):
        ws2.row_dimensions[r].height = 20
        
    ws2.freeze_panes = 'D18'
    
    # Colors/Tab Colors
    if zone == 'CTG.A':
        ws1.sheet_properties.tabColor = '00F2FE'
        ws2.sheet_properties.tabColor = '00F2FE'
    else:
        ws1.sheet_properties.tabColor = 'A855F7'
        ws2.sheet_properties.tabColor = 'A855F7'
        
    # Lay out columns:
    # A: Spacer
    # B: DATE
    # C: SH, SELF
    # D..: FM wise total summaries
    # Then all FM columns side-by-side.
    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 14  # DATE
    ws2.column_dimensions['C'].width = 13  # SH SELF
    
    # Group metadata elements
    num_fms = len(zone_fms)
    
    # Summary header
    summary_start_col = 4
    summary_end_col = 3 + num_fms
    
    ws2.merge_cells(start_row=5, start_column=summary_start_col, end_row=5, end_column=summary_end_col)
    c_sum_hdr = ws2.cell(row=5, column=summary_start_col, value=f"FM WISE TOTAL CASH IN HAND, {zone}")
    c_sum_hdr.font = font_total_h
    c_sum_hdr.fill = fill_tot_head
    c_sum_hdr.alignment = align_center
    
    # Format summary columns headers
    for col_idx in range(summary_start_col, summary_end_col + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18
        ws2.cell(row=5, column=col_idx).fill = fill_tot_head
        ws2.cell(row=5, column=col_idx).border = bd_total
        
    # SH, SELF header
    ws2.merge_cells(start_row=5, start_column=3, end_row=11, end_column=3)
    c_sh_hdr = ws2.cell(row=5, column=3, value="SH, SELF")
    c_sh_hdr.font = font_total_h
    c_sh_hdr.fill = fill_tot_head
    c_sh_hdr.alignment = align_center
    for r in range(5, 12):
        ws2.cell(row=r, column=3).fill = fill_tot_head
        ws2.cell(row=r, column=3).border = bd_total
        
    # Populate FM names in Row 6 (merged down to 11) for summary columns
    for f_idx, (fm_name, fm_data) in enumerate(zone_fms.items()):
        col_idx = summary_start_col + f_idx
        fm_clean_name = fm_name.split(',')[0].strip()
        
        ws2.merge_cells(start_row=6, start_column=col_idx, end_row=11, end_column=col_idx)
        c_fm_summary_hdr = ws2.cell(row=6, column=col_idx, value=fm_clean_name)
        c_fm_summary_hdr.font = font_total_h
        c_fm_summary_hdr.fill = fill_tot_head
        c_fm_summary_hdr.alignment = align_center
        for r in range(6, 12):
            ws2.cell(row=r, column=col_idx).fill = fill_tot_head
            ws2.cell(row=r, column=col_idx).border = bd_total
            
    # Now build the individual FM columns side-by-side starting from summary_end_col + 1
    start_data_col = summary_end_col + 1
    
    # Store mappings of which summary column sums which individual columns
    fm_col_ranges = []
    
    curr_col = start_data_col
    for fm_name, fm_data in zone_fms.items():
        fm_clean_name = fm_name.split(',')[0].strip()
        markets = fm_data['markets']
        
        # Get unique DAs
        das = []
        for m in markets:
            if m['da_name'] and str(m['da_name']).strip().upper() != 'VACANT':
                da_str = str(m['da_name']).strip().upper()
                if da_str not in das:
                    das.append(da_str)
                    
        num_mpos = len(markets)
        num_das = len(das)
        
        fm_start = curr_col
        fm_end = curr_col + 1 + num_mpos + num_das - 1
        fm_col_ranges.append((fm_start, fm_end))
        
        # 1. FM SELF Column (Col curr_col)
        ws2.merge_cells(start_row=11, start_column=curr_col, end_row=14, end_column=curr_col)
        c_fm_self_h = ws2.cell(row=11, column=curr_col, value="FM SELF")
        c_fm_self_h.font = font_hdr
        c_fm_self_h.fill = fill_mid
        c_fm_self_h.alignment = align_center
        for r in range(11, 15):
            ws2.cell(row=r, column=curr_col).fill = fill_mid
            ws2.cell(row=r, column=curr_col).border = bd_hdr
            
        ws2.cell(row=6, column=curr_col, value=zone).font = font_sub
        ws2.cell(row=7, column=curr_col, value=fm_clean_name).font = font_sub
        for r in range(6, 11):
            ws2.cell(row=r, column=curr_col).fill = fill_navy
            ws2.cell(row=r, column=curr_col).border = bd_hdr
            ws2.cell(row=r, column=curr_col).alignment = align_center
            
        ws2.column_dimensions[get_column_letter(curr_col)].width = 13
        
        # 2. MPO Columns (Cols curr_col + 1 to curr_col + num_mpos)
        mpo_start = curr_col + 1
        mpo_end = curr_col + num_mpos
        
        ws2.merge_cells(start_row=5, start_column=mpo_start, end_row=5, end_column=mpo_end)
        c_mpo_h = ws2.cell(row=5, column=mpo_start, value="MPO")
        c_mpo_h.font = font_hdr
        c_mpo_h.fill = fill_mid
        c_mpo_h.alignment = align_center
        
        for c in range(mpo_start, mpo_end + 1):
            ws2.cell(row=5, column=c).fill = fill_mid
            ws2.cell(row=5, column=c).border = bd_hdr
            
        for m_idx, m in enumerate(markets):
            c_idx = mpo_start + m_idx
            ws2.cell(row=6, column=c_idx, value=zone).font = font_sub
            ws2.cell(row=7, column=c_idx, value=fm_clean_name).font = font_sub
            ws2.cell(row=8, column=c_idx, value=m['market_name']).font = font_sub
            ws2.cell(row=9, column=c_idx, value=m['mpo_code']).font = font_sub
            ws2.cell(row=10, column=c_idx, value=m['fm_code']).font = font_sub
            
            for r in range(6, 11):
                ws2.cell(row=r, column=c_idx).fill = fill_navy
                ws2.cell(row=r, column=c_idx).border = bd_hdr
                ws2.cell(row=r, column=c_idx).alignment = align_center
                
            ws2.cell(row=11, column=c_idx, value=m['mpo_name']).font = font_name
            ws2.cell(row=11, column=c_idx).fill = fill_dark
            ws2.cell(row=11, column=c_idx).border = bd_hdr
            ws2.cell(row=11, column=c_idx).alignment = align_center
            
            ws2.cell(row=12, column=c_idx, value=m['desig']).font = font_sub
            ws2.cell(row=13, column=c_idx, value=m['is_vacant']).font = font_sub
            ws2.cell(row=14, column=c_idx, value=m['da_name']).font = font_sub
            for r in range(12, 15):
                ws2.cell(row=r, column=c_idx).fill = fill_navy
                ws2.cell(row=r, column=c_idx).border = bd_hdr
                ws2.cell(row=r, column=c_idx).alignment = align_center
                
            name_len = len(m['mpo_name'] or '')
            ws2.column_dimensions[get_column_letter(c_idx)].width = max(name_len * 1.1, 13)
            
            if m['is_vacant'] == 'Y':
                ws2.column_dimensions[get_column_letter(c_idx)].hidden = True
                
        # 3. DA Columns (Cols curr_col + 1 + num_mpos to curr_col + num_mpos + num_das)
        if num_das > 0:
            da_start = curr_col + 1 + num_mpos
            da_end = curr_col + num_mpos + num_das
            
            ws2.merge_cells(start_row=5, start_column=da_start, end_row=5, end_column=da_end)
            c_da_h = ws2.cell(row=5, column=da_start, value="DA")
            c_da_h.font = font_hdr
            c_da_h.fill = fill_mid
            c_da_h.alignment = align_center
            
            for c in range(da_start, da_end + 1):
                ws2.cell(row=5, column=c).fill = fill_mid
                ws2.cell(row=5, column=c).border = bd_hdr
                
            for d_idx, da_name in enumerate(das):
                c_idx = da_start + d_idx
                ws2.cell(row=11, column=c_idx, value=da_name).font = font_name
                ws2.cell(row=11, column=c_idx).fill = fill_dark
                ws2.cell(row=11, column=c_idx).border = bd_hdr
                ws2.cell(row=11, column=c_idx).alignment = align_center
                
                # Apply headers styling and borders
                for r in range(5, 11):
                    # For DA columns, rows 6-10 are styled blank headers
                    if r >= 6:
                        ws2.cell(row=r, column=c_idx).fill = fill_navy
                        ws2.cell(row=r, column=c_idx).border = bd_hdr
                for r in range(12, 15):
                    ws2.cell(row=r, column=c_idx).fill = fill_navy
                    ws2.cell(row=r, column=c_idx).border = bd_hdr
                    
                name_len = len(da_name or '')
                ws2.column_dimensions[get_column_letter(c_idx)].width = max(name_len * 1.1, 13)
                
        curr_col = fm_end + 1
        
    # DATE Header for Zonal Summary tab
    ws2.merge_cells(start_row=11, start_column=2, end_row=14, end_column=2)
    c_date_h = ws2.cell(row=11, column=2, value="DATE")
    c_date_h.font = font_hdr
    c_date_h.fill = fill_mid
    c_date_h.alignment = align_center
    for r in range(11, 15):
        ws2.cell(row=r, column=2).fill = fill_mid
        ws2.cell(row=r, column=2).border = bd_hdr
        
    # Merged title block spanning from B2 to the last data column
    last_col_idx = curr_col - 1
    ws2.merge_cells(start_row=2, start_column=2, end_row=3, end_column=last_col_idx)
    title_cell = ws2.cell(row=2, column=2, value="CASH IN HAND")
    title_cell.font = font_title
    title_cell.fill = fill_void
    title_cell.alignment = align_center
    for r in range(2, 4):
        for c in range(2, last_col_idx + 1):
            cell = ws2.cell(row=r, column=c)
            cell.fill = fill_void
            left = side('medium', B_NEON) if c == 2 else None
            right = side('medium', B_NEON) if c == last_col_idx else None
            top = side('medium', B_NEON) if r == 2 else None
            bottom = side('medium', B_NEON) if r == 3 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            
    # Rows 15 to 17 buffer and formulas
    for r in range(15, 18):
        # Date column is empty in these rows
        for c in range(2, last_col_idx + 1):
            ws2.cell(row=r, column=c).border = bd_data
            
        # SH SELF
        ws2.cell(row=r, column=3, value=0).font = font_total_d
        ws2.cell(row=r, column=3).alignment = align_center
        ws2.cell(row=r, column=3).fill = fill_tot_data
        ws2.cell(row=r, column=3).border = bd_total
        
        # Summaries
        for f_idx, (f_start, f_end) in enumerate(fm_col_ranges):
            col_idx = summary_start_col + f_idx
            ws2.cell(row=r, column=col_idx, value=f"=SUM({get_column_letter(f_start)}{r}:{get_column_letter(f_end)}{r})").font = font_total_d
            ws2.cell(row=r, column=col_idx).alignment = align_center
            ws2.cell(row=r, column=col_idx).fill = fill_tot_data
            ws2.cell(row=r, column=col_idx).border = bd_total
            
    # Rows 18 to 48 data and formulas
    for r_idx, date_val in enumerate(dates):
        row_num = 18 + r_idx
        row_fill = fill_za if (r_idx % 2 == 0) else fill_zb
        
        c_val = ws2.cell(row=row_num, column=2, value=date_val)
        c_val.font = font_date
        c_val.alignment = align_center
        c_val.border = bd_date
        c_val.fill = row_fill
        
        # SH, SELF
        sh_cell = ws2.cell(row=row_num, column=3, value=0)
        sh_cell.font = font_total_d
        sh_cell.alignment = align_right
        sh_cell.border = bd_total
        sh_cell.fill = fill_tot_data
        sh_cell.number_format = '#,##0'
        
        # Zonal FM total summary columns
        for f_idx, (f_start, f_end) in enumerate(fm_col_ranges):
            col_idx = summary_start_col + f_idx
            sum_cell = ws2.cell(row=row_num, column=col_idx, value=f"=SUM({get_column_letter(f_start)}{row_num}:{get_column_letter(f_end)}{row_num})")
            sum_cell.font = font_total_d
            sum_cell.alignment = align_right
            sum_cell.border = bd_total
            sum_cell.fill = fill_tot_data
            sum_cell.number_format = '#,##0'
            
        # Individual columns
        for c in range(start_data_col, last_col_idx + 1):
            cell = ws2.cell(row=row_num, column=c)
            cell.border = bd_data
            cell.alignment = align_right
            cell.number_format = '#,##0'
            cell.font = font_body
            cell.fill = row_fill
            
    # Hide metadata rows in summary sheet
    for r in [6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17]:
        ws2.row_dimensions[r].hidden = True
        
    wb.save(output_path)
    wb.close()
    print(f"Generated Zonal Summary Excel: {output_path}")

def main():
    wb_local = openpyxl.load_workbook(r"c:\Users\Irak\Desktop\Cash in Hand and Dic Adjustment\Cash in Hand.xlsx")
    ws_local = wb_local["FM wise DA and MPO Names"]
    
    fm_groups = {}
    for r in range(2, ws_local.max_row + 1):
        depot = ws_local.cell(row=r, column=1).value
        zone = ws_local.cell(row=r, column=2).value
        market = ws_local.cell(row=r, column=3).value
        mpo_name = ws_local.cell(row=r, column=4).value
        fm_am_zone = ws_local.cell(row=r, column=5).value
        vacant = ws_local.cell(row=r, column=6).value
        desig = ws_local.cell(row=r, column=7).value
        mpo_code = ws_local.cell(row=r, column=8).value
        fm_code = ws_local.cell(row=r, column=9).value
        da_names = [ws_local.cell(row=r, column=c).value for c in range(10, 14)]
        da_names = [da for da in da_names if da and str(da).strip() and str(da).strip().upper() != 'VACANT']
        
        if not fm_am_zone or str(fm_am_zone).strip() == "":
            continue
            
        fm_am_zone = str(fm_am_zone).strip()
        if ',' in fm_am_zone:
            parts = fm_am_zone.split(',')
            cleaned_fm_part = clean_person_name(parts[0])
            fm_am_zone_clean = f"{cleaned_fm_part}, {parts[1].strip()}"
        else:
            fm_am_zone_clean = clean_person_name(fm_am_zone)
            
        if fm_am_zone_clean not in fm_groups:
            fm_groups[fm_am_zone_clean] = {
                'depot': depot,
                'zone': str(zone).strip() if zone else "",
                'markets': []
            }
            
        is_vacant = (vacant == 'Y') or (mpo_name and 'vacant' in str(mpo_name).lower())
        mpo_clean = 'VACANT' if is_vacant else clean_person_name(mpo_name)
        da_clean = clean_person_name(da_names[0]) if da_names else None
        
        fm_groups[fm_am_zone_clean]['markets'].append({
            'market_name': market,
            'mpo_name': mpo_clean,
            'desig': desig,
            'mpo_code': mpo_code,
            'fm_code': fm_code,
            'is_vacant': 'Y' if is_vacant else None,
            'da_name': da_clean
        })
    wb_local.close()
    
    # Filter valid FMs
    valid_fms = {}
    for fm_name, fm_data in fm_groups.items():
        non_vacant_mpos = [m for m in fm_data['markets'] if not m['is_vacant']]
        if non_vacant_mpos:
            valid_fms[fm_name] = fm_data
            
    # Group valid FMs by Zone
    zone_groups = {}
    for fm_name, fm_data in valid_fms.items():
        zone = fm_data['zone']
        if zone not in zone_groups:
            zone_groups[zone] = {}
        zone_groups[zone][fm_name] = fm_data
        
    base_dir = r"c:\Users\Irak\Desktop\Cash in Hand and Dic Adjustment"
    for zone, zone_fms in zone_groups.items():
        # Sort zone_fms according to our custom key
        def fm_sort_key(item):
            fm_name = item[0]
            fm_clean = fm_name.split(',')[0].strip().upper()
            if zone == 'CTG.B':
                order = ["JAINAL ABEDIN AKHAND", "FIROZ AHMED", "NARAYAN DAS", "SHAHJAHAN"]
            else:
                order = ["MONIR UDDIN", "RAFIQUL MOULA", "VACANT, KHAGRACHARI"]
            for idx, name in enumerate(order):
                if name in fm_clean:
                    return idx
            return 99
            
        sorted_zone_fms = dict(sorted(zone_fms.items(), key=fm_sort_key))
        
        zone_dir = os.path.join(base_dir, zone)
        os.makedirs(zone_dir, exist_ok=True)
        
        output_file = os.path.join(zone_dir, f"{zone} CASH IN HAND.xlsx")
        generate_zonal_file(zone, sorted_zone_fms, output_file)

if __name__ == "__main__":
    main()
