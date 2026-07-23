"""
Script to rearrange Bhaluka from JPUR zone to MYM.A zone
"""
import os, json, shutil
import openpyxl

BASE_DIR = r"C:\Users\Irak\Desktop\deskTop\Cash in Hand and Dic Adjustment"
RAW_DATA_PATH = os.path.join(BASE_DIR, "raw_compiled_data.json")

# Step 1: Update raw_compiled_data.json
print("=== Step 1: Updating raw_compiled_data.json ===")
with open(RAW_DATA_PATH, 'r', encoding='utf-8') as f:
    raw_data = json.load(f)

updated_count = 0
for rec in raw_data.get('records', []):
    m_name = str(rec.get('market_name', '')).strip()
    if m_name in ['SEED STORE', 'BHALUKA', 'GAFFORGAON-1', 'GAFFORGAON-2']:
        rec['zone'] = 'MYM.A'
        updated_count += 1

with open(RAW_DATA_PATH, 'w', encoding='utf-8') as f:
    json.dump(raw_data, f, indent=2, ensure_ascii=False)

print(f"Updated {updated_count} market records from JPUR to MYM.A in raw_compiled_data.json.")

# Step 2: Move local Excel file from JPUR to MYM.A
print("\n=== Step 2: Moving local Excel file ===")
jpur_bhaluka_file = os.path.join(BASE_DIR, "JPUR", "VACANT, BHALUKA.xlsx")
mym_a_dir = os.path.join(BASE_DIR, "MYM.A")
os.makedirs(mym_a_dir, exist_ok=True)
mym_a_bhaluka_file = os.path.join(mym_a_dir, "VACANT, BHALUKA.xlsx")

if os.path.exists(jpur_bhaluka_file):
    # Load and update zone label inside the workbook
    wb = openpyxl.load_workbook(jpur_bhaluka_file)
    ws = wb.active
    # Update zone references in row 6 and row 13
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=6, column=col).value == 'JPUR':
            ws.cell(row=6, column=col, value='MYM.A')
        if ws.cell(row=13, column=col).value == 'JPUR':
            ws.cell(row=13, column=col, value='MYM.A')
    
    wb.save(mym_a_bhaluka_file)
    wb.close()
    os.remove(jpur_bhaluka_file)
    print(f"Moved {jpur_bhaluka_file} -> {mym_a_bhaluka_file} and updated zone label to MYM.A.")

print("\n=== Step 3: Local file rearrangement complete ===")
