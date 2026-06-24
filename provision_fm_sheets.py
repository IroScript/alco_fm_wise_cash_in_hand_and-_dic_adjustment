import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import re
import time
import json
import gspread
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# Directory Settings
BASE_DIR = r"c:\Users\Irak\Desktop\Cash in Hand and Dic Adjustment"
CLIENT_SECRET_FILE = os.path.join(BASE_DIR, "FieldEdit", "client_secret_866102064521-5g6tq5989nqs97ehgse7n6fl1o9pslt5.apps.googleusercontent.com.json")
TOKEN_FILE = os.path.join(BASE_DIR, "FieldEdit", "token.json")
PARENT_FOLDER_ID = "1iOFeqywnIZ_yVclg_Em2U1npPtsokfGk"
EMAIL_SHEET_ID = "1f5SFvhH8Bjb3OUlpof68teBktHuYyVELioxLv_KWXJo"

# Google APIs Setup
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

def get_oauth_credentials():
    creds = None
    if os.path.exists(TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception:
            pass
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())
    return creds

creds = get_oauth_credentials()
gc = gspread.authorize(creds)
drive_service = build('drive', 'v3', credentials=creds)

def clean_person_name(name):
    if not name:
        return ""
    name = str(name).strip()
    name = re.sub(r'^(MR|MD|MRS|MST|DR)\.?\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(MR|MD|MRS|MST|DR)\.?\s+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', ' ', name)
    return name.strip().upper()

def get_email_mappings():
    print("Fetching email mappings from Google Sheet...")
    sheet = gc.open_by_key(EMAIL_SHEET_ID)
    ws = sheet.worksheet("EMAIL_2026")
    rows = ws.get_all_values()
    
    headers = [h.strip().upper() for h in rows[0]]
    fm_col = headers.index('FM/AM, ZONE')
    email_col = headers.index('EMAIL')
    boss_email_col = headers.index('SM/DSM MAIL') if 'SM/DSM MAIL' in headers else -1
    boss_name_col = headers.index('SM NAME') if 'SM NAME' in headers else -1

    mappings = {}
    for r in rows[1:]:
        if len(r) <= fm_col or not r[fm_col]:
            continue
        fm_full = str(r[fm_col]).strip()
        # Extract FM Name without zone suffix (e.g. "RASHIDUL ISLAM, TANG" -> "RASHIDUL ISLAM")
        fm_name = fm_full.split(',')[0].strip().upper()
        fm_name_clean = clean_person_name(fm_name)
        
        email = str(r[email_col]).strip() if len(r) > email_col else ""
        boss_email = str(r[boss_email_col]).strip() if boss_email_col != -1 and len(r) > boss_email_col else ""
        boss_name = str(r[boss_name_col]).strip() if boss_name_col != -1 and len(r) > boss_name_col else ""
        
        mappings[fm_name_clean] = {
            'email': email,
            'boss_email': boss_email,
            'boss_name': boss_name
        }
    return mappings

def get_or_create_drive_folder(name, parent_id):
    query = f"name = '{name}' and mimeType = 'application/vnd.google-apps.folder' and '{parent_id}' in parents and trashed = false"
    response = drive_service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    files = response.get('files', [])
    if files:
        return files[0]['id']
    
    # Create folder
    file_metadata = {
        'name': name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }
    folder = drive_service.files().create(body=file_metadata, fields='id').execute()
    print(f"Created Google Drive folder: {name} (ID: {folder['id']})")
    return folder['id']

def share_file(file_id, email, role='writer'):
    if not email or '@' not in email:
        return
    try:
        user_permission = {
            'type': 'user',
            'role': role,
            'emailAddress': email
        }
        drive_service.permissions().create(
            fileId=file_id,
            body=user_permission,
            fields='id'
        ).execute()
        print(f"Shared file {file_id} with {email} as {role}")
    except Exception as e:
        print(f"Error sharing file {file_id} with {email}: {e}")

def create_local_excel(fm_name, fm_data):
    # ── 50K AD Design System ──────────────────────────────
    FONT_FAMILY   = 'Aptos'
    C_VOID        = '060816'
    C_DEEP_NAVY   = '0D1425'
    C_MIDNIGHT    = '1E293B'
    C_DARK_SURF   = '0F172A'
    C_ZEBRA_A     = 'FFFFFF'
    C_ZEBRA_B     = 'F8FAFC'
    C_TOTAL_DATA  = 'ECFDF5'
    C_TOTAL_HEAD  = '065F46'
    T_NEON        = '00F2FE'
    T_WHITE       = 'FFFFFF'
    T_SLATE       = 'CBD5E1'
    T_INK         = '0F172A'
    T_MINT        = 'A7F3D0'
    T_TOTAL_DARK  = '064E3B'
    T_DATE        = '475569'
    B_NEON        = '0E7490'
    B_SLATE       = '334155'
    B_LIGHT       = 'E2E8F0'
    B_TOTAL       = '6EE7B7'

    font_title    = Font(name=FONT_FAMILY, size=18, bold=True,  color=T_NEON)
    font_hdr      = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
    font_sub      = Font(name=FONT_FAMILY, size=9,  bold=True,  color=T_SLATE)
    font_name     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
    font_date     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_DATE)
    font_body     = Font(name=FONT_FAMILY, size=10, bold=False, color=T_INK)
    font_total_h  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_MINT)
    font_total_d  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_TOTAL_DARK)

    fill_void     = PatternFill('solid', start_color=C_VOID,       end_color=C_VOID)
    fill_navy     = PatternFill('solid', start_color=C_DEEP_NAVY,  end_color=C_DEEP_NAVY)
    fill_mid      = PatternFill('solid', start_color=C_MIDNIGHT,   end_color=C_MIDNIGHT)
    fill_dark     = PatternFill('solid', start_color=C_DARK_SURF,  end_color=C_DARK_SURF)
    fill_za       = PatternFill('solid', start_color=C_ZEBRA_A,    end_color=C_ZEBRA_A)
    fill_zb       = PatternFill('solid', start_color=C_ZEBRA_B,    end_color=C_ZEBRA_B)
    fill_tot_data = PatternFill('solid', start_color=C_TOTAL_DATA, end_color=C_TOTAL_DATA)
    fill_tot_head = PatternFill('solid', start_color=C_TOTAL_HEAD, end_color=C_TOTAL_HEAD)

    def side(style='thin', color=B_LIGHT):
        return Side(style=style, color=color)

    bd_data  = Border(left=side(), right=side(), top=side(), bottom=side())
    bd_hdr   = Border(left=side('thin',B_SLATE), right=side('thin',B_SLATE), top=side('thin',B_SLATE), bottom=side('thin',B_SLATE))
    bd_total = Border(left=side('thin',B_TOTAL), right=side('thin',B_TOTAL), top=side('thin',B_TOTAL), bottom=side('thin',B_TOTAL))
    bd_date  = Border(left=side('medium',B_SLATE), right=side('thin',B_LIGHT), top=side('thin',B_LIGHT), bottom=side('thin',B_LIGHT))

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right  = Alignment(horizontal='right', vertical='center')

    # Dates for JUN'26 (30 days)
    dates = [f"{d} JUN'26" for d in range(30, 0, -1)]

    zone = fm_data['zone']
    markets = fm_data['markets']
    das = []
    for m in markets:
        if m['da_name'] and str(m['da_name']).strip().upper() != 'VACANT':
            da_str = str(m['da_name']).strip().upper()
            if da_str not in das:
                das.append(da_str)
                
    num_mpos = len(markets)
    num_das = len(das)
    total_cols = 1 + 1 + num_mpos + num_das + 1
    total_col = 4 + num_mpos + num_das
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JUN'26"
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_view.zoomScale = 90
    
    # Style Tab Color
    if zone == 'CTG.A':
        ws.sheet_properties.tabColor = '00F2FE'
    else:
        ws.sheet_properties.tabColor = 'A855F7'
        
    # Row Heights
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[11].height = 26
    for r in range(18, 18 + len(dates)):
        ws.row_dimensions[r].height = 20
        
    ws.freeze_panes = 'D18'
    
    # Title Block
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=total_cols+1)
    title_cell = ws.cell(row=2, column=2, value="CASH IN HAND")
    title_cell.font = font_title
    title_cell.fill = fill_void
    title_cell.alignment = align_center
    
    for r in range(2, 4):
        for c in range(2, total_cols + 2):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_void
            left = side('medium', B_NEON) if c == 2 else None
            right = side('medium', B_NEON) if c == total_cols + 1 else None
            top = side('medium', B_NEON) if r == 2 else None
            bottom = side('medium', B_NEON) if r == 3 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            
    mpo_start = 4
    mpo_end = 3 + num_mpos
    da_start = 4 + num_mpos
    da_end = 3 + num_mpos + num_das
    
    ws.merge_cells(start_row=5, start_column=mpo_start, end_row=5, end_column=mpo_end)
    c_mpo = ws.cell(row=5, column=mpo_start, value="MPO")
    c_mpo.font = font_hdr
    c_mpo.fill = fill_mid
    c_mpo.alignment = align_center
    
    if num_das > 0:
        ws.merge_cells(start_row=5, start_column=da_start, end_row=5, end_column=da_end)
        c_da = ws.cell(row=5, column=da_start, value="DA")
        c_da.font = font_hdr
        c_da.fill = fill_mid
        c_da.alignment = align_center
        
    ws.merge_cells(start_row=5, start_column=total_col, end_row=11, end_column=total_col)
    c_tot_header = ws.cell(row=5, column=total_col, value="TOTAL CASH IN HAND")
    c_tot_header.font = font_total_h
    c_tot_header.fill = fill_tot_head
    c_tot_header.alignment = align_center
    
    for r in range(5, 12):
        ws.cell(row=r, column=total_col).fill = fill_tot_head
        ws.cell(row=r, column=total_col).border = bd_total
    for c in range(mpo_start, mpo_end + 1):
        ws.cell(row=5, column=c).fill = fill_mid
        ws.cell(row=5, column=c).border = bd_hdr
    if num_das > 0:
        for c in range(da_start, da_end + 1):
            ws.cell(row=5, column=c).fill = fill_mid
            ws.cell(row=5, column=c).border = bd_hdr
            
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        ws.cell(row=6, column=c_idx, value=zone).font = font_sub
        ws.cell(row=7, column=c_idx, value=fm_name).font = font_sub
        ws.cell(row=8, column=c_idx, value=m['market_name']).font = font_sub
        ws.cell(row=9, column=c_idx, value=m['mpo_code']).font = font_sub
        ws.cell(row=10, column=c_idx, value=m['fm_code']).font = font_sub
        
        for r in range(6, 11):
            ws.cell(row=r, column=c_idx).fill = fill_navy
            ws.cell(row=r, column=c_idx).border = bd_hdr
            ws.cell(row=r, column=c_idx).alignment = align_center
            
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        ws.cell(row=11, column=c_idx, value=m['mpo_name']).font = font_name
        ws.cell(row=11, column=c_idx).fill = fill_dark
        ws.cell(row=11, column=c_idx).border = bd_hdr
        ws.cell(row=11, column=c_idx).alignment = align_center
        
    for d_idx, da_name in enumerate(das):
        c_idx = da_start + d_idx
        ws.cell(row=11, column=c_idx, value=da_name).font = font_name
        ws.cell(row=11, column=c_idx).fill = fill_dark
        ws.cell(row=11, column=c_idx).border = bd_hdr
        ws.cell(row=11, column=c_idx).alignment = align_center
        
    ws.merge_cells(start_row=11, start_column=2, end_row=14, end_column=2)
    c_date_h = ws.cell(row=11, column=2, value="DATE")
    c_date_h.font = font_hdr
    c_date_h.fill = fill_mid
    c_date_h.alignment = align_center
    
    ws.merge_cells(start_row=11, start_column=3, end_row=14, end_column=3)
    c_fm_h = ws.cell(row=11, column=3, value="FM SELF")
    c_fm_h.font = font_hdr
    c_fm_h.fill = fill_mid
    c_fm_h.alignment = align_center
    
    for r in range(11, 15):
        ws.cell(row=r, column=2).fill = fill_mid
        ws.cell(row=r, column=2).border = bd_hdr
        ws.cell(row=r, column=3).fill = fill_mid
        ws.cell(row=r, column=3).border = bd_hdr
        
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        ws.cell(row=12, column=c_idx, value=m['desig']).font = font_sub
        ws.cell(row=13, column=c_idx, value=m['is_vacant']).font = font_sub
        ws.cell(row=14, column=c_idx, value=m['da_name']).font = font_sub
        for r in range(12, 15):
            ws.cell(row=r, column=c_idx).fill = fill_navy
            ws.cell(row=r, column=c_idx).border = bd_hdr
            ws.cell(row=r, column=c_idx).alignment = align_center
            
    for r in range(15, 18):
        c_tot = ws.cell(row=r, column=total_col, value=f"=SUM(C{r}:{get_column_letter(total_col-1)}{r})")
        c_tot.font = font_total_d
        c_tot.alignment = align_center
        c_tot.fill = fill_tot_data
        c_tot.border = bd_total
        for c in range(2, total_col):
            ws.cell(row=r, column=c).border = bd_data
            
    for r_idx, date_val in enumerate(dates):
        row_num = 18 + r_idx
        row_fill = fill_za if (r_idx % 2 == 0) else fill_zb
        
        c_val = ws.cell(row=row_num, column=2, value=date_val)
        c_val.font = font_date
        c_val.alignment = align_center
        c_val.border = bd_date
        c_val.fill = row_fill
        
        for c in range(3, total_col):
            cell = ws.cell(row=row_num, column=c)
            cell.border = bd_data
            cell.alignment = align_right
            cell.number_format = '#,##0'
            cell.font = font_body
            cell.fill = row_fill
            
        c_tot = ws.cell(row=row_num, column=total_col, value=f"=SUM(C{row_num}:{get_column_letter(total_col-1)}{row_num})")
        c_tot.font = font_total_d
        c_tot.alignment = align_right
        c_tot.border = bd_total
        c_tot.fill = fill_tot_data
        c_tot.number_format = '#,##0'
        
    for m_idx, m in enumerate(markets):
        if m['is_vacant'] == 'Y':
            col_letter = get_column_letter(mpo_start + m_idx)
            ws.column_dimensions[col_letter].hidden = True
            
    for r in [6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17]:
        ws.row_dimensions[r].hidden = True
        
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 13
    
    for m_idx, m in enumerate(markets):
        c_let = get_column_letter(mpo_start + m_idx)
        name_len = len(m['mpo_name'] or '')
        ws.column_dimensions[c_let].width = max(name_len * 1.1, 13)
        
    for d_idx, da_name in enumerate(das):
        c_let = get_column_letter(da_start + d_idx)
        name_len = len(da_name or '')
        ws.column_dimensions[c_let].width = max(name_len * 1.1, 13)
        
    ws.column_dimensions[get_column_letter(total_col)].width = 18
    
    os.makedirs(os.path.join(BASE_DIR, "temp_sheets"), exist_ok=True)
    temp_path = os.path.join(BASE_DIR, "temp_sheets", f"{fm_name}.xlsx")
    wb.save(temp_path)
    wb.close()
    return temp_path

def main():
    start_time = time.time()
    wb_local = openpyxl.load_workbook(os.path.join(BASE_DIR, "Cash in Hand.xlsx"))
    ws_local = wb_local["FM wise DA and MPO Names"]
    
    fm_groups = {}
    for r in range(2, ws_local.max_row + 1):
        depot = ws_local.cell(row=r, column=1).value
        zone = ws_local.cell(row=r, column=2).value
        market = ws_local.cell(row=r, column=3).value
        mpo_name = ws_local.cell(row=r, column=4).value
        fm_am_zone = ws_local.cell(row=r, column=5).value
        vacant = ws_local.cell(row=r, column=6).value
        desig = ws_local.cell(row=r, column=7).value
        mpo_code = ws_local.cell(row=r, column=8).value
        fm_code = ws_local.cell(row=r, column=9).value
        da_names = [ws_local.cell(row=r, column=c).value for c in range(10, 14)]
        da_names = [da for da in da_names if da and str(da).strip() and str(da).strip().upper() != 'VACANT']

        if not fm_am_zone or str(fm_am_zone).strip() == "":
            continue

        fm_am_zone = str(fm_am_zone).strip()
        if ',' in fm_am_zone:
            parts = fm_am_zone.split(',')
            cleaned_fm_part = clean_person_name(parts[0])
            fm_am_zone_clean = f"{cleaned_fm_part}, {parts[1].strip()}"
        else:
            fm_am_zone_clean = clean_person_name(fm_am_zone)
        
        if fm_am_zone_clean not in fm_groups:
            fm_groups[fm_am_zone_clean] = {
                'depot': depot,
                'zone': str(zone).strip() if zone else "",
                'markets': []
            }
        
        is_vacant = (vacant == 'Y') or (mpo_name and 'vacant' in str(mpo_name).lower())
        mpo_clean = 'VACANT' if is_vacant else clean_person_name(mpo_name)
        da_clean = clean_person_name(da_names[0]) if da_names else None
        
        fm_groups[fm_am_zone_clean]['markets'].append({
            'market_name': market,
            'mpo_name': mpo_clean,
            'desig': desig,
            'mpo_code': mpo_code,
            'fm_code': fm_code,
            'is_vacant': 'Y' if is_vacant else None,
            'da_name': da_clean
        })

    wb_local.close()

    valid_fms = {}
    for fm_name, fm_data in fm_groups.items():
        non_vacant_mpos = [m for m in fm_data['markets'] if not m['is_vacant']]
        if non_vacant_mpos:
            valid_fms[fm_name] = fm_data

    print(f"Parsed {len(valid_fms)} valid FMs from Cash in Hand.xlsx")
    email_mappings = get_email_mappings()
    registry = []
    
    # We will process first 3 for safe dry run verification
    fms_to_process = list(valid_fms.keys())[:3]
    print(f"Processing first {len(fms_to_process)} FMs for safety dry-run...")
    
    for fm_name in fms_to_process:
        fm_clean_name = fm_name.split(',')[0].strip()
        fm_data = valid_fms[fm_name]
        zone = fm_data['zone']
        
        mapping = email_mappings.get(clean_person_name(fm_clean_name), {'email': '', 'boss_email': '', 'boss_name': ''})
        fm_email = mapping['email']
        boss_email = mapping['boss_email']
        boss_name = mapping['boss_name']
        
        print(f"\n--- Processing FM: {fm_clean_name} | Zone: {zone} ---")
        
        local_path = create_local_excel(fm_clean_name, fm_data)
        zone_folder_id = get_or_create_drive_folder(zone, PARENT_FOLDER_ID)
        
        file_metadata = {
            'name': f"CASH IN HAND - {fm_clean_name}",
            'mimeType': 'application/vnd.google-apps.spreadsheet',
            'parents': [zone_folder_id]
        }
        media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        
        uploaded_file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id, webViewLink'
        ).execute()
        
        sheet_id = uploaded_file.get('id')
        web_link = uploaded_file.get('webViewLink')
        print(f"Uploaded Google Sheet successfully. ID: {sheet_id}")
        
        if fm_email:
            share_file(sheet_id, fm_email, role='writer')
        if boss_email:
            share_file(sheet_id, boss_email, role='writer')
            
        registry.append({
            'FM Name': fm_clean_name,
            'Zone': zone,
            'Sheet ID': sheet_id,
            'URL': web_link,
            'FM Email': fm_email,
            'Boss Name': boss_name,
            'Boss Email': boss_email
        })
        
        try:
            if os.path.exists(local_path):
                os.remove(local_path)
        except Exception as e:
            print(f"Non-fatal clean up error: {e}")
            
        time.sleep(1)

    print("\nWriting Master Registry to Google Sheets...")
    registry_name = "Master_Registry_Cash_In_Hand"
    
    query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
    res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
    files = res.get('files', [])
    if files:
        reg_sheet = gc.open_by_key(files[0]['id'])
        reg_ws = reg_sheet.get_worksheet(0)
        reg_ws.clear()
        print("Cleared existing Master Registry sheet.")
    else:
        file_metadata = {
            'name': registry_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet',
            'parents': [PARENT_FOLDER_ID]
        }
        reg_file = drive_service.files().create(body=file_metadata, fields='id').execute()
        reg_sheet = gc.open_by_key(reg_file['id'])
        reg_ws = reg_sheet.get_worksheet(0)
        print("Created new Master Registry sheet.")
        
    header = ['FM Name', 'Zone', 'Sheet ID', 'URL', 'FM Email', 'Boss Name', 'Boss Email']
    rows_to_write = [header]
    for r in registry:
        rows_to_write.append([r['FM Name'], r['Zone'], r['Sheet ID'], r['URL'], r['FM Email'], r['Boss Name'], r['Boss Email']])
        
    reg_ws.update('A1', rows_to_write)
    print("Master Registry updated successfully.")
    print(f"Finished in {time.time() - start_time:.2f} seconds.")

if __name__ == "__main__":
    main()
