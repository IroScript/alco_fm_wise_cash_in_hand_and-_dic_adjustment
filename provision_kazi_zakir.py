"""
Kazi Zakir Hossain (SLT Zone) - Provisioning & Verification Script
"""
import os, json, time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = r"C:\Users\Irak\Desktop\deskTop\Cash in Hand and Dic Adjustment"

FM_NAME    = "KAZI ZAKIR HOSSAIN"
ZONE       = "SLT"
MONTH_STR  = "JUL'26"
NUM_DAYS   = 31

MARKETS = [
    {"market_name": "ZOKIGONJ-1",  "mpo_name": "UTTOM BISWAS", "desig": "MPO", "mpo_code": "5010", "fm_code": "5110", "is_vacant": "N"},
    {"market_name": "ZOKIGONJ-2",  "mpo_name": "UTTOM BISWAS", "desig": "MPO", "mpo_code": "5020", "fm_code": "5030", "is_vacant": "N"},
    {"market_name": "BEANIBAZAR",  "mpo_name": "UTTOM BISWAS", "desig": "MPO", "mpo_code": "5040", "fm_code": "5041", "is_vacant": "N"},
    {"market_name": "KULAURA",     "mpo_name": "AMARJIT DAS",  "desig": "AFM", "mpo_code": "4001", "fm_code": "4002", "is_vacant": "N"},
    {"market_name": "BARALAKHA",   "mpo_name": "PAPPU DAS",    "desig": "MPO", "mpo_code": "4003", "fm_code": "4007", "is_vacant": "N"},
]

DAS = [
    {"da_name": "MAIYNUL ISLAM"},
    {"da_name": "RAKIB HOSSAIN"}
]

# EXACT COLOR PALETTE
FONT_FAM = 'Aptos'
C_VOID   = '060816';  C_NAVY = '0D1425';  C_MID = '1E293B'
C_ZA     = 'FFFFFF';  C_ZB   = 'F8FAFC';  C_TOT_D = 'ECFDF5';  C_TOT_H = '065F46'
T_NEON   = '00F2FE';  T_WHT  = 'FFFFFF';  T_SLATE = 'CBD5E1';  T_INK   = '0F172A'
T_MINT   = 'A7F3D0';  T_TOTD = '064E3B';  T_DATE  = '475569'
B_NEON   = '0E7490';  B_SLT  = '334155';  B_LIGHT = 'E2E8F0';  B_TOTAL = '6EE7B7'

font_title   = Font(name=FONT_FAM, size=26, bold=True,  color=T_NEON)
font_hdr     = Font(name=FONT_FAM, size=10, bold=True,  color=T_WHT)
font_sub     = Font(name=FONT_FAM, size=9,  bold=True,  color=T_SLATE)
font_name    = Font(name=FONT_FAM, size=10, bold=True,  color=T_WHT)
font_date    = Font(name=FONT_FAM, size=10, bold=True,  color=T_DATE)
font_body    = Font(name=FONT_FAM, size=10, bold=False, color=T_INK)
font_tot_h   = Font(name=FONT_FAM, size=10, bold=True,  color=T_MINT)
font_tot_d   = Font(name=FONT_FAM, size=10, bold=True,  color=T_TOTD)

fill_void = PatternFill('solid', fgColor=C_VOID)
fill_navy = PatternFill('solid', fgColor=C_NAVY)
fill_mid  = PatternFill('solid', fgColor=C_MID)
fill_totH = PatternFill('solid', fgColor=C_TOT_H)
fill_totD = PatternFill('solid', fgColor=C_TOT_D)

def S(style='thin', color=B_LIGHT):
    return Side(style=style, color=color)
bd_data  = Border(left=S(), right=S(), top=S(), bottom=S())
bd_hdr   = Border(left=S('thin',B_SLT), right=S('thin',B_SLT), top=S('thin',B_SLT), bottom=S('thin',B_SLT))
bd_tot   = Border(left=S('thin',B_TOTAL), right=S('thin',B_TOTAL), top=S('thin',B_TOTAL), bottom=S('thin',B_TOTAL))
bd_date  = Border(left=S('medium',B_SLT), right=S('thin',B_LIGHT), top=S('thin',B_LIGHT), bottom=S('thin',B_LIGHT))
bd_title = Border(left=S('medium',B_NEON), right=S('medium',B_NEON), top=S('medium',B_NEON), bottom=S('medium',B_NEON))

align_c  = Alignment(horizontal='center', vertical='center', wrap_text=True)

def generate_kazi_zakir_excel():
    num_mpos = len(MARKETS)
    num_das  = len(DAS)
    total_col_idx = 4 + num_mpos + num_das
    total_col_ltr = get_column_letter(total_col_idx)
    last_inp_col  = total_col_idx - 1
    last_inp_ltr  = get_column_letter(last_inp_col)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = MONTH_STR
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = '334155'
    ws.freeze_panes = 'C18'

    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[11].height = 26
    for r in range(18, 18 + NUM_DAYS):
        ws.row_dimensions[r].height = 20

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 14
    for ci in range(4, total_col_idx):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.column_dimensions[total_col_ltr].width = 18

    # Title Block
    title_cell = ws.cell(row=2, column=3, value="CASH IN HAND")
    title_cell.font      = font_title
    title_cell.alignment = align_c
    title_cell.fill      = fill_void
    title_cell.border    = bd_title
    ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=total_col_idx)

    for r in range(1, 5):
        for c in range(1, total_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_void

    # Category Headers
    mpo_s = 4; mpo_e = 3 + num_mpos
    da_s  = mpo_e + 1; da_e = da_s + num_das - 1

    mpo_cell = ws.cell(row=5, column=mpo_s, value="MPO")
    mpo_cell.font = font_hdr; mpo_cell.alignment = align_c; mpo_cell.fill = fill_mid; mpo_cell.border = bd_hdr
    ws.merge_cells(start_row=5, start_column=mpo_s, end_row=5, end_column=mpo_e)

    da_cell = ws.cell(row=5, column=da_s, value="DA")
    da_cell.font = font_hdr; da_cell.alignment = align_c; da_cell.fill = fill_mid; da_cell.border = bd_hdr
    ws.merge_cells(start_row=5, start_column=da_s, end_row=5, end_column=da_e)

    tot_cell = ws.cell(row=5, column=total_col_idx, value="TOTAL CASH IN HAND")
    tot_cell.font = font_tot_h; tot_cell.alignment = align_c; tot_cell.fill = fill_totH; tot_cell.border = bd_tot
    ws.merge_cells(start_row=5, start_column=total_col_idx, end_row=11, end_column=total_col_idx)
    for r in range(6, 12):
        c = ws.cell(row=r, column=total_col_idx)
        c.fill = fill_totH; c.border = bd_tot

    for c in range(2, total_col_idx):
        cell = ws.cell(row=5, column=c)
        if cell.fill == PatternFill() or not cell.fill.fgColor or cell.fill.fgColor.rgb == '00000000':
            cell.fill = fill_mid
        cell.border = bd_hdr

    # Hidden Metadata Rows 6-10
    for r in range(6, 11):
        ws.row_dimensions[r].hidden = True
        for c in range(2, total_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_navy; cell.font = font_sub; cell.alignment = align_c; cell.border = bd_hdr

    for c in range(3, total_col_idx):
        ws.cell(row=6, column=c, value=ZONE)
        ws.cell(row=7, column=c, value=FM_NAME)

    for i, m in enumerate(MARKETS):
        ws.cell(row=8, column=mpo_s + i, value=m['market_name'])
        ws.cell(row=9, column=mpo_s + i, value=m['mpo_code'])
        ws.cell(row=10, column=mpo_s + i, value=m['fm_code'])

    # Row 11 Labels
    date_hdr = ws.cell(row=11, column=2, value="DATE")
    date_hdr.font = font_name; date_hdr.alignment = align_c; date_hdr.fill = fill_navy; date_hdr.border = bd_hdr
    ws.merge_cells(start_row=11, start_column=2, end_row=17, end_column=2)

    fm_self = ws.cell(row=11, column=3, value="FM SELF")
    fm_self.font = font_name; fm_self.alignment = align_c; fm_self.fill = fill_navy; fm_self.border = bd_hdr
    ws.cell(row=12, column=3).fill = fill_mid; ws.cell(row=12, column=3).border = bd_hdr

    for i, m in enumerate(MARKETS):
        c = ws.cell(row=11, column=mpo_s + i, value=m['mpo_name'])
        c.font = font_name; c.alignment = align_c; c.fill = fill_navy; c.border = bd_hdr

    for i, d in enumerate(DAS):
        c = ws.cell(row=11, column=da_s + i, value=d['da_name'])
        c.font = font_name; c.alignment = align_c; c.fill = fill_navy; c.border = bd_hdr

    ws.row_dimensions[12].hidden = True
    for c in range(2, total_col_idx + 1):
        cell = ws.cell(row=12, column=c)
        cell.fill = fill_mid; cell.border = bd_hdr

    for r in range(13, 18):
        ws.row_dimensions[r].hidden = True
        for c in range(2, total_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_navy; cell.font = font_sub; cell.alignment = align_c; cell.border = bd_hdr

    for c in range(3, total_col_idx):
        ws.cell(row=13, column=c, value=ZONE)
        ws.cell(row=14, column=c, value=FM_NAME)

    ws.cell(row=15, column=3, value="FM")
    for i, m in enumerate(MARKETS):
        ws.cell(row=15, column=mpo_s + i, value=m['desig'])
        ws.cell(row=16, column=mpo_s + i, value=m['mpo_code'])
    for i, d in enumerate(DAS):
        ws.cell(row=15, column=da_s + i, value="DA")
        ws.cell(row=17, column=da_s + i, value=d['da_name'])

    # Data Rows
    for idx in range(NUM_DAYS):
        r = 18 + idx
        d_val = f"{NUM_DAYS - idx} {MONTH_STR}"
        zr = PatternFill('solid', fgColor=C_ZA if idx % 2 == 0 else C_ZB)

        b = ws.cell(row=r, column=2, value=d_val)
        b.font = font_date; b.alignment = align_c; b.fill = zr; b.border = bd_date

        for c in range(3, total_col_idx):
            cell = ws.cell(row=r, column=c)
            cell.font = font_body; cell.alignment = align_c; cell.fill = zr
            cell.border = bd_data; cell.number_format = '#,##0'

        t = ws.cell(row=r, column=total_col_idx, value=f"=SUM(C{r}:{last_inp_ltr}{r})")
        t.font = font_tot_d; t.alignment = align_c; t.fill = fill_totD; t.border = bd_tot
        t.number_format = '#,##0'

    local_dir = os.path.join(BASE_DIR, ZONE)
    os.makedirs(local_dir, exist_ok=True)
    local_path = os.path.join(local_dir, f"{FM_NAME}.xlsx")
    wb.save(local_path)
    wb.close()
    print(f"Generated local Excel file successfully: {local_path}")
    return local_path

if __name__ == '__main__':
    generate_kazi_zakir_excel()
