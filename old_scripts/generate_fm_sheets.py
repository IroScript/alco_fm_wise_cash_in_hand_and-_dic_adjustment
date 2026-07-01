import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

# Load data from local Excel file
wb_local = openpyxl.load_workbook(r"c:\Users\Irak\Desktop\Cash in Hand and Dic Adjustment\Cash in Hand.xlsx")
ws_local = wb_local["FM wise DA and MPO Names"]

# Parse FM groups
fm_groups = {}

import re

def clean_person_name(name):
    if not name:
        return ""
    name = str(name).strip()
    # Remove common prefix titles like MR, MD, MRS, MST, DR (with optional dot)
    name = re.sub(r'^(MR|MD|MRS|MST|DR)\.?\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(MR|MD|MRS|MST|DR)\.?\s+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', ' ', name)
    return name.strip().upper()

for r in range(2, ws_local.max_row + 1):
    depot = ws_local.cell(row=r, column=1).value
    zone = ws_local.cell(row=r, column=2).value
    market = ws_local.cell(row=r, column=3).value
    mpo_name = ws_local.cell(row=r, column=4).value
    fm_am_zone = ws_local.cell(row=r, column=5).value
    vacant = ws_local.cell(row=r, column=6).value
    desig = ws_local.cell(row=r, column=7).value
    mpo_code = ws_local.cell(row=r, column=8).value
    fm_code = ws_local.cell(row=r, column=9).value
    da_names = [ws_local.cell(row=r, column=c).value for c in range(10, 14)]
    da_names = [da for da in da_names if da and str(da).strip() and str(da).strip().upper() != 'VACANT']

    if not fm_am_zone or str(fm_am_zone).strip() == "":
        continue

    # Clean FM name parts but keep zone suffix
    fm_am_zone = str(fm_am_zone).strip()
    if ',' in fm_am_zone:
        parts = fm_am_zone.split(',')
        cleaned_fm_part = clean_person_name(parts[0])
        fm_am_zone_clean = f"{cleaned_fm_part}, {parts[1].strip()}"
    else:
        fm_am_zone_clean = clean_person_name(fm_am_zone)
    
    if fm_am_zone_clean not in fm_groups:
        fm_groups[fm_am_zone_clean] = {
            'depot': depot,
            'zone': str(zone).strip() if zone else "",
            'markets': []
        }
    
    is_vacant = (vacant == 'Y') or (mpo_name and 'vacant' in str(mpo_name).lower())
    
    mpo_clean = 'VACANT' if is_vacant else clean_person_name(mpo_name)
    da_clean = clean_person_name(da_names[0]) if da_names else None
    
    fm_groups[fm_am_zone_clean]['markets'].append({
        'market_name': market,
        'mpo_name': mpo_clean,
        'desig': desig,
        'mpo_code': mpo_code,
        'fm_code': fm_code,
        'is_vacant': 'Y' if is_vacant else None,
        'da_name': da_clean
    })

# Filter valid FMs (with at least one non-vacant MPO)
valid_fms = {}
for fm_name, fm_data in fm_groups.items():
    non_vacant_mpos = [m for m in fm_data['markets'] if not m['is_vacant']]
    if non_vacant_mpos:
        valid_fms[fm_name] = fm_data

# Create output folders and files
base_dir = r"c:\Users\Irak\Desktop\Cash in Hand and Dic Adjustment"

# ── 50K AD Design System ──────────────────────────────
FONT_FAMILY   = 'Aptos'

# Backgrounds
C_VOID        = '060816'
C_DEEP_NAVY   = '0D1425'
C_MIDNIGHT    = '1E293B'
C_DARK_SURF   = '0F172A'
C_ZEBRA_A     = 'FFFFFF'
C_ZEBRA_B     = 'F8FAFC'
C_TOTAL_DATA  = 'ECFDF5'
C_TOTAL_HEAD  = '065F46'

# Text
T_NEON        = '00F2FE'
T_WHITE       = 'FFFFFF'
T_SLATE       = 'CBD5E1'
T_INK         = '0F172A'
T_MINT        = 'A7F3D0'
T_TOTAL_DARK  = '064E3B'
T_DATE        = '475569'

# Borders
B_NEON        = '0E7490'
B_SLATE       = '334155'
B_LIGHT       = 'E2E8F0'
B_TOTAL       = '6EE7B7'

# Fonts
font_title    = Font(name=FONT_FAMILY, size=18, bold=True,  color=T_NEON)
font_hdr      = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
font_sub      = Font(name=FONT_FAMILY, size=9,  bold=True,  color=T_SLATE)
font_name     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
font_date     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_DATE)
font_body     = Font(name=FONT_FAMILY, size=10, bold=False, color=T_INK)
font_total_h  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_MINT)
font_total_d  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_TOTAL_DARK)

# Fills
fill_void     = PatternFill('solid', start_color=C_VOID,       end_color=C_VOID)
fill_navy     = PatternFill('solid', start_color=C_DEEP_NAVY,  end_color=C_DEEP_NAVY)
fill_mid      = PatternFill('solid', start_color=C_MIDNIGHT,   end_color=C_MIDNIGHT)
fill_dark     = PatternFill('solid', start_color=C_DARK_SURF,  end_color=C_DARK_SURF)
fill_za       = PatternFill('solid', start_color=C_ZEBRA_A,    end_color=C_ZEBRA_A)
fill_zb       = PatternFill('solid', start_color=C_ZEBRA_B,    end_color=C_ZEBRA_B)
fill_tot_data = PatternFill('solid', start_color=C_TOTAL_DATA, end_color=C_TOTAL_DATA)
fill_tot_head = PatternFill('solid', start_color=C_TOTAL_HEAD, end_color=C_TOTAL_HEAD)

# Border helpers
def side(style='thin', color=B_LIGHT):
    return Side(style=style, color=color)

bd_data  = Border(left=side(), right=side(), top=side(), bottom=side())
bd_hdr   = Border(left=side('thin',B_SLATE), right=side('thin',B_SLATE),
                  top=side('thin',B_SLATE),  bottom=side('thin',B_SLATE))
bd_title = Border(left=side('medium',B_NEON), right=side('medium',B_NEON),
                  top=side('medium',B_NEON),  bottom=side('medium',B_NEON))
bd_total = Border(left=side('thin',B_TOTAL), right=side('thin',B_TOTAL),
                  top=side('thin',B_TOTAL),  bottom=side('thin',B_TOTAL))
bd_date  = Border(left=side('medium',B_SLATE), right=side('thin',B_LIGHT),
                  top=side('thin',B_LIGHT),   bottom=side('thin',B_LIGHT))
# ──────────────────────────────────────────────────────

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_right  = Alignment(horizontal='right', vertical='center')

dates = [f"{d} JUN'26" for d in range(31, 0, -1)]

for fm_name, fm_data in valid_fms.items():
    zone = fm_data['zone']
    markets = fm_data['markets']
    
    # Get unique DAs associated with this FM (non-vacant)
    das = []
    for m in markets:
        if m['da_name'] and str(m['da_name']).strip().upper() != 'VACANT':
            da_str = str(m['da_name']).strip().upper()
            if da_str not in das:
                das.append(da_str)
                
    num_mpos = len(markets)
    num_das = len(das)
    
    total_cols = 1 + 1 + num_mpos + num_das + 1
    total_col = 4 + num_mpos + num_das
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CASH IN HAND_format"
    ws.views.sheetView[0].showGridLines = True
    
    # Set zoom and tab color
    ws.sheet_view.zoomScale = 90
    if zone == 'CTG.A':
        ws.sheet_properties.tabColor = '00F2FE'  # Neon Cyan
    else:
        ws.sheet_properties.tabColor = 'A855F7'  # Neon Purple
        
    # Row Heights
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[11].height = 26
    for r in range(18, 49):
        ws.row_dimensions[r].height = 20
        
    # Lock Name + Date Headers (Freeze D18)
    ws.freeze_panes = 'D18'
    
    # Title Block (Merged B2:K3 equivalent)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=total_cols+1)
    title_cell = ws.cell(row=2, column=2, value="CASH IN HAND")
    title_cell.font = font_title
    title_cell.fill = fill_void
    title_cell.alignment = align_center
    
    # Apply border and fill to all title block cells
    for r in range(2, 4):
        for c in range(2, total_cols + 2):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_void
            left = side('medium', B_NEON) if c == 2 else None
            right = side('medium', B_NEON) if c == total_cols + 1 else None
            top = side('medium', B_NEON) if r == 2 else None
            bottom = side('medium', B_NEON) if r == 3 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            
    # Row 5: MPO & DA headers
    mpo_start = 4
    mpo_end = 3 + num_mpos
    da_start = 4 + num_mpos
    da_end = 3 + num_mpos + num_das
    
    ws.merge_cells(start_row=5, start_column=mpo_start, end_row=5, end_column=mpo_end)
    c_mpo = ws.cell(row=5, column=mpo_start, value="MPO")
    c_mpo.font = font_hdr
    c_mpo.fill = fill_mid
    c_mpo.alignment = align_center
    
    if num_das > 0:
        ws.merge_cells(start_row=5, start_column=da_start, end_row=5, end_column=da_end)
        c_da = ws.cell(row=5, column=da_start, value="DA")
        c_da.font = font_hdr
        c_da.fill = fill_mid
        c_da.alignment = align_center
        
    # TOTAL CASH IN HAND Header (K5:K11 merged)
    ws.merge_cells(start_row=5, start_column=total_col, end_row=11, end_column=total_col)
    c_tot_header = ws.cell(row=5, column=total_col, value="TOTAL CASH IN HAND")
    c_tot_header.font = font_total_h
    c_tot_header.fill = fill_tot_head
    c_tot_header.alignment = align_center
    
    # Fill remaining cells in headers
    for r in range(5, 12):
        ws.cell(row=r, column=total_col).fill = fill_tot_head
        ws.cell(row=r, column=total_col).border = bd_total
    for c in range(mpo_start, mpo_end + 1):
        ws.cell(row=5, column=c).fill = fill_mid
        ws.cell(row=5, column=c).border = bd_hdr
    if num_das > 0:
        for c in range(da_start, da_end + 1):
            ws.cell(row=5, column=c).fill = fill_mid
            ws.cell(row=5, column=c).border = bd_hdr
            
    # Populate metadata rows (Row 6 to 10)
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        
        ws.cell(row=6, column=c_idx, value=zone).font = font_sub
        ws.cell(row=7, column=c_idx, value=fm_name).font = font_sub
        ws.cell(row=8, column=c_idx, value=m['market_name']).font = font_sub
        ws.cell(row=9, column=c_idx, value=m['mpo_code']).font = font_sub
        ws.cell(row=10, column=c_idx, value=m['fm_code']).font = font_sub
        
        # Apply DEEP_NAVY fill to metadata rows
        for r in range(6, 11):
            ws.cell(row=r, column=c_idx).fill = fill_navy
            ws.cell(row=r, column=c_idx).border = bd_hdr
            ws.cell(row=r, column=c_idx).alignment = align_center
            
    # Row 11: Name Row (DARK_SURFACE bg)
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        ws.cell(row=11, column=c_idx, value=m['mpo_name']).font = font_name
        ws.cell(row=11, column=c_idx).fill = fill_dark
        ws.cell(row=11, column=c_idx).border = bd_hdr
        ws.cell(row=11, column=c_idx).alignment = align_center
        
    for d_idx, da_name in enumerate(das):
        c_idx = da_start + d_idx
        ws.cell(row=11, column=c_idx, value=da_name).font = font_name
        ws.cell(row=11, column=c_idx).fill = fill_dark
        ws.cell(row=11, column=c_idx).border = bd_hdr
        ws.cell(row=11, column=c_idx).alignment = align_center
        
    # DATE (B11:B14 merged)
    ws.merge_cells(start_row=11, start_column=2, end_row=14, end_column=2)
    c_date_h = ws.cell(row=11, column=2, value="DATE")
    c_date_h.font = font_hdr
    c_date_h.fill = fill_mid
    c_date_h.alignment = align_center
    
    # FM SELF (C11:C14 merged)
    ws.merge_cells(start_row=11, start_column=3, end_row=14, end_column=3)
    c_fm_h = ws.cell(row=11, column=3, value="FM SELF")
    c_fm_h.font = font_hdr
    c_fm_h.fill = fill_mid
    c_fm_h.alignment = align_center
    
    for r in range(11, 15):
        ws.cell(row=r, column=2).fill = fill_mid
        ws.cell(row=r, column=2).border = bd_hdr
        ws.cell(row=r, column=3).fill = fill_mid
        ws.cell(row=r, column=3).border = bd_hdr
        
    # Rows 12 to 14 metadata / hidden info for MPOs
    for m_idx, m in enumerate(markets):
        c_idx = mpo_start + m_idx
        ws.cell(row=12, column=c_idx, value=m['desig']).font = font_sub
        ws.cell(row=13, column=c_idx, value=m['is_vacant']).font = font_sub
        ws.cell(row=14, column=c_idx, value=m['da_name']).font = font_sub
        for r in range(12, 15):
            ws.cell(row=r, column=c_idx).fill = fill_navy
            ws.cell(row=r, column=c_idx).border = bd_hdr
            ws.cell(row=r, column=c_idx).alignment = align_center
            
    # Rows 15 to 17 (hidden buffer rows, formulas active in total col)
    for r in range(15, 18):
        c_tot = ws.cell(row=r, column=total_col, value=f"=SUM(C{r}:{get_column_letter(total_col-1)}{r})")
        c_tot.font = font_total_d
        c_tot.alignment = align_center
        c_tot.fill = fill_tot_data
        c_tot.border = bd_total
        for c in range(2, total_col):
            ws.cell(row=r, column=c).border = bd_data
            
    # Rows 18 to 48 (dates & formulas)
    for r_idx, date_val in enumerate(dates):
        row_num = 18 + r_idx
        # Even rows: Pure White (ZEBRA_A), Odd rows: Zebra Light (ZEBRA_B)
        row_fill = fill_za if (r_idx % 2 == 0) else fill_zb
        
        c_val = ws.cell(row=row_num, column=2, value=date_val)
        c_val.font = font_date
        c_val.alignment = align_center
        c_val.border = bd_date
        c_val.fill = row_fill
        
        for c in range(3, total_col):
            cell = ws.cell(row=row_num, column=c)
            cell.border = bd_data
            cell.alignment = align_right
            cell.number_format = '#,##0'
            cell.font = font_body
            cell.fill = row_fill
            
        c_tot = ws.cell(row=row_num, column=total_col, value=f"=SUM(C{row_num}:{get_column_letter(total_col-1)}{row_num})")
        c_tot.font = font_total_d
        c_tot.alignment = align_right
        c_tot.border = bd_total
        c_tot.fill = fill_tot_data
        c_tot.number_format = '#,##0'
        
    # Hide vacant MPO columns
    for m_idx, m in enumerate(markets):
        if m['is_vacant'] == 'Y':
            col_letter = get_column_letter(mpo_start + m_idx)
            ws.column_dimensions[col_letter].hidden = True
            
    # Hide metadata rows
    for r in [6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17]:
        ws.row_dimensions[r].hidden = True
        
    # Set explicit column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 13
    
    # MPO & DA columns width: max(name_length * 1.1, 13), min=13
    for m_idx, m in enumerate(markets):
        c_let = get_column_letter(mpo_start + m_idx)
        name_len = len(m['mpo_name'] or '')
        ws.column_dimensions[c_let].width = max(name_len * 1.1, 13)
        
    for d_idx, da_name in enumerate(das):
        c_let = get_column_letter(da_start + d_idx)
        name_len = len(da_name or '')
        ws.column_dimensions[c_let].width = max(name_len * 1.1, 13)
        
    # TOTAL column: width 18 (fixed)
    ws.column_dimensions[get_column_letter(total_col)].width = 18
    
    # Save file
    zone_dir = os.path.join(base_dir, zone)
    os.makedirs(zone_dir, exist_ok=True)
    
    fm_filename = fm_name.split(',')[0].strip()
    # Replace invalid chars just in case
    for char in ['/', '\\', '?', '*', ':', '|', '"', '<', '>']:
        fm_filename = fm_filename.replace(char, '')
        
    file_path = os.path.join(zone_dir, f"{fm_filename}.xlsx")
    wb.save(file_path)
    print(f"Generated futuristic 50k sheet: {file_path}")

print("All FM-wise files updated with premium 50k AD styling!")
