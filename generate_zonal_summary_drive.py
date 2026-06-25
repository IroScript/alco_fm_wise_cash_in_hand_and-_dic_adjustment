import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import re
import time
import gspread
from google.oauth2.credentials import Credentials
import googleapiclient.discovery
from googleapiclient.http import MediaFileUpload

# Directory Settings
BASE_DIR = r"c:\Users\Irak\Desktop\Cash in Hand and Dic Adjustment"
TOKEN_FILE = os.path.join(BASE_DIR, "FieldEdit", "token.json")
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]
PARENT_FOLDER_ID = "1iOFeqywnIZ_yVclg_Em2U1npPtsokfGk"

# ── 50K AD Design System ──────────────────────────────
FONT_FAMILY   = 'Aptos'

C_VOID        = '060816'
C_DEEP_NAVY   = '0D1425'
C_MIDNIGHT    = '1E293B'
C_DARK_SURF   = '0F172A'
C_ZEBRA_A     = 'FFFFFF'
C_ZEBRA_B     = 'F8FAFC'
C_TOTAL_DATA  = 'ECFDF5'
C_TOTAL_HEAD  = '065F46'

T_NEON        = '00F2FE'
T_WHITE       = 'FFFFFF'
T_SLATE       = 'CBD5E1'
T_INK         = '0F172A'
T_MINT        = 'A7F3D0'
T_TOTAL_DARK  = '064E3B'
T_DATE        = '475569'

B_NEON        = '0E7490'
B_SLATE       = '334155'
B_LIGHT       = 'E2E8F0'
B_TOTAL       = '6EE7B7'

# Fonts
font_title    = Font(name=FONT_FAMILY, size=18, bold=True,  color=T_NEON)
font_hdr      = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
font_sub      = Font(name=FONT_FAMILY, size=9,  bold=True,  color=T_SLATE)
font_name     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
font_date     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_DATE)
font_body     = Font(name=FONT_FAMILY, size=10, bold=False, color=T_INK)
font_total_h  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_MINT)
font_total_d  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_TOTAL_DARK)

# Fills
fill_void     = PatternFill('solid', start_color=C_VOID,       end_color=C_VOID)
fill_navy     = PatternFill('solid', start_color=C_DEEP_NAVY,  end_color=C_DEEP_NAVY)
fill_mid      = PatternFill('solid', start_color=C_MIDNIGHT,   end_color=C_MIDNIGHT)
fill_dark     = PatternFill('solid', start_color=C_DARK_SURF,  end_color=C_DARK_SURF)
fill_za       = PatternFill('solid', start_color=C_ZEBRA_A,    end_color=C_ZEBRA_A)
fill_zb       = PatternFill('solid', start_color=C_ZEBRA_B,    end_color=C_ZEBRA_B)
fill_tot_data = PatternFill('solid', start_color=C_TOTAL_DATA, end_color=C_TOTAL_DATA)
fill_tot_head = PatternFill('solid', start_color=C_TOTAL_HEAD, end_color=C_TOTAL_HEAD)

# Border helpers
def side(style='thin', color=B_LIGHT):
    return Side(style=style, color=color)

bd_data  = Border(left=side(), right=side(), top=side(), bottom=side())
bd_hdr   = Border(left=side('thin', B_SLATE), right=side('thin', B_SLATE),
                  top=side('thin', B_SLATE),  bottom=side('thin', B_SLATE))
bd_total = Border(left=side('thin', B_TOTAL), right=side('thin', B_TOTAL),
                  top=side('thin', B_TOTAL),  bottom=side('thin', B_TOTAL))
bd_date  = Border(left=side('medium', B_SLATE), right=side('thin', B_LIGHT),
                  top=side('thin', B_LIGHT),   bottom=side('thin', B_LIGHT))

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_right  = Alignment(horizontal='right', vertical='center')

# 31 JUL'26 down to 1 JUL'26
dates = [f"{d} JUL'26" for d in range(31, 0, -1)]

def clean_person_name(name):
    if not name:
        return ""
    name = str(name).strip()
    name = re.sub(r'^(MR|MD|MRS|MST|DR)\.?\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(MR|MD|MRS|MST|DR)\.?\s+', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', ' ', name)
    return name.strip().upper()

def get_or_create_drive_folder(drive_service, name, parent_id):
    query = f"name = '{name}' and mimeType = 'application/vnd.google-apps.folder' and '{parent_id}' in parents and trashed = false"
    response = drive_service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    files = response.get('files', [])
    if files:
        return files[0]['id']
    
    file_metadata = {
        'name': name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }
    folder = drive_service.files().create(body=file_metadata, fields='id').execute()
    return folder['id']

def generate_local_zonal_excel(zone, zone_fms, registry_map, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JUL'26"
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_view.zoomScale = 90
    
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[11].height = 26
    for r in range(18, 49):
        ws.row_dimensions[r].height = 20
        
    ws.freeze_panes = 'E18'
    ws.sheet_properties.tabColor = '00F2FE' if zone == 'CTG.A' else 'A855F7'
    
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14  # DATE
    ws.column_dimensions['C'].width = 18  # Zone Summary
    ws.column_dimensions['D'].width = 13  # SH, SELF
    
    num_fms = len(zone_fms)
    zone_summary_col = 3
    sh_self_col = 4
    summary_start_col = 5
    summary_end_col = 4 + num_fms
    
    # Summary header
    ws.merge_cells(start_row=5, start_column=summary_start_col, end_row=5, end_column=summary_end_col)
    c_sum_hdr = ws.cell(row=5, column=summary_start_col, value=f"FM WISE TOTAL CASH IN HAND, {zone}")
    c_sum_hdr.font = font_total_h
    c_sum_hdr.fill = fill_tot_head
    c_sum_hdr.alignment = align_center
    
    for col_idx in range(summary_start_col, summary_end_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.cell(row=5, column=col_idx).fill = fill_tot_head
        ws.cell(row=5, column=col_idx).border = bd_total
        
    # Zone Summary header
    ws.merge_cells(start_row=5, start_column=zone_summary_col, end_row=11, end_column=zone_summary_col)
    c_zone_hdr = ws.cell(row=5, column=zone_summary_col, value=f"ZONE SUMMARY\n{zone}")
    c_zone_hdr.font = font_total_h
    c_zone_hdr.fill = fill_tot_head
    c_zone_hdr.alignment = align_center
    for r in range(5, 12):
        ws.cell(row=r, column=zone_summary_col).fill = fill_tot_head
        ws.cell(row=r, column=zone_summary_col).border = bd_total
        
    # SH, SELF header
    ws.merge_cells(start_row=5, start_column=sh_self_col, end_row=11, end_column=sh_self_col)
    c_sh_hdr = ws.cell(row=5, column=sh_self_col, value="SH, SELF")
    c_sh_hdr.font = font_total_h
    c_sh_hdr.fill = fill_tot_head
    c_sh_hdr.alignment = align_center
    for r in range(5, 12):
        ws.cell(row=r, column=sh_self_col).fill = fill_tot_head
        ws.cell(row=r, column=sh_self_col).border = bd_total
        
    # FM Names in summary headers
    for f_idx, (fm_name, fm_data) in enumerate(zone_fms.items()):
        col_idx = summary_start_col + f_idx
        fm_clean_name = fm_name.split(',')[0].strip()
        
        ws.merge_cells(start_row=6, start_column=col_idx, end_row=11, end_column=col_idx)
        c_fm_summary_hdr = ws.cell(row=6, column=col_idx, value=fm_clean_name)
        c_fm_summary_hdr.font = font_total_h
        c_fm_summary_hdr.fill = fill_tot_head
        c_fm_summary_hdr.alignment = align_center
        for r in range(6, 12):
            ws.cell(row=r, column=col_idx).fill = fill_tot_head
            ws.cell(row=r, column=col_idx).border = bd_total
            
    # Build individual columns side-by-side
    start_data_col = summary_end_col + 1
    fm_col_ranges = []
    
    curr_col = start_data_col
    for fm_name, fm_data in zone_fms.items():
        fm_clean_name = fm_name.split(',')[0].strip()
        markets = fm_data['markets']
        
        # Get unique DAs
        das = []
        for m in markets:
            if m['da_name'] and str(m['da_name']).strip().upper() != 'VACANT':
                da_str = str(m['da_name']).strip().upper()
                if da_str not in das:
                    das.append(da_str)
                    
        num_mpos = len(markets)
        num_das = len(das)
        
        fm_start = curr_col
        fm_end = curr_col + 1 + num_mpos + num_das - 1
        fm_col_ranges.append((fm_start, fm_end, fm_clean_name))
        
        # FM SELF Column
        ws.merge_cells(start_row=11, start_column=curr_col, end_row=14, end_column=curr_col)
        c_fm_self_h = ws.cell(row=11, column=curr_col, value="FM SELF")
        c_fm_self_h.font = font_hdr
        c_fm_self_h.fill = fill_mid
        c_fm_self_h.alignment = align_center
        for r in range(11, 15):
            ws.cell(row=r, column=curr_col).fill = fill_mid
            ws.cell(row=r, column=curr_col).border = bd_hdr
            
        ws.cell(row=6, column=curr_col, value=zone).font = font_sub
        ws.cell(row=7, column=curr_col, value=fm_clean_name).font = font_sub
        for r in range(6, 11):
            ws.cell(row=r, column=curr_col).fill = fill_navy
            ws.cell(row=r, column=curr_col).border = bd_hdr
            ws.cell(row=r, column=curr_col).alignment = align_center
        ws.column_dimensions[get_column_letter(curr_col)].width = 13
        
        # MPO Columns
        mpo_start = curr_col + 1
        mpo_end = curr_col + num_mpos
        
        ws.merge_cells(start_row=5, start_column=mpo_start, end_row=5, end_column=mpo_end)
        c_mpo_h = ws.cell(row=5, column=mpo_start, value="MPO")
        c_mpo_h.font = font_hdr
        c_mpo_h.fill = fill_mid
        c_mpo_h.alignment = align_center
        
        for c in range(mpo_start, mpo_end + 1):
            ws.cell(row=5, column=c).fill = fill_mid
            ws.cell(row=5, column=c).border = bd_hdr
            
        for m_idx, m in enumerate(markets):
            c_idx = mpo_start + m_idx
            ws.cell(row=6, column=c_idx, value=zone).font = font_sub
            ws.cell(row=7, column=c_idx, value=fm_clean_name).font = font_sub
            ws.cell(row=8, column=c_idx, value=m['market_name']).font = font_sub
            ws.cell(row=9, column=c_idx, value=m['mpo_code']).font = font_sub
            ws.cell(row=10, column=c_idx, value=m['fm_code']).font = font_sub
            
            for r in range(6, 11):
                ws.cell(row=r, column=c_idx).fill = fill_navy
                ws.cell(row=r, column=c_idx).border = bd_hdr
                ws.cell(row=r, column=c_idx).alignment = align_center
                
            ws.cell(row=11, column=c_idx, value=m['mpo_name']).font = font_name
            ws.cell(row=11, column=c_idx).fill = fill_dark
            ws.cell(row=11, column=c_idx).border = bd_hdr
            ws.cell(row=11, column=c_idx).alignment = align_center
            
            ws.cell(row=12, column=c_idx, value=m['desig']).font = font_sub
            ws.cell(row=13, column=c_idx, value=m['is_vacant']).font = font_sub
            ws.cell(row=14, column=c_idx, value=m['da_name']).font = font_sub
            for r in range(12, 15):
                ws.cell(row=r, column=c_idx).fill = fill_navy
                ws.cell(row=r, column=c_idx).border = bd_hdr
                ws.cell(row=r, column=c_idx).alignment = align_center
                
            name_len = len(m['mpo_name'] or '')
            ws.column_dimensions[get_column_letter(c_idx)].width = max(name_len * 1.1, 13)
            
            if m['is_vacant'] == 'Y':
                ws.column_dimensions[get_column_letter(c_idx)].hidden = True
                
        # DA Columns
        if num_das > 0:
            da_start = curr_col + 1 + num_mpos
            da_end = curr_col + num_mpos + num_das
            
            ws.merge_cells(start_row=5, start_column=da_start, end_row=5, end_column=da_end)
            c_da_h = ws.cell(row=5, column=da_start, value="DA")
            c_da_h.font = font_hdr
            c_da_h.fill = fill_mid
            c_da_h.alignment = align_center
            
            for c in range(da_start, da_end + 1):
                ws.cell(row=5, column=c).fill = fill_mid
                ws.cell(row=5, column=c).border = bd_hdr
                
            for d_idx, da_name in enumerate(das):
                c_idx = da_start + d_idx
                ws.cell(row=11, column=c_idx, value=da_name).font = font_name
                ws.cell(row=11, column=c_idx).fill = fill_dark
                ws.cell(row=11, column=c_idx).border = bd_hdr
                ws.cell(row=11, column=c_idx).alignment = align_center
                
                for r in range(5, 11):
                    if r >= 6:
                        ws.cell(row=r, column=c_idx).fill = fill_navy
                        ws.cell(row=r, column=c_idx).border = bd_hdr
                for r in range(12, 15):
                    ws.cell(row=r, column=c_idx).fill = fill_navy
                    ws.cell(row=r, column=c_idx).border = bd_hdr
                    
                name_len = len(da_name or '')
                ws.column_dimensions[get_column_letter(c_idx)].width = max(name_len * 1.1, 13)
                
        curr_col = fm_end + 1
        
    # DATE Header
    ws.merge_cells(start_row=11, start_column=2, end_row=14, end_column=2)
    c_date_h = ws.cell(row=11, column=2, value="DATE")
    c_date_h.font = font_hdr
    c_date_h.fill = fill_mid
    c_date_h.alignment = align_center
    for r in range(11, 15):
        ws.cell(row=r, column=2).fill = fill_mid
        ws.cell(row=r, column=2).border = bd_hdr
        
    last_col_idx = curr_col - 1
    
    # Title block
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=last_col_idx)
    title_cell = ws.cell(row=2, column=2, value="CASH IN HAND")
    title_cell.font = font_title
    title_cell.fill = fill_void
    title_cell.alignment = align_center
    for r in range(2, 4):
        for c in range(2, last_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_void
            left = side('medium', B_NEON) if c == 2 else None
            right = side('medium', B_NEON) if c == last_col_idx else None
            top = side('medium', B_NEON) if r == 2 else None
            bottom = side('medium', B_NEON) if r == 3 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            
    # Rows 15 to 17 buffer and summary formulas
    for r in range(15, 18):
        for c in range(2, last_col_idx + 1):
            ws.cell(row=r, column=c).border = bd_data
            
        # Zone Summary
        ws.cell(row=r, column=zone_summary_col, value=f"=SUM({get_column_letter(sh_self_col)}{r}:{get_column_letter(summary_end_col)}{r})").font = font_total_d
        ws.cell(row=r, column=zone_summary_col).alignment = align_center
        ws.cell(row=r, column=zone_summary_col).fill = fill_tot_data
        ws.cell(row=r, column=zone_summary_col).border = bd_total
        
        # SH, SELF
        ws.cell(row=r, column=sh_self_col, value=0).font = font_total_d
        ws.cell(row=r, column=sh_self_col).alignment = align_center
        ws.cell(row=r, column=sh_self_col).fill = fill_tot_data
        ws.cell(row=r, column=sh_self_col).border = bd_total
        
        for f_idx, (f_start, f_end, _) in enumerate(fm_col_ranges):
            col_idx = summary_start_col + f_idx
            ws.cell(row=r, column=col_idx, value=f"=SUM({get_column_letter(f_start)}{r}:{get_column_letter(f_end)}{r})").font = font_total_d
            ws.cell(row=r, column=col_idx).alignment = align_center
            ws.cell(row=r, column=col_idx).fill = fill_tot_data
            ws.cell(row=r, column=col_idx).border = bd_total
            
    # Rows 18 to 48 data & formulas
    for r_idx, date_val in enumerate(dates):
        row_num = 18 + r_idx
        row_fill = fill_za if (r_idx % 2 == 0) else fill_zb
        
        c_val = ws.cell(row=row_num, column=2, value=date_val)
        c_val.font = font_date
        c_val.alignment = align_center
        c_val.border = bd_date
        c_val.fill = row_fill
        
        # Zone Summary
        zs_cell = ws.cell(row=row_num, column=zone_summary_col, value=f"=SUM({get_column_letter(sh_self_col)}{row_num}:{get_column_letter(summary_end_col)}{row_num})")
        zs_cell.font = font_total_d
        zs_cell.alignment = align_right
        zs_cell.border = bd_total
        zs_cell.fill = fill_tot_data
        zs_cell.number_format = '#,##0'
        
        # SH, SELF
        sh_cell = ws.cell(row=row_num, column=sh_self_col, value=0)
        sh_cell.font = font_total_d
        sh_cell.alignment = align_right
        sh_cell.border = bd_total
        sh_cell.fill = fill_tot_data
        sh_cell.number_format = '#,##0'
        
        # Summaries
        for f_idx, (f_start, f_end, _) in enumerate(fm_col_ranges):
            col_idx = summary_start_col + f_idx
            sum_cell = ws.cell(row=row_num, column=col_idx, value=f"=SUM({get_column_letter(f_start)}{row_num}:{get_column_letter(f_end)}{row_num})")
            sum_cell.font = font_total_d
            sum_cell.alignment = align_right
            sum_cell.border = bd_total
            sum_cell.fill = fill_tot_data
            sum_cell.number_format = '#,##0'
            
        # Format empty individual columns (which will be populated by IMPORTRANGE in row 18)
        for c in range(start_data_col, last_col_idx + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.border = bd_data
            cell.alignment = align_right
            cell.number_format = '#,##0'
            cell.font = font_body
            cell.fill = row_fill
            
    # Write the IMPORTRANGE formulas in row 18 for each FM
    for f_start, f_end, fm_clean_name in fm_col_ranges:
        fm_url = registry_map.get(fm_clean_name)
        if fm_url:
            # We want to import columns C to total_col-1 from the FM sheet
            # The count of columns is (f_end - f_start + 1)
            # The columns in the FM sheet are C (3) to C + count - 1
            col_count = f_end - f_start + 1
            fm_last_col_letter = get_column_letter(3 + col_count - 1)
            
            importrange_formula = f'=IMPORTRANGE("{fm_url}", "JUN\'26!C18:{fm_last_col_letter}48")'
            ws.cell(row=18, column=f_start, value=importrange_formula)
            print(f"Set IMPORTRANGE for {fm_clean_name} at cell {get_column_letter(f_start)}18")
            
    # Hide metadata rows in summary sheet
    for r in [6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17]:
        ws.row_dimensions[r].hidden = True
        
    # Freeze the DATE column (B) and metadata/title rows (1-17)
    ws.freeze_panes = 'C18'
        
    wb.save(output_path)
    wb.close()

def main():
    # 1. Load Registry
    print("Reading registry from Master_Registry_Cash_In_Hand...")
    creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    gc = gspread.authorize(creds)
    drive_service = googleapiclient.discovery.build('drive', 'v3', credentials=creds)
    
    registry_name = "Master_Registry_Cash_In_Hand"
    query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
    res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
    files = res.get('files', [])
    if not files:
        print("Error: Master Registry sheet not found in Drive!")
        return
        
    reg_sheet = gc.open_by_key(files[0]['id'])
    reg_ws = reg_sheet.get_worksheet(0)
    rows = reg_ws.get_all_records()
    
    registry_map = {}
    for r in rows:
        fm = r.get('FM Name')
        url = r.get('URL')
        if fm and url:
            registry_map[clean_person_name(fm)] = url
            
    # 2. Parse Force hierarchy
    wb_local = openpyxl.load_workbook(os.path.join(BASE_DIR, "Cash in Hand.xlsx"))
    ws_local = wb_local["FM wise DA and MPO Names"]
    
    fm_groups = {}
    for r in range(2, ws_local.max_row + 1):
        depot = ws_local.cell(row=r, column=1).value
        zone = ws_local.cell(row=r, column=2).value
        market = ws_local.cell(row=r, column=3).value
        mpo_name = ws_local.cell(row=r, column=4).value
        fm_am_zone = ws_local.cell(row=r, column=5).value
        vacant = ws_local.cell(row=r, column=6).value
        desig = ws_local.cell(row=r, column=7).value
        mpo_code = ws_local.cell(row=r, column=8).value
        fm_code = ws_local.cell(row=r, column=9).value
        da_names = [ws_local.cell(row=r, column=c).value for c in range(10, 14)]
        da_names = [da for da in da_names if da and str(da).strip() and str(da).strip().upper() != 'VACANT']
        
        if not fm_am_zone or str(fm_am_zone).strip() == "":
            continue
            
        fm_am_zone = str(fm_am_zone).strip()
        if ',' in fm_am_zone:
            parts = fm_am_zone.split(',')
            cleaned_fm_part = clean_person_name(parts[0])
            fm_am_zone_clean = f"{cleaned_fm_part}, {parts[1].strip()}"
        else:
            fm_am_zone_clean = clean_person_name(fm_am_zone)
            
        if fm_am_zone_clean not in fm_groups:
            fm_groups[fm_am_zone_clean] = {
                'depot': depot,
                'zone': str(zone).strip() if zone else "",
                'markets': []
            }
            
        is_vacant = (vacant == 'Y') or (mpo_name and 'vacant' in str(mpo_name).lower())
        mpo_clean = 'VACANT' if is_vacant else clean_person_name(mpo_name)
        da_clean = clean_person_name(da_names[0]) if da_names else None
        
        fm_groups[fm_am_zone_clean]['markets'].append({
            'market_name': market,
            'mpo_name': mpo_clean,
            'desig': desig,
            'mpo_code': mpo_code,
            'fm_code': fm_code,
            'is_vacant': 'Y' if is_vacant else None,
            'da_name': da_clean
        })
    wb_local.close()
    
    # Filter valid FMs
    valid_fms = {}
    for fm_name, fm_data in fm_groups.items():
        non_vacant_mpos = [m for m in fm_data['markets'] if not m['is_vacant']]
        if non_vacant_mpos:
            valid_fms[fm_name] = fm_data
            
    # Group valid FMs by Zone
    zone_groups = {}
    for fm_name, fm_data in valid_fms.items():
        zone = fm_data['zone']
        if zone not in zone_groups:
            zone_groups[zone] = {}
        zone_groups[zone][fm_name] = fm_data

    # Fetch SH email per zone from EMAIL_2026 sheet
    sh_by_zone = {}
    try:
        EMAIL_SHEET_ID = "1f5SFvhH8Bjb3OUlpof68teBktHuYyVELioxLv_KWXJo"
        email_ws = gc.open_by_key(EMAIL_SHEET_ID).worksheet("EMAIL_2026")
        email_rows = email_ws.get_all_values()
        email_headers = [h.strip().upper() for h in email_rows[0]]
        z_col = email_headers.index('ZONE') if 'ZONE' in email_headers else -1
        sh_col = email_headers.index('SH EMAIL') if 'SH EMAIL' in email_headers else -1
        for er in email_rows[1:]:
            if z_col != -1 and sh_col != -1 and len(er) > max(z_col, sh_col):
                z = str(er[z_col]).strip().upper()
                sh = str(er[sh_col]).strip()
                if z and sh and '@' in sh and z not in sh_by_zone:
                    sh_by_zone[z] = sh
        print(f"SH emails mapped: {sh_by_zone}")
    except Exception as e:
        print(f"Could not fetch SH emails: {e}")

    for zone, zone_fms in zone_groups.items():
        print(f"\n--- Processing Zonal Summary for {zone} ---")
        
        # Sort FMs to match sample order exactly
        def fm_sort_key(item):
            fm_name = item[0]
            fm_clean = fm_name.split(',')[0].strip().upper()
            if zone == 'CTG.B':
                order = ["JAINAL ABEDIN AKHAND", "FIROZ AHMED", "NARAYAN DAS", "SHAHJAHAN"]
            else:
                order = ["MONIR UDDIN", "RAFIQUL MOULA", "VACANT, KHAGRACHARI"]
            for idx, name in enumerate(order):
                if name in fm_clean:
                    return idx
            return 99
            
        sorted_zone_fms = dict(sorted(zone_fms.items(), key=fm_sort_key))
        
        # Create folder if not exists on Drive
        zone_folder_id = get_or_create_drive_folder(drive_service, zone, PARENT_FOLDER_ID)
        
        local_excel_path = os.path.join(BASE_DIR, f"{zone}_CASH_IN_HAND_Summary_temp.xlsx")
        generate_local_zonal_excel(zone, sorted_zone_fms, registry_map, local_excel_path)
        
        # Check if Google Sheet already exists in this Zone folder
        sheet_name = f"{zone} CASH IN HAND"
        query = f"name = '{sheet_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{zone_folder_id}' in parents and trashed = false"
        res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
        files = res.get('files', [])
        
        if files:
            # Trash old one
            drive_service.files().update(fileId=files[0]['id'], body={'trashed': True}).execute()
            print(f"Trashed old summary sheet: {sheet_name}")
            
        # Upload as a new Google Sheet
        file_metadata = {
            'name': sheet_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet',
            'parents': [zone_folder_id]
        }
        media = MediaFileUpload(local_excel_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        uploaded_file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id, webViewLink'
        ).execute()
        
        zonal_sheet_id = uploaded_file.get('id')
        print(f"Uploaded Zonal Summary Google Sheet successfully. ID: {zonal_sheet_id}")

        # Apply sheet protection: lock whole sheet, only D18:D48 editable for SH
        sh_email = sh_by_zone.get(zone.upper(), '')
        try:
            sheets_api = googleapiclient.discovery.build('sheets', 'v4', credentials=creds)
            ss_meta = sheets_api.spreadsheets().get(spreadsheetId=zonal_sheet_id, fields='sheets(properties(sheetId,title))').execute()
            real_sheet_id = ss_meta['sheets'][0]['properties']['sheetId']
            requests_body = [{
                'addProtectedRange': {
                    'protectedRange': {
                        'range': {'sheetId': real_sheet_id},
                        'description': f'Zonal Summary {zone} locked - SH edits only D18:D48',
                        'warningOnly': False,
                        'editors': {'users': [sh_email] if sh_email else []}
                    }
                }
            }, {
                'addProtectedRange': {
                    'protectedRange': {
                        'range': {
                            'sheetId': real_sheet_id,
                            'startRowIndex': 17, 'endRowIndex': 48,
                            'startColumnIndex': 3, 'endColumnIndex': 4
                        },
                        'description': f'SH editable range for {zone}',
                        'warningOnly': False,
                        'editors': {'users': [sh_email] if sh_email else []}
                    }
                }
            }]
            sheets_api.spreadsheets().batchUpdate(spreadsheetId=zonal_sheet_id, body={'requests': requests_body}).execute()
            print(f"Applied protection: locked all except D18:D48 for SH {sh_email}")

            # Add data validation: D18:D48 only accepts numeric values >= 0
            try:
                validation_requests = [{
                    'setDataValidation': {
                        'range': {
                            'sheetId': real_sheet_id,
                            'startRowIndex': 17, 'endRowIndex': 48,
                            'startColumnIndex': 3, 'endColumnIndex': 4
                        },
                        'rule': {
                            'condition': {
                                'type': 'NUMBER_GREATER_THAN_EQ',
                                'values': [{'userEnteredValue': '0'}]
                            },
                            'inputMessage': 'Enter a positive number (0 or greater). Letters and text are not allowed.',
                            'strict': True,
                            'showCustomUi': True
                        }
                    }
                }]
                sheets_api.spreadsheets().batchUpdate(spreadsheetId=zonal_sheet_id, body={'requests': validation_requests}).execute()
                print(f"Applied number-only validation on D18:D48 for {zone} summary (SH: {sh_email})")
            except Exception as ve:
                print(f"Non-fatal: could not apply data validation to {zone}: {ve}")
        except Exception as e:
            print(f"Non-fatal: could not apply protection to {zone} summary: {e}")

        # Share zonal summary with SH (writer so they can edit D18:D48)
        if sh_email:
            try:
                user_permission = {
                    'type': 'user',
                    'role': 'writer',
                    'emailAddress': sh_email
                }
                drive_service.permissions().create(
                    fileId=zonal_sheet_id,
                    body=user_permission,
                    fields='id'
                ).execute()
                print(f"Shared Zonal Summary {zone} with SH {sh_email}")
            except Exception as e:
                print(f"Error sharing zonal summary with SH: {e}")

        # Delete local temp file
        try:
            if os.path.exists(local_excel_path):
                os.remove(local_excel_path)
        except Exception as e:
            print(f"Non-fatal clean up error: {e}")
            
        time.sleep(1)

if __name__ == "__main__":
    main()
