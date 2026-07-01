import os
import re
import time
import datetime
import calendar
import json
import socket
import threading
import webbrowser
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import gspread
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# Directory Settings
BASE_DIR = r"c:\Users\Irak\Desktop\Cash in Hand and Dic Adjustment"
CLIENT_SECRET_FILE = os.path.join(BASE_DIR, "FieldEdit", "client_secret_866102064521-5g6tq5989nqs97ehgse7n6fl1o9pslt5.apps.googleusercontent.com.json")
TOKEN_FILE = os.path.join(BASE_DIR, "FieldEdit", "token.json")
PARENT_FOLDER_ID = "1iOFeqywnIZ_yVclg_Em2U1npPtsokfGk"
EMAIL_SHEET_ID = "1f5SFvhH8Bjb3OUlpof68teBktHuYyVELioxLv_KWXJo"
DATA_SHEET_ID = "1ywTyruBLxNXz6pjsGgufNstb0hOsrM9P-ER65iVvqN8"  # Operational Field Force Sheet
COMPANY_MASTER_SHEET_ID = "1Q4utivZ5OpgDznqlqElYU-HWNnZYI71YYpcZKcSM3xY"  # Company Hardcoded Master Sheet for Sync Cross-Checking

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

# Shared GUI log utility
log_callback = None

def log_message(msg):
    print(msg)
    if log_callback:
        log_callback(msg + "\n")

# Online / Offline Verification Engine
def check_online_status(timeout=3):
    log_message("Checking internet connection (Online/Offline status)...")
    try:
        # Try connecting to Google DNS or public server
        socket.setdefaulttimeout(timeout)
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect(("8.8.8.8", 53))
        s.close()
        socket.setdefaulttimeout(None) # Reset timeout so Google API requests do not time out!
        log_message("Internet connection verified: ONLINE 🟢")
        return True
    except Exception as e:
        socket.setdefaulttimeout(None)
        log_message("Error: No internet connection detected: OFFLINE 🔴")
        try:
            messagebox.showerror("Offline Alert", "System is Offline!\nPlease check your internet connection and try again.")
        except Exception:
            pass
        return False

# Asia/Dhaka Timezone Utility (UTC+06:00)
def get_dhaka_today():
    dhaka_tz = datetime.timezone(datetime.timedelta(hours=6))
    now_dhaka = datetime.datetime.now(dhaka_tz)
    return now_dhaka.date()

# Oauth Credentials Loader
def get_oauth_credentials():
    creds = None
    if os.path.exists(TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception:
            creds = None
    if not creds or not creds.valid:
        refreshed = False
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                refreshed = True
            except Exception as e:
                print(f"Warning: OAuth token refresh failed ({e}). Re-authenticating via browser...")
                creds = None
        if not refreshed:
            if os.path.exists(TOKEN_FILE):
                try:
                    os.remove(TOKEN_FILE)
                except Exception:
                    pass
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'w') as token:
            token.write(creds.to_json())
    return creds

# String Cleaners
def clean_person_name(name, zone=None):
    if not name:
        return ""
    name = str(name).strip().upper()
    name = re.sub(r'^(MR|MD|MRS|MST|DR)\.?\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(MR|MD|MRS|MST|DR)\.?\s+', '', name, flags=re.IGNORECASE)
    if zone:
        z = str(zone).strip().upper()
        if name.endswith(',' + z):
            name = name[:-len(',' + z)].strip()
        elif name.endswith(', ' + z):
            name = name[:-len(', ' + z)].strip()
        elif name.endswith('(' + z + ')'):
            name = name[:-len('(' + z + ')')].strip()
    elif ',' in name and 'VACANT' not in name and 'SH' not in name and 'ARM' not in name and 'RSM' not in name:
        name = name.split(',')[0].strip()
    name = re.sub(r'\s+', ' ', name)
    return name.strip()

# Date / Month Parsing & Calculations
def parse_month_str(m_str):
    try:
        parts = m_str.split("'")
        if len(parts) == 2:
            m_abbr, y_short = parts[0].upper(), parts[1]
            m_num = datetime.datetime.strptime(m_abbr, "%b").month
            year = 2000 + int(y_short)
            return (year, m_num)
    except Exception:
        pass
    return None

def format_month_str(year, month):
    m_abbr = calendar.month_abbr[month].upper()
    year_short = str(year)[2:]
    return f"{m_abbr}'{year_short}"

def get_next_month_info(target_date=None):
    if target_date is None:
        target_date = get_dhaka_today()
    first_next = target_date.replace(day=1) + datetime.timedelta(days=32)
    next_month_date = first_next.replace(day=1)
    year = next_month_date.year
    month = next_month_date.month
    month_str = format_month_str(year, month)
    _, num_days = calendar.monthrange(year, month)
    return month_str, num_days

def get_current_month_info(target_date=None):
    if target_date is None:
        target_date = get_dhaka_today()
    year = target_date.year
    month = target_date.month
    month_str = format_month_str(year, month)
    _, num_days = calendar.monthrange(year, month)
    return month_str, num_days

# Universal Registry Header Lookup Helper
def get_reg_col_indices(headers):
    headers = [str(h).strip().upper() for h in headers] if headers else []
    def get_c(possible, fall):
        for p in possible:
            for idx, h in enumerate(headers):
                if p == h or (p in h and len(h) < len(p) + 8):
                    return idx
        return fall
    return {
        'fm': get_c(['FM NAME', 'FM/AM'], 0),
        'zone': get_c(['ZONE'], 1),
        'sheet_id': get_c(['SHEET ID', 'ID'], 2),
        'url': get_c(['URL', 'LINK'], 3),
        'fm_email': get_c(['FM EMAIL', 'EMAIL'], 4),
        'boss_name': get_c(['BOSS NAME', 'SM NAME'], 5),
        'boss_email': get_c(['BOSS EMAIL', 'SM/DSM MAIL'], 6),
        'sh_email': get_c(['SH EMAIL', 'SH MAIL'], 7)
    }

# Google Sheets Data Fetcher
def fetch_master_data(gc):
    log_message("Fetching master data from Google Sheets (MPO/FM)...")
    sheet = gc.open_by_key(DATA_SHEET_ID)
    ws = None
    for w in sheet.worksheets():
        if "MPO" in w.title.upper() or "FM" in w.title.upper():
            ws = w
            break
    if ws is None:
        ws = sheet.get_worksheet(0)
    rows = ws.get_all_values()
    
    headers = [str(h).strip().upper() for h in rows[0]] if rows else []
    def get_col_idx(possible_names, fallback):
        for n in possible_names:
            for idx, h in enumerate(headers):
                if n == h or (n in h and len(h) < len(n) + 8):
                    return idx
        return fallback

    depot_col = get_col_idx(['DEPOT', 'DEPOT NAME'], 0)
    zone_col = get_col_idx(['ZONE', 'ZONE NAME'], 1)
    market_col = get_col_idx(['MARKET', 'MARKET NAME', 'TERRITORY'], 2)
    mpo_col = get_col_idx(['MPO', 'MPO NAME', 'MPO/AM'], 3)
    fm_col = get_col_idx(['FM/AM, ZONE', 'FM/AM', 'FM NAME', 'AM NAME'], 4)
    vacant_col = get_col_idx(['VACANT', 'VACANT STATUS'], 5)
    desig_col = get_col_idx(['DESIG', 'DESIGNATION'], 6)
    mpo_code_col = get_col_idx(['MPO CODE', 'MPO ID'], 7)
    fm_code_col = get_col_idx(['FM CODE', 'FM ID', 'AM CODE'], 8)
    
    da_cols = [idx for idx, h in enumerate(headers) if 'DA' in h and ('NAME' in h or len(h) <= 5)]
    if not da_cols:
        da_cols = list(range(9, min(len(headers), 13))) if len(headers) > 9 else [9, 10, 11, 12]

    fm_groups = {}
    for r in rows[1:]:
        max_req_col = max([depot_col, zone_col, market_col, mpo_col, fm_col, vacant_col, desig_col, mpo_code_col, fm_code_col])
        if len(r) <= max_req_col:
            continue
        depot = r[depot_col]
        zone = r[zone_col]
        market = r[market_col]
        mpo_name = r[mpo_col]
        fm_am_zone = r[fm_col]
        vacant = r[vacant_col]
        desig = r[desig_col]
        mpo_code = r[mpo_code_col]
        fm_code = r[fm_code_col]
        da_names = [r[c] for c in da_cols if c < len(r) and r[c] and str(r[c]).strip() and str(r[c]).strip().upper() != 'VACANT']

        if not fm_am_zone or str(fm_am_zone).strip() == "":
            continue

        fm_am_zone = str(fm_am_zone).strip()
        fm_am_zone_clean = clean_person_name(fm_am_zone, zone)
        
        if fm_am_zone_clean not in fm_groups:
            fm_groups[fm_am_zone_clean] = {
                'depot': depot,
                'zone': str(zone).strip() if zone else "",
                'markets': []
            }
        
        is_vacant = (vacant == 'Y') or (mpo_name and 'vacant' in str(mpo_name).lower())
        mpo_clean = 'VACANT' if is_vacant else clean_person_name(mpo_name)
        da_clean = clean_person_name(da_names[0]) if da_names else None
        
        fm_groups[fm_am_zone_clean]['markets'].append({
            'market_name': market,
            'mpo_name': mpo_clean,
            'desig': desig,
            'mpo_code': mpo_code,
            'fm_code': fm_code,
            'is_vacant': 'Y' if is_vacant else None,
            'da_name': da_clean
        })

    valid_fms = {}
    for fm_name, fm_data in fm_groups.items():
        non_vacant_mpos = [m for m in fm_data['markets'] if not m['is_vacant']]
        if non_vacant_mpos:
            valid_fms[fm_name] = fm_data

    log_message(f"Successfully loaded {len(valid_fms)} valid FMs from MPO/FM sheet.")
    return valid_fms

# Company Master Sheet Cross-Checking Engine
def fetch_company_master_data(gc):
    log_message("Fetching Company Hardcoded Master Sheet for cross-checking...")
    sheet = gc.open_by_key(COMPANY_MASTER_SHEET_ID)
    ws = None
    for w in sheet.worksheets():
        if "FIELD" in w.title.upper() or "MPO" in w.title.upper() or "FM" in w.title.upper():
            ws = w
            break
    if ws is None:
        ws = sheet.get_worksheet(0)
    rows = ws.get_all_values()
    
    headers = [str(h).strip().upper() for h in rows[0]] if rows else []
    def get_col_idx(possible_names, fallback):
        for n in possible_names:
            for idx, h in enumerate(headers):
                if n == h or (n in h and len(h) < len(n) + 8):
                    return idx
        return fallback

    depot_col = get_col_idx(['DEPOT', 'DEPOT NAME'], 0)
    zone_col = get_col_idx(['ZONE', 'ZONE NAME'], 1)
    market_col = get_col_idx(['MARKET', 'MARKET NAME', 'TERRITORY'], 3)
    mpo_col = get_col_idx(['MPO', 'MPO NAME', 'RX CODE', 'OLD CODE'], 6)
    fm_col = get_col_idx(['FM/AM, ZONE', 'FM/AM', 'FM NAME', 'AM NAME'], 7)
    vacant_col = get_col_idx(['VACANT', 'VACANT STATUS', "VACANT (JUN'26)?", "VACANT (JAN'26)?"], 8)

    company_fms = {}
    company_markets = set()
    
    for r in rows[1:]:
        # CRITICAL RULE FROM USER: "NICHER DATA WILL NOT BE CONSIDERED IF A COLUMN IS ACTALLY BLANK.... MAANE MSTER FIELD FORCE ER NICHER DATA NEVBE NAA JODI A COLUMN BLANK THAKE JEMON A455 BLANK ER NICHER DATA IS NOT CONSIDERABLE"
        if not r or len(r) <= depot_col or not str(r[depot_col]).strip():
            break
            
        max_req_col = max([depot_col, zone_col, market_col, fm_col])
        if len(r) <= max_req_col:
            continue
            
        zone_str = str(r[zone_col]).strip().upper()
        market_str = str(r[market_col]).strip()
        fm_raw = str(r[fm_col]).strip()
        if not fm_raw or not zone_str or not market_str:
            continue
            
        vacant_val = str(r[vacant_col]).strip().upper() if len(r) > vacant_col else ''
        mpo_val = str(r[mpo_col]).strip() if len(r) > mpo_col else ''
        is_vacant = (vacant_val == 'Y') or (mpo_val and 'vacant' in mpo_val.lower()) or 'vacant' in fm_raw.lower()
        if is_vacant:
            continue
            
        fm_clean = clean_person_name(fm_raw, zone_str)
        if fm_clean not in company_fms:
            company_fms[fm_clean] = {'zone': zone_str, 'markets': set()}
        company_fms[fm_clean]['markets'].add(market_str.upper())
        company_markets.add((zone_str, market_str.upper()))
        
    return company_fms, company_markets

def perform_company_master_sync_check(gc, selected_zones):
    log_message("Cross-checking Operational Field Force with Company Hardcoded Master Sheet...")
    try:
        op_fms = fetch_master_data(gc)
        op_fm_names = set(op_fms.keys())
        op_markets = set()
        for fm_name, f_data in op_fms.items():
            if f_data['zone'] in selected_zones:
                for mkt in f_data['markets']:
                    if not mkt.get('is_vacant'):
                        op_markets.add((f_data['zone'], mkt['market_name'].strip().upper()))
        
        comp_fms, comp_markets = fetch_company_master_data(gc)
        
        missing_fms_in_op = []
        for fm_name, c_data in comp_fms.items():
            if c_data['zone'] in selected_zones and fm_name not in op_fm_names:
                missing_fms_in_op.append(f"[{c_data['zone']}] {fm_name}")
                
        missing_markets_in_op = []
        for z, m_name in comp_markets:
            if z in selected_zones and (z, m_name) not in op_markets:
                missing_markets_in_op.append(f"[{z}] {m_name}")
                
        if missing_fms_in_op or missing_markets_in_op:
            warn_msg = "⚠️ MASTER SHEET SYNC DISCREPANCY DETECTED!\n\n"
            warn_msg += "The following items were found in the Company Hardcoded Master Sheet but are MISSING from your Operational Field Force Sheet:\n\n"
            if missing_fms_in_op:
                warn_msg += f"• Missing FMs ({len(missing_fms_in_op)}):\n"
                for m_fm in sorted(missing_fms_in_op)[:10]:
                    warn_msg += f"  - {m_fm}\n"
                if len(missing_fms_in_op) > 10:
                    warn_msg += f"  ...and {len(missing_fms_in_op)-10} more.\n"
                warn_msg += "\n"
            if missing_markets_in_op:
                warn_msg += f"• Missing Markets ({len(missing_markets_in_op)}):\n"
                for m_mkt in sorted(missing_markets_in_op)[:15]:
                    warn_msg += f"  - {m_mkt}\n"
                if len(missing_markets_in_op) > 15:
                    warn_msg += f"  ...and {len(missing_markets_in_op)-15} more.\n"
                warn_msg += "\n"
            warn_msg += "Please update your Operational Field Force Sheet to ensure consistency!\n\nDo you want to proceed with the system execution anyway?"
            
            log_message("\n" + "="*70)
            log_message("⚠️ WARNING: COMPANY MASTER vs OPERATIONAL SHEET DISCREPANCY")
            log_message(f"Missing FMs count: {len(missing_fms_in_op)} | Missing Markets count: {len(missing_markets_in_op)}")
            for m_fm in sorted(missing_fms_in_op):
                log_message(f"  [Missing FM] {m_fm}")
            for m_mkt in sorted(missing_markets_in_op):
                log_message(f"  [Missing Market] {m_mkt}")
            log_message("="*70 + "\n")
            
            return messagebox.askyesno("⚠️ Master Sheet Sync Warning", warn_msg, icon="warning")
        else:
            log_message("✔ Operational Field Force is 100% in sync with Company Master Sheet.")
            return True
    except Exception as e:
        log_message(f"Notice: Could not cross-check with Company Master Sheet ({e}). Proceeding...")
        return True

def get_email_mappings(gc):
    log_message("Fetching email mappings from Google Sheet...")
    sheet = gc.open_by_key(EMAIL_SHEET_ID)
    ws = None
    for w in sheet.worksheets():
        if "EMAIL" in w.title.upper():
            ws = w
            break
    if ws is None:
        ws = sheet.get_worksheet(0)
    rows = ws.get_all_values()
    
    headers = [h.strip().upper() for h in rows[0]] if rows else []
    def get_em_col(possible_names, fallback=-1):
        for n in possible_names:
            for idx, h in enumerate(headers):
                if n == h or n in h:
                    return idx
        return fallback

    fm_col = get_em_col(['FM/AM, ZONE', 'FM/AM', 'FM NAME'], 0)
    email_col = get_em_col(['EMAIL', 'MAIL ID', 'FM EMAIL'], 1)
    sh_email_col = get_em_col(['SH EMAIL', 'SH MAIL', 'RH EMAIL'], -1)
    boss_email_col = get_em_col(['SM/DSM MAIL', 'BOSS EMAIL', 'SM MAIL', 'DSM MAIL', 'BOSS MAIL'], -1)
    boss_name_col = get_em_col(['SM NAME', 'BOSS NAME', 'DSM NAME'], -1)
    zone_col = get_em_col(['ZONE', 'ZONE NAME'], -1)

    sh_by_zone = {}
    for r in rows[1:]:
        if zone_col != -1 and sh_email_col != -1 and len(r) > max(zone_col, sh_email_col):
            zone = str(r[zone_col]).strip().upper()
            sh_email = str(r[sh_email_col]).strip()
            if zone and sh_email and '@' in sh_email and zone not in sh_by_zone:
                sh_by_zone[zone] = sh_email

    mappings = {}
    for r in rows[1:]:
        if len(r) <= fm_col or not r[fm_col]:
            continue
        fm_full = str(r[fm_col]).strip()
        zone = str(r[zone_col]).strip().upper() if zone_col != -1 and len(r) > zone_col else ""
        fm_name_clean = clean_person_name(fm_full, zone)
        
        email = str(r[email_col]).strip() if len(r) > email_col else ""
        boss_email = str(r[boss_email_col]).strip() if boss_email_col != -1 and len(r) > boss_email_col else ""
        boss_name = str(r[boss_name_col]).strip() if boss_name_col != -1 and len(r) > boss_name_col else ""
        sh_email = sh_by_zone.get(zone, "")
        
        val = {
            'email': email,
            'boss_email': boss_email,
            'boss_name': boss_name,
            'sh_email': sh_email,
            'zone': zone
        }
        if zone:
            mappings[f"{zone}___{fm_name_clean}"] = val
        if fm_name_clean not in mappings or ('VACANT' not in fm_name_clean and fm_name_clean not in ['SH', 'ARM', 'RSM']):
            mappings[fm_name_clean] = val
    return mappings, sh_by_zone

def lookup_email_mapping(email_mappings, fm_name, zone=None, sh_by_zone=None):
    cn = clean_person_name(fm_name, zone)
    if zone:
        z = str(zone).strip().upper()
        res = email_mappings.get(f"{z}___{cn}")
        if res:
            return res
        if sh_by_zone and cn == 'SH' and z in sh_by_zone:
            return {'email': sh_by_zone[z], 'boss_email': '', 'boss_name': '', 'sh_email': sh_by_zone[z], 'zone': z}
        if 'VACANT' in cn or cn in ['SH', 'ARM', 'RSM']:
            return {'email': '', 'boss_email': '', 'boss_name': '', 'sh_email': sh_by_zone.get(z, '') if sh_by_zone else '', 'zone': z}
    return email_mappings.get(cn, {'email': '', 'boss_email': '', 'boss_name': '', 'sh_email': '', 'zone': ''})

# Drive API Helper
def get_or_create_drive_folder(drive_service, name, parent_id):
    query = f"name = '{name}' and mimeType = 'application/vnd.google-apps.folder' and '{parent_id}' in parents and trashed = false"
    response = drive_service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    files = response.get('files', [])
    if files:
        return files[0]['id']
    
    file_metadata = {
        'name': name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }
    folder = drive_service.files().create(body=file_metadata, fields='id').execute()
    log_message(f"Created Google Drive folder: {name} (ID: {folder['id']})")
    return folder['id']

def share_file(drive_service, file_id, email, role='writer'):
    if not email or '@' not in email:
        return
    try:
        user_permission = {
            'type': 'user',
            'role': role,
            'emailAddress': email
        }
        # Explicitly enable Google Drive email notifications to ensure recipients receive the sheet link
        drive_service.permissions().create(
            fileId=file_id,
            body=user_permission,
            sendNotificationEmail=True,
            fields='id'
        ).execute()
        log_message(f"Shared file {file_id} with {email} as {role} (email invite sent)")
    except Exception as e:
        log_message(f"Error sharing file {file_id} with {email}: {e}")

# Sheet Locking Helper (Protected Range)
def lock_sheet(sheets_service, spreadsheet_id, sheet_id, owner_email, message="Locked archive"):
    try:
        body = {
            "requests": [
                {
                    "addProtectedRange": {
                        "protectedRange": {
                            "range": {
                                "sheetId": sheet_id
                            },
                            "description": message,
                            "warningOnly": False,
                            "editors": {
                                "users": [owner_email]
                            }
                        }
                    }
                }
            ]
        }
        sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=body).execute()
        log_message(f"Successfully locked sheet ID: {sheet_id} in Spreadsheet: {spreadsheet_id} (Restricted to owner {owner_email})")
    except Exception as e:
        log_message(f"Error locking sheet ID {sheet_id}: {e}")

# Tab Sorting & Dynamic Read-Only Protection Engine
def sort_and_protect_spreadsheet_tabs(gc, sheets_service, sheet_id, editors_list=None, today_date=None):
    try:
        if today_date is None:
            today_date = get_dhaka_today()
        today_year = today_date.year
        today_month = today_date.month

        ss = gc.open_by_key(sheet_id)
        worksheets = ss.worksheets()
        
        month_tabs = []
        other_tabs = []
        for ws in worksheets:
            parsed = parse_month_str(ws.title)
            if parsed:
                month_tabs.append((parsed[0], parsed[1], ws))
            else:
                other_tabs.append(ws)
                
        # Sort month tabs descending by (year, month) so LATEST (newest/future) month comes FIRST (leftmost)!
        month_tabs.sort(key=lambda x: (x[0], x[1]), reverse=True)
        sorted_worksheets = [x[2] for x in month_tabs] + other_tabs

        requests = []
        for idx, ws in enumerate(sorted_worksheets):
            requests.append({
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": ws.id,
                        "index": idx
                    },
                    "fields": "index"
                }
            })

        meta = sheets_service.spreadsheets().get(
            spreadsheetId=sheet_id,
            fields="sheets(properties(sheetId,title),protectedRanges(protectedRangeId))"
        ).execute()

        for s_meta in meta.get('sheets', []):
            s_id = s_meta['properties']['sheetId']
            title = s_meta['properties']['title']
            
            for pr in s_meta.get('protectedRanges', []):
                requests.append({"deleteProtectedRange": {"protectedRangeId": pr['protectedRangeId']}})
                
            parsed = parse_month_str(title)
            if parsed:
                yr, mo = parsed
                is_old_month = (yr < today_year) or (yr == today_year and mo < today_month)
                if is_old_month:
                    # TOTAL READ ONLY -> ZERO UNPROTECTED RANGES
                    requests.append({
                        "addProtectedRange": {
                            "protectedRange": {
                                "range": {"sheetId": s_id},
                                "description": f"ARCHIVED TOTAL READ ONLY ({title})",
                                "warningOnly": False,
                                "unprotectedRanges": [],
                                "editors": {"users": []}
                            }
                        }
                    })
                else:
                    try:
                        ws_obj = ss.worksheet(title)
                        row_11 = [str(x).strip().upper() for x in ws_obj.row_values(11)]
                        try:
                            t_col_idx = row_11.index("TOTAL CASH IN HAND") + 1
                        except Exception:
                            t_col_idx = 11
                    except Exception:
                        t_col_idx = 11
                        
                    _, num_days = calendar.monthrange(yr, mo)
                    editors = [e for e in (editors_list or []) if e]
                    requests.append({
                        "addProtectedRange": {
                            "protectedRange": {
                                "range": {"sheetId": s_id},
                                "description": f"Locked headers, dates and formula columns ({title})",
                                "warningOnly": False,
                                "unprotectedRanges": [{
                                    "sheetId": s_id,
                                    "startRowIndex": 17,
                                    "endRowIndex": 17 + num_days,
                                    "startColumnIndex": 2,
                                    "endColumnIndex": t_col_idx - 1
                                }],
                                "editors": {"users": editors}
                            }
                        }
                    })

        if requests:
            sheets_service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body={"requests": requests}).execute()
        log_message(f"  ✔ Sorted tabs (latest first) & applied dynamic read-only protection for spreadsheet ID: {sheet_id}")
    except Exception as e:
        log_message(f"  Warning during tab sorting & locking for {sheet_id}: {e}")

# Existing Sheet Detection Engine
def check_existing_sheets_flow(gc, drive_service, selected_zones, target_month_str):
    log_message(f"Checking if sheets for {target_month_str} already exist in Google Drive...")
    try:
        registry_name = "Master_Registry_Cash_In_Hand"
        query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
        res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
        files = res.get('files', [])
        if not files:
            return False # No registry means no sheets exist yet
            
        reg_sheet = gc.open_by_key(files[0]['id'])
        reg_ws = reg_sheet.get_worksheet(0)
        reg_rows = reg_ws.get_all_values()
        if len(reg_rows) < 2:
            return False
            
        reg_cols = get_reg_col_indices(reg_rows[0])
        z_col, sid_col = reg_cols['zone'], reg_cols['sheet_id']
        
        # Check first matching FM sheet in selected zones
        for r in reg_rows[1:]:
            if len(r) > max(z_col, sid_col) and r[z_col] in selected_zones and r[sid_col]:
                sheet_id = r[sid_col]
                try:
                    ss = gc.open_by_key(sheet_id)
                    for ws in ss.worksheets():
                        if ws.title.upper() == target_month_str.upper():
                            log_message(f"Found existing tab '{target_month_str}' in sheet ID: {sheet_id}")
                            return True
                    break
                except Exception:
                    continue
    except Exception as e:
        log_message(f"Error checking existing sheets: {e}")
    return False

# ==========================================
# ADVANCED CLOUD SNAPSHOT & VERSION CONTROL ENGINE (WITH 6-MONTH AUTO-PURGE)
# ==========================================
ARCHIVE_FOLDER_NAME = "SYSTEM_VERSION_CONTROL_ARCHIVE (Auto-Purged after 6 Months)"
VERSION_LOG_NAME = "Master_Version_History_Log (Git_History)"

def get_or_create_archive_folder(drive_service):
    q = f"name = '{ARCHIVE_FOLDER_NAME}' and mimeType = 'application/vnd.google-apps.folder' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
    res = drive_service.files().list(q=q, spaces='drive', fields='files(id)').execute()
    files = res.get('files', [])
    if files:
        return files[0]['id']
    else:
        meta = {'name': ARCHIVE_FOLDER_NAME, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [PARENT_FOLDER_ID]}
        folder = drive_service.files().create(body=meta, fields='id').execute()
        log_message(f"Created dedicated Version Control Archive Folder ID: {folder['id']}")
        return folder['id']

def get_or_create_version_log(drive_service, gc, archive_folder_id):
    q = f"name = '{VERSION_LOG_NAME}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{archive_folder_id}' in parents and trashed = false"
    res = drive_service.files().list(q=q, spaces='drive', fields='files(id)').execute()
    files = res.get('files', [])
    if files:
        return gc.open_by_key(files[0]['id'])
    else:
        meta = {'name': VERSION_LOG_NAME, 'mimeType': 'application/vnd.google-apps.spreadsheet', 'parents': [archive_folder_id]}
        file_obj = drive_service.files().create(body=meta, fields='id, webViewLink').execute()
        ss = gc.open_by_key(file_obj['id'])
        ws = ss.get_worksheet(0)
        ws.update_title("Version_History_Log")
        headers = [["Timestamp", "Action Type", "Target File / Person", "Target Month", "Snapshot File Name", "Snapshot File ID", "Snapshot Link (Click to Restore)"]]
        ws.update(range_name='A1:G1', values=headers)
        ws.format('A1:G1', {'textFormat': {'bold': True}, 'backgroundColor': {'red': 0.1, 'green': 0.2, 'blue': 0.4}})
        log_message(f"Created Master Version History Log ID: {file_obj['id']}")
        return ss

def purge_old_snapshots(drive_service, gc, archive_folder_id):
    try:
        # Check files older than 180 days (6 months)
        six_months_ago = (datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=180)).strftime('%Y-%m-%dT%H:%M:%SZ')
        q = f"'{archive_folder_id}' in parents and createdTime < '{six_months_ago}' and trashed = false and name != '{VERSION_LOG_NAME}'"
        res = drive_service.files().list(q=q, spaces='drive', fields='files(id, name, createdTime)').execute()
        old_files = res.get('files', [])
        if old_files:
            log_message(f"Auto-Purge Engine: Found {len(old_files)} snapshot(s) older than 6 months -> Purging to save Drive storage...")
            for f in old_files:
                try:
                    drive_service.files().delete(fileId=f['id']).execute()
                    log_message(f"  ✔ Purged expired snapshot: {f['name']} (created: {f['createdTime'][:10]})")
                except Exception as ex:
                    log_message(f"  Error purging {f['name']}: {ex}")
    except Exception as e:
        log_message(f"Error during snapshot auto-purge: {e}")

def create_version_snapshot(drive_service, gc, file_id, file_name, action_type, target_month):
    try:
        log_message(f"🛡️ Taking Cloud Snapshot for version control before [{action_type}] on '{file_name}'...")
        archive_folder_id = get_or_create_archive_folder(drive_service)
        
        # Run 6-month auto-purge check silently
        purge_old_snapshots(drive_service, gc, archive_folder_id)
        
        timestamp_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        clean_ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        snapshot_name = f"[SNAPSHOT: {action_type} - {target_month}] {file_name} ({clean_ts})"
        
        copy_meta = {'name': snapshot_name, 'parents': [archive_folder_id]}
        copied_file = drive_service.files().copy(fileId=file_id, body=copy_meta, fields='id, webViewLink').execute()
        snap_id = copied_file['id']
        snap_link = copied_file.get('webViewLink', f"https://docs.google.com/spreadsheets/d/{snap_id}")
        
        # Record in log
        try:
            log_ss = get_or_create_version_log(drive_service, gc, archive_folder_id)
            log_ws = log_ss.get_worksheet(0)
            row_data = [timestamp_str, action_type, file_name, target_month, snapshot_name, snap_id, snap_link]
            log_ws.append_row(row_data)
        except Exception as log_err:
            log_message(f"  Warning: Could not update Version Log spreadsheet: {log_err}")
            
        log_message(f"  ✔ Snapshot archived successfully! ID: {snap_id}")
        return snap_id, snap_link
    except Exception as e:
        log_message(f"  Warning: Failed to create snapshot for {file_name}: {e}")
        return None, None

# Master Backup Copy Discovery & Auto-Generation Engine
def get_or_create_master_backups(drive_service, gc):
    log_message("Checking Google Drive for ONE Master Backup Copies...")
    mpo_backup_name = "Backup_Copy_DreamApps_MPO_FM_Codes"
    mail_backup_name = "Backup_Copy_Mail_Address_Master"
    
    # Check MPO Backup
    q_mpo = f"name = '{mpo_backup_name}' and '{PARENT_FOLDER_ID}' in parents and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
    res_mpo = drive_service.files().list(q=q_mpo, spaces='drive', fields='files(id, webViewLink)').execute()
    files_mpo = res_mpo.get('files', [])
    
    mpo_just_created = False
    if not files_mpo:
        log_message(f"First run detected: Generating ONE Master Backup Copy for Field Force -> '{mpo_backup_name}'...")
        copy_meta = {'name': mpo_backup_name, 'parents': [PARENT_FOLDER_ID]}
        mpo_file = drive_service.files().copy(fileId=DATA_SHEET_ID, body=copy_meta, fields='id, webViewLink').execute()
        mpo_backup_id = mpo_file['id']
        mpo_just_created = True
        log_message(f"Created Field Force Backup Copy ID: {mpo_backup_id}")
    else:
        mpo_backup_id = files_mpo[0]['id']
        log_message(f"Found existing Field Force Backup Copy ID: {mpo_backup_id}")

    # Check Mail Backup
    q_mail = f"name = '{mail_backup_name}' and '{PARENT_FOLDER_ID}' in parents and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
    res_mail = drive_service.files().list(q=q_mail, spaces='drive', fields='files(id, webViewLink)').execute()
    files_mail = res_mail.get('files', [])
    
    mail_just_created = False
    if not files_mail:
        log_message(f"First run detected: Generating ONE Master Backup Copy for Mail Address -> '{mail_backup_name}'...")
        copy_meta = {'name': mail_backup_name, 'parents': [PARENT_FOLDER_ID]}
        mail_file = drive_service.files().copy(fileId=EMAIL_SHEET_ID, body=copy_meta, fields='id, webViewLink').execute()
        mail_backup_id = mail_file['id']
        mail_just_created = True
        log_message(f"Created Mail Address Backup Copy ID: {mail_backup_id}")
    else:
        mail_backup_id = files_mail[0]['id']
        log_message(f"Found existing Mail Address Backup Copy ID: {mail_backup_id}")

    return {
        'mpo_id': mpo_backup_id,
        'mpo_just_created': mpo_just_created,
        'mail_id': mail_backup_id,
        'mail_just_created': mail_just_created
    }

# Email Modification Detection Engine
def check_for_changed_emails(gc, drive_service, selected_zones, backups_info=None):
    log_message("Checking for Email ID modifications...")
    if backups_info and backups_info.get('mail_just_created'):
        log_message("Mail Address Master Backup Copy was just generated on this run -> No email modification differences to report.")
        return [], None, []
    try:
        email_mappings, sh_by_zone = get_email_mappings(gc)
        registry_name = "Master_Registry_Cash_In_Hand"
        query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
        res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
        files = res.get('files', [])
        if not files:
            return [], None, []
            
        reg_sheet = gc.open_by_key(files[0]['id'])
        reg_ws = reg_sheet.get_worksheet(0)
        reg_rows = reg_ws.get_all_values()
        if len(reg_rows) < 2:
            return [], reg_ws, reg_rows

        reg_cols = get_reg_col_indices(reg_rows[0]) if reg_rows else {}
        fm_col, z_col, sid_col = reg_cols.get('fm', 0), reg_cols.get('zone', 1), reg_cols.get('sheet_id', 2)
        fme_col, bosse_col, she_col = reg_cols.get('fm_email', 4), reg_cols.get('boss_email', 6), reg_cols.get('sh_email', 7)

        changed_items = []
        for idx, r in enumerate(reg_rows[1:], start=2):
            if len(r) <= max(fm_col, z_col, sid_col):
                continue
            fm_name = r[fm_col]
            zone = r[z_col]
            sheet_id = r[sid_col]
            if zone not in selected_zones:
                continue
                
            old_fm_email = r[fme_col].strip() if len(r) > fme_col else ''
            old_boss_email = r[bosse_col].strip() if len(r) > bosse_col else ''
            old_sh_email = r[she_col].strip() if len(r) > she_col else ''
            
            mapping = lookup_email_mapping(email_mappings, fm_name, zone, sh_by_zone)
            new_fm_email = mapping['email'].strip()
            new_boss_email = mapping['boss_email'].strip()
            new_sh_email = mapping.get('sh_email', '').strip() or sh_by_zone.get(zone, '').strip()
            
            changes = []
            if new_fm_email and new_fm_email != old_fm_email:
                changes.append(('FM Email', old_fm_email, new_fm_email))
            if new_boss_email and new_boss_email != old_boss_email:
                changes.append(('Boss Email', old_boss_email, new_boss_email))
            if new_sh_email and new_sh_email != old_sh_email:
                changes.append(('SH Email', old_sh_email, new_sh_email))
                
            if changes:
                changed_items.append({
                    'row_idx': idx,
                    'fm_name': fm_name,
                    'zone': zone,
                    'sheet_id': sheet_id,
                    'changes': changes,
                    'new_fm': new_fm_email,
                    'new_boss': new_boss_email,
                    'new_sh': new_sh_email
                })
        return changed_items, reg_ws, reg_rows
    except Exception as e:
        log_message(f"Error checking email modifications: {e}")
        return [], None, []

def execute_email_permissions_update(drive_service, changed_items, reg_ws, reg_rows, dry_run=False):
    if not changed_items:
        return
    log_message("\n--- Executing Email Sharing Permissions & Registry Update ---")
    try:
        for item in changed_items:
            fm_name = item['fm_name']
            zone = item['zone']
            sheet_id = item['sheet_id']
            log_message(f"Updating sharing permissions for {fm_name} ({zone})...")
            
            if dry_run:
                log_message(f"  [DRY-RUN] Would share sheet {sheet_id} with new emails.")
                continue
                
            if item['new_fm']:
                share_file(drive_service, sheet_id, item['new_fm'], role='writer')
            if item['new_boss']:
                share_file(drive_service, sheet_id, item['new_boss'], role='reader')
            if item['new_sh']:
                share_file(drive_service, sheet_id, item['new_sh'], role='writer')
                
            row_idx = item['row_idx']
            log_message(f"  Updating Master Registry row {row_idx} with new emails...")
            try:
                reg_ws.update(f"E{row_idx}:H{row_idx}", [[item['new_fm'], "Done", item['new_boss'], item['new_sh']]])
            except Exception as ex:
                log_message(f"  Error updating registry row {row_idx}: {ex}")
                
        log_message("Email sharing permissions updated successfully!")
    except Exception as e:
        log_message(f"Error executing email updates: {e}")

# Missing Markets Detection Engine
def check_for_missing_markets(gc, drive_service, selected_zones, backups_info=None):
    log_message("Checking for Missing Markets / FMs...")
    if backups_info and backups_info.get('mpo_just_created'):
        log_message("Field Force Master Backup Copy was just generated on this run -> No cross-check differences to report.")
        try:
            valid_fms = fetch_master_data(gc)
            email_mappings, sh_by_zone = get_email_mappings(gc)
            return [], email_mappings, sh_by_zone, None
        except Exception:
            return [], {}, {}, None
    try:
        valid_fms = fetch_master_data(gc)
        email_mappings, sh_by_zone = get_email_mappings(gc)
        
        registry_name = "Master_Registry_Cash_In_Hand"
        query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
        res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
        files = res.get('files', [])
        
        existing_fm_names = set()
        reg_ws = None
        if files:
            reg_sheet = gc.open_by_key(files[0]['id'])
            reg_ws = reg_sheet.get_worksheet(0)
            reg_rows = reg_ws.get_all_values()
            reg_cols = get_reg_col_indices(reg_rows[0]) if reg_rows else {}
            fm_col = reg_cols.get('fm', 0)
            for r in reg_rows[1:]:
                if len(r) > fm_col and r[fm_col]:
                    existing_fm_names.add(r[fm_col].strip().upper())
                    
        missing_fms = []
        for fm_name, fm_data in valid_fms.items():
            if fm_data['zone'] in selected_zones:
                clean_n = fm_name.split(',')[0].strip().upper()
                if clean_n not in existing_fm_names:
                    missing_fms.append((clean_n, fm_data))
        return missing_fms, email_mappings, sh_by_zone, reg_ws
    except Exception as e:
        log_message(f"Error checking missing markets: {e}")
        return [], {}, {}, None

def execute_missing_markets_provisioning(gc, drive_service, sheets_service, missing_fms, email_mappings, sh_by_zone, reg_ws, month_str, num_days, dry_run=False):
    if not missing_fms:
        return
    log_message("\n--- Executing Missing Markets Provisioning ---")
    try:
        new_registry_entries = []
        for fm_clean_name, fm_data in missing_fms:
            zone = fm_data['zone']
            mapping = lookup_email_mapping(email_mappings, fm_clean_name, zone, sh_by_zone)
            fm_email = mapping['email']
            boss_email = mapping['boss_email']
            boss_name = mapping['boss_name']
            sh_email = mapping.get('sh_email', '') or sh_by_zone.get(zone, '')

            log_message(f"--- Provisioning missing market: {fm_clean_name} ({zone}) ---")
            local_path = create_local_excel(fm_clean_name, fm_data, month_str, num_days)
            
            if dry_run:
                log_message(f"  [DRY-RUN] Would upload and share missing sheet for {fm_clean_name}")
                continue

            zone_folder_id = get_or_create_drive_folder(drive_service, zone, PARENT_FOLDER_ID)
            file_title = f"CASH IN HAND - {fm_clean_name}"
            
            file_metadata = {
                'name': file_title,
                'mimeType': 'application/vnd.google-apps.spreadsheet',
                'parents': [zone_folder_id]
            }
            media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
            created_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
            sheet_id = created_file.get('id')
            sheet_url = created_file.get('webViewLink')
            
            share_file(drive_service, sheet_id, fm_email, role='writer')
            share_file(drive_service, sheet_id, boss_email, role='writer')
            share_file(drive_service, sheet_id, sh_email, role='reader')
            
            # Apply cell locking
            try:
                ss = gc.open_by_key(sheet_id)
                ws = ss.get_worksheet(0)
                try:
                    total_col_idx = [str(x).strip().upper() for x in ws.row_values(11)].index("TOTAL CASH IN HAND") + 1
                except Exception:
                    total_col_idx = 11
                
                editors_list = [e for e in [boss_email, sh_email] if e]
                reqs = [{
                    "setDataValidation": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 17,
                            "endRowIndex": 17 + num_days,
                            "startColumnIndex": 2,
                            "endColumnIndex": total_col_idx - 1
                        },
                        "rule": {
                            "condition": {
                                "type": "NUMBER_GREATER_THAN_EQ",
                                "values": [{"userEnteredValue": "0"}]
                            },
                            "inputMessage": "Enter a positive number or zero.",
                            "strict": True,
                            "showCustomUi": True
                        }
                    }
                }]
                sheets_service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body={"requests": reqs}).execute()
                sort_and_protect_spreadsheet_tabs(gc, sheets_service, sheet_id, editors_list)
            except Exception as e:
                log_message(f"Error locking cells for {fm_clean_name}: {e}")

            new_registry_entries.append([fm_clean_name, zone, sheet_id, sheet_url, fm_email, boss_name, boss_email, sh_email])

        if new_registry_entries and reg_ws and not dry_run:
            log_message("Appending new markets to Master Registry...")
            reg_ws.append_rows(new_registry_entries, value_input_option='USER_ENTERED')
            
            # Re-update Boss Summary sheets for the affected bosses to dynamically insert new columns!
            log_message("Updating Boss Summary Sheets with newly added markets...")
            reg_rows = reg_ws.get_all_values()
            headers = reg_rows[0]
            registry_records = []
            for r in reg_rows[1:]:
                rec = {}
                for idx, h in enumerate(headers):
                    if idx < len(r):
                        rec[h] = r[idx]
                registry_records.append(rec)
            update_boss_summary_sheets(drive_service, gc, sheets_service, registry_records, month_str, num_days, dry_run=False)

        log_message("Missing markets provisioning completed!")
    except Exception as e:
        log_message(f"Error executing missing markets: {e}")

# Formatting & Creation of Local Workbook
def create_local_excel(fm_name, fm_data, month_str, num_days):
    FONT_FAMILY   = 'Aptos'
    C_VOID        = '060816'
    C_DEEP_NAVY   = '0D1425'
    C_MIDNIGHT    = '1E293B'
    C_DARK_SURF   = '0F172A'
    C_ZEBRA_A     = 'FFFFFF'
    C_ZEBRA_B     = 'F8FAFC'
    C_TOTAL_DATA  = 'ECFDF5'
    C_TOTAL_HEAD  = '065F46'
    T_NEON        = '00F2FE'
    T_WHITE       = 'FFFFFF'
    T_SLATE       = 'CBD5E1'
    T_INK         = '0F172A'
    T_MINT        = 'A7F3D0'
    T_TOTAL_DARK  = '064E3B'
    T_DATE        = '475569'
    B_NEON        = '0E7490'
    B_SLATE       = '334155'
    B_LIGHT       = 'E2E8F0'
    B_TOTAL       = '6EE7B7'

    font_title    = Font(name=FONT_FAMILY, size=18, bold=True,  color=T_NEON)
    font_hdr      = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
    font_sub      = Font(name=FONT_FAMILY, size=9,  bold=True,  color=T_SLATE)
    font_name     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
    font_date     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_DATE)
    font_body     = Font(name=FONT_FAMILY, size=10, bold=False, color=T_INK)
    font_total_h  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_MINT)
    font_total_d  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_TOTAL_DARK)

    fill_void     = PatternFill('solid', start_color=C_VOID,       end_color=C_VOID)
    fill_navy     = PatternFill('solid', start_color=C_DEEP_NAVY,  end_color=C_DEEP_NAVY)
    fill_mid      = PatternFill('solid', start_color=C_MIDNIGHT,   end_color=C_MIDNIGHT)
    fill_dark     = PatternFill('solid', start_color=C_DARK_SURF,  end_color=C_DARK_SURF)
    fill_za       = PatternFill('solid', start_color=C_ZEBRA_A,    end_color=C_ZEBRA_A)
    fill_zb       = PatternFill('solid', start_color=C_ZEBRA_B,    end_color=C_ZEBRA_B)
    fill_tot_data = PatternFill('solid', start_color=C_TOTAL_DATA, end_color=C_TOTAL_DATA)
    fill_tot_head = PatternFill('solid', start_color=C_TOTAL_HEAD, end_color=C_TOTAL_HEAD)

    def side(style='thin', color=B_LIGHT):
        return Side(style=style, color=color)

    bd_data  = Border(left=side(), right=side(), top=side(), bottom=side())
    bd_hdr   = Border(left=side('thin',B_SLATE), right=side('thin',B_SLATE), top=side('thin',B_SLATE), bottom=side('thin',B_SLATE))
    bd_total = Border(left=side('thin',B_TOTAL), right=side('thin',B_TOTAL), top=side('thin',B_TOTAL), bottom=side('thin',B_TOTAL))
    bd_date  = Border(left=side('medium',B_SLATE), right=side('thin',B_LIGHT), top=side('thin',B_LIGHT), bottom=side('thin',B_LIGHT))

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right  = Alignment(horizontal='right', vertical='center')

    dates = [f"{d} {month_str}" for d in range(num_days, 0, -1)]

    zone = fm_data['zone']
    markets = fm_data['markets']
    das = []
    for m in markets:
        if m['da_name'] and str(m['da_name']).strip().upper() != 'VACANT':
            da_str = str(m['da_name']).strip().upper()
            if da_str not in das:
                das.append(da_str)
                
    num_mpos = len(markets)
    num_das = len(das)
    total_cols = 1 + 1 + num_mpos + num_das + 1
    total_col = 4 + num_mpos + num_das
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month_str
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_view.zoomScale = 90
    
    if zone == 'CTG.A':
        ws.sheet_properties.tabColor = '00F2FE'
    else:
        ws.sheet_properties.tabColor = 'A855F7'
        
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[11].height = 26
    for r in range(18, 18 + len(dates)):
        ws.row_dimensions[r].height = 20
        
    ws.freeze_panes = 'D18'
    
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=total_cols+1)
    title_cell = ws.cell(row=2, column=2, value="CASH IN HAND")
    title_cell.font = font_title
    title_cell.fill = fill_void
    title_cell.alignment = align_center
    
    for r in range(2, 4):
        for c in range(2, total_cols + 2):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_void
            left = side('medium', B_NEON) if c == 2 else None
            right = side('medium', B_NEON) if c == total_cols + 1 else None
            top = side('medium', B_NEON) if r == 2 else None
            bottom = side('medium', B_NEON) if r == 3 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            
    mpo_start = 4
    mpo_end = 3 + num_mpos
    da_start = 4 + num_mpos
    da_end = 3 + num_mpos + num_das
    
    ws.merge_cells(start_row=5, start_column=mpo_start, end_row=5, end_column=mpo_end)
    c_mpo = ws.cell(row=5, column=mpo_start, value="MPO")
    c_mpo.font = font_hdr
    c_mpo.fill = fill_mid
    c_mpo.alignment = align_center
    
    if num_das > 0:
        ws.merge_cells(start_row=5, start_column=da_start, end_row=5, end_column=da_end)
        c_da = ws.cell(row=5, column=da_start, value="DA")
        c_da.font = font_hdr
        c_da.fill = fill_mid
        c_da.alignment = align_center
        
    ws.merge_cells(start_row=5, start_column=total_col, end_row=11, end_column=total_col)
    c_tot_header = ws.cell(row=5, column=total_col, value="TOTAL CASH IN HAND")
    c_tot_header.font = font_total_h
    c_tot_header.fill = fill_tot_head
    c_tot_header.alignment = align_center
    
    for r in range(5, 12):
        ws.cell(row=r, column=total_col).fill = fill_tot_head
        ws.cell(row=r, column=total_col).border = bd_total
    for c in range(mpo_start, mpo_end + 1):
        ws.cell(row=5, column=c).fill = fill_mid
        ws.cell(row=5, column=c).border = bd_hdr
    if num_das > 0:
        for c in range(da_start, da_end + 1):
            ws.cell(row=5, column=c).fill = fill_mid
            ws.cell(row=5, column=c).border = bd_hdr
            
    for m_idx, m in enumerate(markets):
        col = 4 + m_idx
        ws.cell(row=6, column=col, value=zone).font = font_sub
        ws.cell(row=7, column=col, value=fm_name).font = font_sub
        ws.cell(row=8, column=col, value=m['market_name']).font = font_sub
        ws.cell(row=9, column=col, value=m['mpo_code']).font = font_sub
        ws.cell(row=10, column=col, value=m['fm_code']).font = font_sub
        
        c_name = ws.cell(row=11, column=col, value=m['mpo_name'])
        c_name.font = font_name
        c_name.fill = fill_navy
        c_name.alignment = align_center
        c_name.border = bd_hdr
        
        for r in range(6, 11):
            ws.cell(row=r, column=col).fill = fill_navy
            ws.cell(row=r, column=col).border = bd_hdr
            
    for d_idx, da_val in enumerate(das):
        col = 4 + num_mpos + d_idx
        ws.cell(row=6, column=col, value=zone).font = font_sub
        ws.cell(row=7, column=col, value=fm_name).font = font_sub
        ws.cell(row=8, column=col, value="").font = font_sub
        ws.cell(row=9, column=col, value="").font = font_sub
        ws.cell(row=10, column=col, value="").font = font_sub
        
        c_name = ws.cell(row=11, column=col, value=da_val)
        c_name.font = font_name
        c_name.fill = fill_navy
        c_name.alignment = align_center
        c_name.border = bd_hdr
        
        for r in range(6, 11):
            ws.cell(row=r, column=col).fill = fill_navy
            ws.cell(row=r, column=col).border = bd_hdr
            
    # Row 12: Blank separator
    for c in range(2, total_cols + 2):
        ws.cell(row=12, column=c).fill = fill_mid
        ws.cell(row=12, column=c).border = bd_hdr
        
    # Rows 13-17: DA metadata
    for m_idx, m in enumerate(markets):
        col = 4 + m_idx
        ws.cell(row=13, column=col, value=zone).font = font_sub
        ws.cell(row=14, column=col, value=fm_name).font = font_sub
        ws.cell(row=15, column=col, value=m['desig']).font = font_sub
        ws.cell(row=16, column=col, value=m['mpo_code']).font = font_sub
        ws.cell(row=17, column=col, value=m['da_name']).font = font_sub
        
        for r in range(13, 18):
            ws.cell(row=r, column=col).fill = fill_navy
            ws.cell(row=r, column=col).border = bd_hdr
            
    for d_idx, da_val in enumerate(das):
        col = 4 + num_mpos + d_idx
        ws.cell(row=13, column=col, value=zone).font = font_sub
        ws.cell(row=14, column=col, value=fm_name).font = font_sub
        ws.cell(row=15, column=col, value="DA").font = font_sub
        ws.cell(row=16, column=col, value="").font = font_sub
        ws.cell(row=17, column=col, value=da_val).font = font_sub
        
        for r in range(13, 18):
            ws.cell(row=r, column=col).fill = fill_navy
            ws.cell(row=r, column=col).border = bd_hdr
            
    # Date Label Header in Column B (Row 11 to 17)
    ws.merge_cells(start_row=11, start_column=2, end_row=17, end_column=2)
    c_date_h = ws.cell(row=11, column=2, value="DATE")
    c_date_h.font = font_date
    c_date_h.alignment = align_center
    for r in range(11, 18):
        ws.cell(row=r, column=2).fill = fill_za
        ws.cell(row=r, column=2).border = bd_hdr

    # Column C: FM SELF Header & Metadata (Rows 5 to 17)
    ws.cell(row=5, column=3).fill = fill_mid
    ws.cell(row=5, column=3).border = bd_hdr

    ws.cell(row=6, column=3, value=zone).font = font_sub
    ws.cell(row=7, column=3, value=fm_name).font = font_sub
    for r in range(6, 11):
        ws.cell(row=r, column=3).fill = fill_navy
        ws.cell(row=r, column=3).border = bd_hdr
        ws.cell(row=r, column=3).alignment = align_center

    c_fmself = ws.cell(row=11, column=3, value="FM SELF")
    c_fmself.font = font_name
    c_fmself.fill = fill_navy
    c_fmself.alignment = align_center
    c_fmself.border = bd_hdr

    ws.cell(row=12, column=3).fill = fill_mid
    ws.cell(row=12, column=3).border = bd_hdr

    ws.cell(row=13, column=3, value=zone).font = font_sub
    ws.cell(row=14, column=3, value=fm_name).font = font_sub
    ws.cell(row=15, column=3, value="FM").font = font_sub
    ws.cell(row=16, column=3, value="").font = font_sub
    ws.cell(row=17, column=3, value="SELF").font = font_sub
    for r in range(13, 18):
        ws.cell(row=r, column=3).fill = fill_navy
        ws.cell(row=r, column=3).border = bd_hdr
        ws.cell(row=r, column=3).alignment = align_center
    ws.column_dimensions['C'].width = 14
    
    # Fill actual date values and formula cells
    for idx, d_val in enumerate(dates):
        r = 18 + idx
        fill_row = fill_zb if idx % 2 == 1 else fill_za
        
        # Column B: Date
        c_date = ws.cell(row=r, column=2, value=d_val)
        c_date.font = font_body
        c_date.alignment = align_center
        c_date.fill = fill_row
        c_date.border = bd_date
        
        # Column C: FM Self
        c_self = ws.cell(row=r, column=3)
        c_self.font = font_body
        c_self.alignment = align_right
        c_self.fill = fill_row
        c_self.border = bd_data
        c_self.number_format = '#,##0'
        
        # Columns D to DAs: data entry fields
        for col_idx in range(4, total_col):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = font_body
            cell.alignment = align_right
            cell.fill = fill_row
            cell.border = bd_data
            cell.number_format = '#,##0'
            
        # Column Total: Formula
        last_input_col_letter = get_column_letter(total_col - 1)
        tot_formula = f"=SUM(C{r}:{last_input_col_letter}{r})"
        c_tot = ws.cell(row=r, column=total_col, value=tot_formula)
        c_tot.font = font_total_d
        c_tot.alignment = align_right
        c_tot.fill = fill_tot_data
        c_tot.border = bd_total
        c_tot.number_format = '#,##0'

    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    for col_idx in range(4, total_col):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions[get_column_letter(total_col)].width = 18
    
    # Hide vacant columns
    for col_idx in range(4, 4 + num_mpos):
        m = markets[col_idx - 4]
        if m['is_vacant']:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True
            
    # Hide metadata rows
    for r in range(6, 11):
        ws.row_dimensions[r].hidden = True
    for r in range(12, 18):
        ws.row_dimensions[r].hidden = True

    # Output path
    os.makedirs(os.path.join(BASE_DIR, zone), exist_ok=True)
    out_path = os.path.join(BASE_DIR, zone, f"{fm_name}.xlsx")
    wb.save(out_path)
    wb.close()
    return out_path

# Local Zonal Summary Sheet generator (similar to generate_zonal_summary_drive)
def create_local_zonal_excel(zone, zone_fms, registry_map, output_path, month_str, num_days):
    FONT_FAMILY   = 'Aptos'
    C_VOID        = '060816'
    C_DEEP_NAVY   = '0D1425'
    C_MIDNIGHT    = '1E293B'
    C_DARK_SURF   = '0F172A'
    C_ZEBRA_A     = 'FFFFFF'
    C_ZEBRA_B     = 'F8FAFC'
    C_TOTAL_DATA  = 'ECFDF5'
    C_TOTAL_HEAD  = '065F46'
    T_NEON        = '00F2FE'
    T_WHITE       = 'FFFFFF'
    T_SLATE       = 'CBD5E1'
    T_INK         = '0F172A'
    T_MINT        = 'A7F3D0'
    T_TOTAL_DARK  = '064E3B'
    T_DATE        = '475569'
    B_NEON        = '0E7490'
    B_SLATE       = '334155'
    B_LIGHT       = 'E2E8F0'
    B_TOTAL       = '6EE7B7'

    font_title    = Font(name=FONT_FAMILY, size=18, bold=True,  color=T_NEON)
    font_hdr      = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
    font_sub      = Font(name=FONT_FAMILY, size=9,  bold=True,  color=T_SLATE)
    font_name     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_WHITE)
    font_date     = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_DATE)
    font_body     = Font(name=FONT_FAMILY, size=10, bold=False, color=T_INK)
    font_total_h  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_MINT)
    font_total_d  = Font(name=FONT_FAMILY, size=10, bold=True,  color=T_TOTAL_DARK)

    fill_void     = PatternFill('solid', start_color=C_VOID,       end_color=C_VOID)
    fill_navy     = PatternFill('solid', start_color=C_DEEP_NAVY,  end_color=C_DEEP_NAVY)
    fill_mid      = PatternFill('solid', start_color=C_MIDNIGHT,   end_color=C_MIDNIGHT)
    fill_dark     = PatternFill('solid', start_color=C_DARK_SURF,  end_color=C_DARK_SURF)
    fill_za       = PatternFill('solid', start_color=C_ZEBRA_A,    end_color=C_ZEBRA_A)
    fill_zb       = PatternFill('solid', start_color=C_ZEBRA_B,    end_color=C_ZEBRA_B)
    fill_tot_data = PatternFill('solid', start_color=C_TOTAL_DATA, end_color=C_TOTAL_DATA)
    fill_tot_head = PatternFill('solid', start_color=C_TOTAL_HEAD, end_color=C_TOTAL_HEAD)

    def side(style='thin', color=B_LIGHT):
        return Side(style=style, color=color)

    bd_data  = Border(left=side(), right=side(), top=side(), bottom=side())
    bd_hdr   = Border(left=side('thin',B_SLATE), right=side('thin',B_SLATE), top=side('thin',B_SLATE), bottom=side('thin',B_SLATE))
    bd_total = Border(left=side('thin',B_TOTAL), right=side('thin',B_TOTAL), top=side('thin',B_TOTAL), bottom=side('thin',B_TOTAL))
    bd_date  = Border(left=side('medium',B_SLATE), right=side('thin',B_LIGHT), top=side('thin',B_LIGHT), bottom=side('thin',B_LIGHT))

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right  = Alignment(horizontal='right', vertical='center')

    dates = [f"{d} {month_str}" for d in range(num_days, 0, -1)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month_str
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_view.zoomScale = 90
    
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[11].height = 26
    for r in range(18, 18 + len(dates)):
        ws.row_dimensions[r].height = 20
        
    ws.freeze_panes = 'E18'
    ws.sheet_properties.tabColor = '00F2FE' if zone == 'CTG.A' else 'A855F7'
    
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 13
    
    num_fms = len(zone_fms)
    zone_summary_col = 3
    sh_self_col = 4
    summary_start_col = 5
    summary_end_col = 4 + num_fms
    
    # Summary header
    ws.merge_cells(start_row=5, start_column=summary_start_col, end_row=5, end_column=summary_end_col)
    c_sum_hdr = ws.cell(row=5, column=summary_start_col, value=f"FM WISE TOTAL CASH IN HAND, {zone}")
    c_sum_hdr.font = font_total_h
    c_sum_hdr.fill = fill_tot_head
    c_sum_hdr.alignment = align_center
    
    for col_idx in range(summary_start_col, summary_end_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.cell(row=5, column=col_idx).fill = fill_tot_head
        ws.cell(row=5, column=col_idx).border = bd_total
        
    ws.merge_cells(start_row=5, start_column=zone_summary_col, end_row=11, end_column=zone_summary_col)
    c_zone_hdr = ws.cell(row=5, column=zone_summary_col, value=f"ZONE SUMMARY\n{zone}")
    c_zone_hdr.font = font_total_h
    c_zone_hdr.fill = fill_tot_head
    c_zone_hdr.alignment = align_center
    for r in range(5, 12):
        ws.cell(row=r, column=zone_summary_col).fill = fill_tot_head
        ws.cell(row=r, column=zone_summary_col).border = bd_total
        
    ws.merge_cells(start_row=5, start_column=sh_self_col, end_row=11, end_column=sh_self_col)
    c_sh_hdr = ws.cell(row=5, column=sh_self_col, value="SH, SELF")
    c_sh_hdr.font = font_total_h
    c_sh_hdr.fill = fill_tot_head
    c_sh_hdr.alignment = align_center
    for r in range(5, 12):
        ws.cell(row=r, column=sh_self_col).fill = fill_tot_head
        ws.cell(row=r, column=sh_self_col).border = bd_total
        
    for f_idx, (f_name, f_data) in enumerate(zone_fms.items()):
        col_idx = summary_start_col + f_idx
        fm_clean_name = f_name.split(',')[0].strip()
        
        ws.merge_cells(start_row=6, start_column=col_idx, end_row=11, end_column=col_idx)
        c_fm_summary_hdr = ws.cell(row=6, column=col_idx, value=fm_clean_name)
        c_fm_summary_hdr.font = font_total_h
        c_fm_summary_hdr.fill = fill_tot_head
        c_fm_summary_hdr.alignment = align_center
        for r in range(6, 12):
            ws.cell(row=r, column=col_idx).fill = fill_tot_head
            ws.cell(row=r, column=col_idx).border = bd_total
            
    start_data_col = summary_end_col + 1
    fm_col_ranges = []
    curr_col = start_data_col
    
    for f_name, f_data in zone_fms.items():
        fm_clean_name = f_name.split(',')[0].strip()
        markets = f_data['markets']
        
        das = []
        for m in markets:
            if m['da_name'] and str(m['da_name']).strip().upper() != 'VACANT':
                da_str = str(m['da_name']).strip().upper()
                if da_str not in das:
                    das.append(da_str)
                    
        num_mpos = len(markets)
        num_das = len(das)
        
        fm_start = curr_col
        fm_end = curr_col + 1 + num_mpos + num_das - 1
        fm_col_ranges.append((fm_start, fm_end, fm_clean_name))
        
        # FM SELF Column
        ws.merge_cells(start_row=11, start_column=curr_col, end_row=14, end_column=curr_col)
        c_fm_self_h = ws.cell(row=11, column=curr_col, value="FM SELF")
        c_fm_self_h.font = font_hdr
        c_fm_self_h.fill = fill_mid
        c_fm_self_h.alignment = align_center
        for r in range(11, 15):
            ws.cell(row=r, column=curr_col).fill = fill_mid
            ws.cell(row=r, column=curr_col).border = bd_hdr
            
        ws.cell(row=6, column=curr_col, value=zone).font = font_sub
        ws.cell(row=7, column=curr_col, value=fm_clean_name).font = font_sub
        for r in range(6, 11):
            ws.cell(row=r, column=curr_col).fill = fill_navy
            ws.cell(row=r, column=curr_col).border = bd_hdr
            ws.cell(row=r, column=curr_col).alignment = align_center
        ws.column_dimensions[get_column_letter(curr_col)].width = 13
        
        # MPOs Columns
        mpo_start = curr_col + 1
        mpo_end = curr_col + num_mpos
        ws.merge_cells(start_row=5, start_column=mpo_start, end_row=5, end_column=mpo_end)
        c_mpo_h = ws.cell(row=5, column=mpo_start, value="MPO")
        c_mpo_h.font = font_hdr
        c_mpo_h.fill = fill_mid
        c_mpo_h.alignment = align_center
        
        for c in range(mpo_start, mpo_end + 1):
            ws.cell(row=5, column=c).fill = fill_mid
            ws.cell(row=5, column=c).border = bd_hdr
            
        for m_idx, m in enumerate(markets):
            c_idx = mpo_start + m_idx
            ws.cell(row=6, column=c_idx, value=zone).font = font_sub
            ws.cell(row=7, column=c_idx, value=fm_clean_name).font = font_sub
            ws.cell(row=8, column=c_idx, value=m['market_name']).font = font_sub
            ws.cell(row=9, column=c_idx, value=m['mpo_code']).font = font_sub
            ws.cell(row=10, column=c_idx, value=m['fm_code']).font = font_sub
            
            for r in range(6, 11):
                ws.cell(row=r, column=c_idx).fill = fill_navy
                ws.cell(row=r, column=c_idx).border = bd_hdr
                ws.cell(row=r, column=c_idx).alignment = align_center
                
            ws.cell(row=11, column=c_idx, value=m['mpo_name']).font = font_name
            ws.cell(row=11, column=c_idx).fill = fill_dark
            ws.cell(row=11, column=c_idx).border = bd_hdr
            ws.cell(row=11, column=c_idx).alignment = align_center
            
            ws.cell(row=12, column=c_idx, value=m['desig']).font = font_sub
            ws.cell(row=13, column=c_idx, value=m['is_vacant']).font = font_sub
            ws.cell(row=14, column=c_idx, value=m['da_name']).font = font_sub
            for r in range(12, 15):
                ws.cell(row=r, column=c_idx).fill = fill_navy
                ws.cell(row=r, column=c_idx).border = bd_hdr
                ws.cell(row=r, column=c_idx).alignment = align_center
                
            name_len = len(m['mpo_name'] or '')
            ws.column_dimensions[get_column_letter(c_idx)].width = max(name_len * 1.1, 13)
            
            if m['is_vacant'] == 'Y':
                ws.column_dimensions[get_column_letter(c_idx)].hidden = True
                
        # DAs Columns
        if num_das > 0:
            da_start = curr_col + 1 + num_mpos
            da_end = curr_col + num_mpos + num_das
            ws.merge_cells(start_row=5, start_column=da_start, end_row=5, end_column=da_end)
            c_da_h = ws.cell(row=5, column=da_start, value="DA")
            c_da_h.font = font_hdr
            c_da_h.fill = fill_mid
            c_da_h.alignment = align_center
            
            for c in range(da_start, da_end + 1):
                ws.cell(row=5, column=c).fill = fill_mid
                ws.cell(row=5, column=c).border = bd_hdr
                
            for d_idx, da_name in enumerate(das):
                c_idx = da_start + d_idx
                ws.cell(row=11, column=c_idx, value=da_name).font = font_name
                ws.cell(row=11, column=c_idx).fill = fill_dark
                ws.cell(row=11, column=c_idx).border = bd_hdr
                ws.cell(row=11, column=c_idx).alignment = align_center
                
                for r in range(5, 11):
                    if r >= 6:
                        ws.cell(row=r, column=c_idx).fill = fill_navy
                        ws.cell(row=r, column=c_idx).border = bd_hdr
                for r in range(12, 15):
                    ws.cell(row=r, column=c_idx).fill = fill_navy
                    ws.cell(row=r, column=c_idx).border = bd_hdr
                    
                name_len = len(da_name or '')
                ws.column_dimensions[get_column_letter(c_idx)].width = max(name_len * 1.1, 13)
                
        curr_col = fm_end + 1
        
    # DATE Header
    ws.merge_cells(start_row=11, start_column=2, end_row=14, end_column=2)
    c_date_h = ws.cell(row=11, column=2, value="DATE")
    c_date_h.font = font_hdr
    c_date_h.fill = fill_mid
    c_date_h.alignment = align_center
    for r in range(11, 15):
        ws.cell(row=r, column=2).fill = fill_mid
        ws.cell(row=r, column=2).border = bd_hdr
        
    last_col_idx = curr_col - 1
    
    # Title Block
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=last_col_idx)
    title_cell = ws.cell(row=2, column=2, value="CASH IN HAND")
    title_cell.font = font_title
    title_cell.fill = fill_void
    title_cell.alignment = align_center
    for r in range(2, 4):
        for c in range(2, last_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_void
            left = side('medium', B_NEON) if c == 2 else None
            right = side('medium', B_NEON) if c == last_col_idx else None
            top = side('medium', B_NEON) if r == 2 else None
            bottom = side('medium', B_NEON) if r == 3 else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            
    # Rows 15-17 summary formulas
    for r in range(15, 18):
        for c in range(2, last_col_idx + 1):
            ws.cell(row=r, column=c).border = bd_data
            
        ws.cell(row=r, column=zone_summary_col, value=f"=SUM({get_column_letter(sh_self_col)}{r}:{get_column_letter(summary_end_col)}{r})").font = font_total_d
        ws.cell(row=r, column=zone_summary_col).alignment = align_center
        ws.cell(row=r, column=zone_summary_col).fill = fill_tot_data
        ws.cell(row=r, column=zone_summary_col).border = bd_total
        
        ws.cell(row=r, column=sh_self_col, value=0).font = font_total_d
        ws.cell(row=r, column=sh_self_col).alignment = align_center
        ws.cell(row=r, column=sh_self_col).fill = fill_tot_data
        ws.cell(row=r, column=sh_self_col).border = bd_total
        
        for f_idx, (f_start, f_end, _) in enumerate(fm_col_ranges):
            col_idx = summary_start_col + f_idx
            ws.cell(row=r, column=col_idx, value=f"=SUM({get_column_letter(f_start)}{r}:{get_column_letter(f_end)}{r})").font = font_total_d
            ws.cell(row=r, column=col_idx).alignment = align_center
            ws.cell(row=r, column=col_idx).fill = fill_tot_data
            ws.cell(row=r, column=col_idx).border = bd_total
            
    # Rows 18-48 data entry & importrange
    for r_idx, date_val in enumerate(dates):
        row_num = 18 + r_idx
        row_fill = fill_za if (r_idx % 2 == 0) else fill_zb
        
        c_val = ws.cell(row=row_num, column=2, value=date_val)
        c_val.font = font_date
        c_val.alignment = align_center
        c_val.border = bd_date
        c_val.fill = row_fill
        
        zs_cell = ws.cell(row=row_num, column=zone_summary_col, value=f"=SUM({get_column_letter(sh_self_col)}{row_num}:{get_column_letter(summary_end_col)}{row_num})")
        zs_cell.font = font_total_d
        zs_cell.alignment = align_right
        zs_cell.border = bd_total
        zs_cell.fill = fill_tot_data
        zs_cell.number_format = '#,##0'
        
        sh_cell = ws.cell(row=row_num, column=sh_self_col, value=0)
        sh_cell.font = font_total_d
        sh_cell.alignment = align_right
        sh_cell.border = bd_total
        sh_cell.fill = fill_tot_data
        sh_cell.number_format = '#,##0'
        
        for f_idx, (f_start, f_end, _) in enumerate(fm_col_ranges):
            col_idx = summary_start_col + f_idx
            sum_cell = ws.cell(row=row_num, column=col_idx, value=f"=SUM({get_column_letter(f_start)}{row_num}:{get_column_letter(f_end)}{row_num})")
            sum_cell.font = font_total_d
            sum_cell.alignment = align_right
            sum_cell.border = bd_total
            sum_cell.fill = fill_tot_data
            sum_cell.number_format = '#,##0'
            
        for c in range(start_data_col, last_col_idx + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.border = bd_data
            cell.alignment = align_right
            cell.number_format = '#,##0'
            cell.font = font_body
            cell.fill = row_fill
            
    # Set IMPORTRANGE formula for each FM starting row 18
    for f_start, f_end, fm_clean_name in fm_col_ranges:
        fm_url = registry_map.get(fm_clean_name)
        if fm_url:
            col_count = f_end - f_start + 1
            fm_last_col_letter = get_column_letter(3 + col_count - 1)
            importrange_formula = f'=IMPORTRANGE("{fm_url}", "{month_str}!C18:{fm_last_col_letter}{18+num_days-1}")'
            ws.cell(row=18, column=f_start, value=importrange_formula)

    # Hide metadata rows
    for r in [6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17]:
        ws.row_dimensions[r].hidden = True
        
    ws.freeze_panes = 'C18'
    wb.save(output_path)
    wb.close()

# Zonal Summary Creation & Drive Uplink Orchestrator
def run_zonal_summaries(selected_zones, month_str, num_days, dry_run=False):
    log_message("\nStarting Zonal Summary Generation process...")
    creds = get_oauth_credentials()
    gc = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    sheets_service = build('sheets', 'v4', credentials=creds)

    valid_fms = fetch_master_data(gc)
    email_mappings, sh_by_zone = get_email_mappings(gc)

    # Load Registry to find uploaded sheets URL map
    registry_name = "Master_Registry_Cash_In_Hand"
    query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
    res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
    files = res.get('files', [])
    if not files:
        log_message("Warning: Master Registry not found! Cannot link worksheets.")
        return
        
    reg_sheet = gc.open_by_key(files[0]['id'])
    reg_ws = reg_sheet.get_worksheet(0)
    registry_records = reg_ws.get_all_records()
    
    registry_map = {}
    for r in registry_records:
        fm = r.get('FM Name')
        url = r.get('URL')
        if fm and url:
            registry_map[clean_person_name(fm)] = url

    # Group valid FMs by Zone
    zone_groups = {}
    for fm_name, fm_data in valid_fms.items():
        z = fm_data['zone']
        if z in selected_zones:
            if z not in zone_groups:
                zone_groups[z] = {}
            zone_groups[z][fm_name] = fm_data

    for zone, zone_fms in zone_groups.items():
        log_message(f"\n--- Processing Zonal Summary: {zone} ---")

        # Sort FMs to match exact sample layouts
        def fm_sort_key(item):
            f_name = item[0]
            fm_clean = f_name.split(',')[0].strip().upper()
            if zone == 'CTG.B':
                order = ["JAINAL ABEDIN AKHAND", "FIROZ AHMED", "NARAYAN DAS", "SHAHJAHAN"]
            else:
                order = ["MONIR UDDIN", "RAFIQUL MOULA", "VACANT, KHAGRACHARI"]
            for idx, name in enumerate(order):
                if name in fm_clean:
                    return idx
            return 99

        sorted_zone_fms = dict(sorted(zone_fms.items(), key=fm_sort_key))

        local_summary_path = os.path.join(BASE_DIR, f"{zone}_CASH_IN_HAND_Summary_temp.xlsx")
        create_local_zonal_excel(zone, sorted_zone_fms, registry_map, local_summary_path, month_str, num_days)
        log_message(f"Generated local summary workbook: {local_summary_path}")

        if dry_run:
            log_message(f"[DRY-RUN] Skip Google Sheets summary upload & sharing for {zone}")
            try:
                os.remove(local_summary_path)
            except Exception:
                pass
            continue

        zone_folder_id = get_or_create_drive_folder(drive_service, zone, PARENT_FOLDER_ID)
        sheet_name = f"{zone} CASH IN HAND"
        sh_email = sh_by_zone.get(zone.upper(), '')

        # Check and trash existing Zonal Summary sheet
        query = f"name = '{sheet_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{zone_folder_id}' in parents and trashed = false"
        res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
        existing_files = res.get('files', [])

        if existing_files:
            drive_service.files().update(fileId=existing_files[0]['id'], body={'trashed': True}).execute()
            log_message(f"Trashed old zonal summary sheet: {sheet_name}")

        # Upload fresh file
        file_metadata = {
            'name': sheet_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet',
            'parents': [zone_folder_id]
        }
        media = MediaFileUpload(local_summary_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        zonal_sheet_id = uploaded_file.get('id')
        log_message(f"Uploaded Zonal Summary Google Sheet successfully. ID: {zonal_sheet_id}")

        # Apply sheet protection (lock all columns except D18:D48 for SH self entry)
        try:
            ss_meta = sheets_service.spreadsheets().get(spreadsheetId=zonal_sheet_id, fields='sheets(properties(sheetId))').execute()
            real_sheet_id = ss_meta['sheets'][0]['properties']['sheetId']
            requests_body = [{
                'addProtectedRange': {
                    'protectedRange': {
                        'range': {'sheetId': real_sheet_id},
                        'description': f'Zonal Summary {zone} locked',
                        'warningOnly': False,
                        'editors': {'users': [sh_email] if sh_email else []}
                    }
                }
            }, {
                'addProtectedRange': {
                    'protectedRange': {
                        'range': {
                            'sheetId': real_sheet_id,
                            'startRowIndex': 17, 'endRowIndex': 17 + num_days,
                            'startColumnIndex': 3, 'endColumnIndex': 4
                        },
                        'description': f'SH editable range for {zone}',
                        'warningOnly': False,
                        'editors': {'users': [sh_email] if sh_email else []}
                    }
                }
            }]
            sheets_service.spreadsheets().batchUpdate(spreadsheetId=zonal_sheet_id, body={'requests': requests_body}).execute()
            log_message(f"Applied protection: locked all except D18:D{18+num_days-1} for SH {sh_email}")

            # Apply data validation on D18:D48
            validation_requests = [{
                'setDataValidation': {
                    'range': {
                        'sheetId': real_sheet_id,
                        'startRowIndex': 17, 'endRowIndex': 17 + num_days,
                        'startColumnIndex': 3, 'endColumnIndex': 4
                    },
                    'rule': {
                        'condition': {
                            'type': 'NUMBER_GREATER_THAN_EQ',
                            'values': [{'userEnteredValue': '0'}]
                        },
                        'inputMessage': 'Enter positive values.',
                        'strict': True,
                        'showCustomUi': True
                    }
                }
            }]
            sheets_service.spreadsheets().batchUpdate(spreadsheetId=zonal_sheet_id, body={'requests': validation_requests}).execute()
        except Exception as protect_err:
            log_message(f"Non-fatal zonal summary protection error: {protect_err}")

        # Share summary sheet with SH
        if sh_email:
            share_file(drive_service, zonal_sheet_id, sh_email, role='writer')

        # Clean local summary temp file
        try:
            os.remove(local_summary_path)
        except Exception:
            pass

        time.sleep(1)

    log_message("Zonal Summary Generation completed!")

# Unified Provisioning Orchestrator
def run_provisioning(selected_zones, month_str, num_days, dry_run=False, existing_action="overwrite"):
    log_message("Starting Generate & Provisioning process...")
    creds = get_oauth_credentials()
    gc = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    sheets_service = build('sheets', 'v4', credentials=creds)

    valid_fms = fetch_master_data(gc)
    email_mappings, sh_by_zone = get_email_mappings(gc)

    # Filter by user selected zones
    filtered_fms = {}
    for fm_name, fm_data in valid_fms.items():
        if fm_data['zone'] in selected_zones:
            filtered_fms[fm_name] = fm_data

    log_message(f"Selected zones: {selected_zones}. Found {len(filtered_fms)} FMs to process.")
    
    registry = []

    for fm_name, fm_data in filtered_fms.items():
        zone = fm_data['zone']
        fm_clean_name = clean_person_name(fm_name, zone)
        mapping = lookup_email_mapping(email_mappings, fm_name, zone, sh_by_zone)
        fm_email = mapping['email']
        boss_email = mapping['boss_email']
        boss_name = mapping['boss_name']
        sh_email = mapping.get('sh_email', '') or sh_by_zone.get(zone, '')

        log_message(f"\n--- Process: {fm_name} ({zone}) ---")
        
        # 1. Create formatted local sheet
        local_path = create_local_excel(fm_name, fm_data, month_str, num_days)
        log_message(f"Generated local file: {local_path}")
        
        if dry_run:
            log_message(f"  [DRY-RUN] Would upload and share: {fm_name} ({zone})")
            continue

        # 2. Get/create Drive folder
        zone_folder_id = get_or_create_drive_folder(drive_service, zone, PARENT_FOLDER_ID)

        # 3. Handle Google Sheet creation/update
        sheet_name = f"CASH IN HAND - {fm_name}"
        query = f"name = '{sheet_name}' and '{zone_folder_id}' in parents and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
        res = drive_service.files().list(q=query, spaces='drive', fields='files(id, webViewLink)').execute()
        files = res.get('files', [])

        if files:
            sheet_id = files[0]['id']
            web_link = files[0]['webViewLink']
            log_message(f"Found existing Google Sheet ID: {sheet_id}")
            
            # Check if tab exists
            ss = gc.open_by_key(sheet_id)
            tab_exists = False
            for ws in ss.worksheets():
                if ws.title == month_str:
                    tab_exists = True
                    break
            
            if tab_exists:
                if existing_action == "skip":
                    log_message(f"Tab '{month_str}' already exists in sheet. Skipping duplicate.")
                    continue
                else:
                    log_message(f"Tab '{month_str}' exists. Executing [{existing_action.upper()}] -> Replacing tab (uploading new tab first)...")
                    if not dry_run:
                        create_version_snapshot(drive_service, gc, sheet_id, sheet_name, f"REPLACE_TAB_{existing_action.upper()}", month_str)
            
            # 1. Add new sheet tab by temporary upload & copy FIRST
            temp_metadata = {'name': 'temp_upload', 'mimeType': 'application/vnd.google-apps.spreadsheet'}
            media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
            temp_file = drive_service.files().create(body=temp_metadata, media_body=media, fields='id').execute()
            
            temp_ss = gc.open_by_key(temp_file['id'])
            temp_ws = temp_ss.get_worksheet(0)
            copied_ws = temp_ws.copy_to(sheet_id)
            
            # 2. Now that the new tab is safely copied inside, we can cleanly delete any old tab with title month_str!
            ss = gc.open_by_key(sheet_id)
            if tab_exists and existing_action != "skip":
                try:
                    for ws in ss.worksheets():
                        if ws.title == month_str and str(ws.id) != str(copied_ws['sheetId']):
                            ss.del_worksheet(ws)
                            log_message(f"  ✔ Removed old tab '{month_str}'.")
                            break
                except Exception as e:
                    log_message(f"  Warning during old tab removal: {e}")
            
            # 3. Rename newly copied tab to month_str and delete default Sheet1 if present and not needed
            target_ws = ss.worksheet(copied_ws['title'])
            target_ws.update_title(month_str)
            
            # Clean up default Sheet1 if we just created this workbook or appended to it
            try:
                for ws in ss.worksheets():
                    if ws.title == "Sheet1" and len(ss.worksheets()) > 1:
                        ss.del_worksheet(ws)
                        break
            except Exception:
                pass
            
            # 4. Clean up temp upload file from Google Drive
            drive_service.files().delete(fileId=temp_file['id']).execute()
            log_message(f"Appended/Updated tab '{month_str}' in spreadsheet.")
        else:
            # Create new file
            file_metadata = {
                'name': sheet_name,
                'mimeType': 'application/vnd.google-apps.spreadsheet',
                'parents': [zone_folder_id]
            }
            media = MediaFileUpload(local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
            uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
            sheet_id = uploaded_file['id']
            web_link = uploaded_file['webViewLink']
            log_message(f"Created new Google Sheet ID: {sheet_id}")

        # 4. Share and Protect
        if fm_email:
            share_file(drive_service, sheet_id, fm_email, role='writer')
        if boss_email:
            share_file(drive_service, sheet_id, boss_email, role='writer')
        if sh_email and sh_email != fm_email:
            share_file(drive_service, sheet_id, sh_email, role='reader')

        # Add data validation rules (positive numbers only in columns C to total_col-1)
        _das = []
        for _m in fm_data['markets']:
            if _m['da_name'] and str(_m['da_name']).strip().upper() != 'VACANT':
                _d = str(_m['da_name']).strip().upper()
                if _d not in _das:
                    _das.append(_d)
        _total_col = 4 + len(fm_data['markets']) + len(_das)
        last_col_letter = get_column_letter(_total_col - 1)
        
        try:
            # Add strict data validation
            ss_api = gc.open_by_key(sheet_id)
            ws_api = ss_api.worksheet(month_str)
            editors_list = [e for e in [boss_email, sh_email] if e]
            validation_requests = [{
                'setDataValidation': {
                    'range': {
                        'sheetId': ws_api.id,
                        'startRowIndex': 17,
                        'endRowIndex': 17 + num_days,
                        'startColumnIndex': 2,
                        'endColumnIndex': _total_col - 1
                    },
                    'rule': {
                        'condition': {
                            'type': 'NUMBER_GREATER_THAN_EQ',
                            'values': [{'userEnteredValue': '0'}]
                        },
                        'inputMessage': 'Enter a positive number (0 or greater).',
                        'strict': True,
                        'showCustomUi': True
                    }
                }
            }]
            sheets_service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body={'requests': validation_requests}).execute()
            sort_and_protect_spreadsheet_tabs(gc, sheets_service, sheet_id, editors_list)
        except Exception as ex:
            log_message(f"Non-fatal data validation setup error: {ex}")

        registry.append({
            'FM Name': fm_clean_name,
            'Zone': zone,
            'Sheet ID': sheet_id,
            'URL': web_link,
            'FM Email': fm_email,
            'Boss Name': boss_name,
            'Boss Email': boss_email,
            'SH Email': sh_email,
        })
        
        # Clean local file
        try:
            os.remove(local_path)
        except Exception:
            pass

        time.sleep(1)

    log_message("\nUpdating Master Registry...")
    registry_name = "Master_Registry_Cash_In_Hand"
    query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
    res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
    files = res.get('files', [])

    if files:
        reg_sheet = gc.open_by_key(files[0]['id'])
        reg_ws = reg_sheet.get_worksheet(0)
    else:
        file_metadata = {
            'name': registry_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet',
            'parents': [PARENT_FOLDER_ID]
        }
        reg_file = drive_service.files().create(body=file_metadata, fields='id').execute()
        reg_sheet = gc.open_by_key(reg_file['id'])
        reg_ws = reg_sheet.get_worksheet(0)

    # Append registry rows to Master Registry
    header = ['FM Name', 'Zone', 'Sheet ID', 'URL', 'FM Email', 'Boss Name', 'Boss Email', 'SH Email']
    existing_rows = reg_ws.get_all_values()
    existing_map = {}
    for i, row in enumerate(existing_rows[1:], start=2):
        if row and row[0]:
            existing_map[row[0]] = i
            
    if not existing_rows or existing_rows[0] != header:
        reg_ws.update('A1', [header])
        
    for r in registry:
        new_row = [r['FM Name'], r['Zone'], r['Sheet ID'], r['URL'], r['FM Email'], r['Boss Name'], r['Boss Email'], r.get('SH Email', '')]
        fm_key = r['FM Name']
        if fm_key in existing_map:
            row_idx = existing_map[fm_key]
            reg_ws.update(f'A{row_idx}', [new_row])
            log_message(f"Updated master registry for {fm_key}")
        else:
            reg_ws.append_row(new_row)
            log_message(f"Appended {fm_key} to master registry")

    log_message("Generate & Provision process completed!")

    # Dynamic invocation of Zonal Summary sheet generator
    run_zonal_summaries(selected_zones, month_str, num_days, dry_run=dry_run)

# Boss Summary Sheet updater during rollover
def update_boss_summary_sheets(drive_service, gc, sheets_service, registry_records, month_str, num_days, dry_run=False):
    log_message(f"\n--- Updating Boss Summary Sheets for {month_str} ---")
    
    # Pre-fetch master data to determine FM column counts in memory
    valid_fms = fetch_master_data(gc)

    # Group registry by Boss Name / Boss Email
    boss_groups = {}
    for r in registry_records:
        boss_name = r.get('Boss Name')
        boss_email = r.get('Boss Email')
        fm_name = r.get('FM Name')
        sheet_url = r.get('URL')
        
        if not boss_name or "VACANT" in boss_name.upper():
            continue
            
        key = (boss_name, boss_email)
        if key not in boss_groups:
            boss_groups[key] = []
        boss_groups[key].append({
            'fm_name': fm_name,
            'sheet_url': sheet_url
        })
        
    for (boss_name, boss_email), fms in boss_groups.items():
        summary_name = f"CASH IN HAND SUMMARY - {boss_name}"
        log_message(f"Creating/Updating summary sheet for Boss: {boss_name} ({boss_email})")
        
        if dry_run:
            log_message(f"  [DRY-RUN] Would create/update Boss Summary sheet: {summary_name}")
            continue

        # Search if summary sheet already exists
        query = f"name = '{summary_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
        res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
        files = res.get('files', [])
        
        if files:
            boss_sheet_id = files[0]['id']
            boss_ss = gc.open_by_key(boss_sheet_id)
        else:
            file_metadata = {
                'name': summary_name,
                'mimeType': 'application/vnd.google-apps.spreadsheet',
                'parents': [PARENT_FOLDER_ID]
            }
            boss_file = drive_service.files().create(body=file_metadata, fields='id').execute()
            boss_sheet_id = boss_file['id']
            boss_ss = gc.open_by_key(boss_sheet_id)
            log_message(f"Created new Boss Summary sheet: {summary_name}")
            
            # Share with Boss
            if boss_email:
                share_file(drive_service, boss_sheet_id, boss_email, role='writer')
                    
        # Check if tab for month_str exists in summary sheet
        ws_exists = False
        for ws in boss_ss.worksheets():
            if ws.title == month_str:
                ws_exists = True
                boss_ws = ws
                break
                
        if not ws_exists:
            boss_ws = boss_ss.add_worksheet(title=month_str, rows=100, cols=20)
            log_message(f"Added tab {month_str} to Boss Summary sheet.")
        else:
            boss_ws.clear()
            log_message(f"Cleared existing tab {month_str} in Boss Summary sheet.")
            
        headers = ["DATE"] + [f["fm_name"] for f in fms]
        date_rows = [f"{d} {month_str}" for d in range(num_days, 0, -1)]
        
        full_rows = [headers]
        
        # Row 2 contains date and IMPORTRANGE formulas (which spill down automatically)
        row2 = [date_rows[0]]
        for fm_info in fms:
            fm_clean_name = clean_person_name(fm_info['fm_name'])
            fm_data = valid_fms.get(fm_clean_name)
            if fm_data:
                # Compute total column in memory
                markets = fm_data['markets']
                das = []
                for m in markets:
                    if m['da_name'] and str(m['da_name']).strip().upper() != 'VACANT':
                        da_str = str(m['da_name']).strip().upper()
                        if da_str not in das:
                            das.append(da_str)
                total_cols_count = 4 + len(markets) + len(das)
                total_col_let = get_column_letter(total_cols_count)
            else:
                total_col_let = "K" # fallback
                
            formula = f'=IMPORTRANGE("{fm_info["sheet_url"]}", "{month_str}!{total_col_let}18:{total_col_let}{18 + num_days - 1}")'
            row2.append(formula)
        full_rows.append(row2)
        
        # Remaining rows only contain the date in Col A
        for d_val in date_rows[1:]:
            full_rows.append([d_val])
            
        boss_ws.update('A1', full_rows, value_input_option='USER_ENTERED')
        log_message(f"Boss summary sheet updated for {boss_name}")

# Unified Monthly Rollover Orchestrator
def run_rollover(selected_zones, current_month_override=None, dry_run=False):
    log_message("Starting Monthly Rollover process...")
    creds = get_oauth_credentials()
    gc = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    sheets_service = build('sheets', 'v4', credentials=creds)

    email_mappings, sh_by_zone = get_email_mappings(gc)

    # Read Master Registry to find the live sheets
    registry_name = "Master_Registry_Cash_In_Hand"
    query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
    res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
    files = res.get('files', [])
    if not files:
        log_message("Error: Master Registry not found! Cannot perform rollover.")
        return
        
    reg_sheet = gc.open_by_key(files[0]['id'])
    reg_ws = reg_sheet.get_worksheet(0)
    reg_rows = reg_ws.get_all_values()

    # Convert rows to records dict
    headers = reg_rows[0]
    registry_records = []
    for r in reg_rows[1:]:
        rec = {}
        for idx, h in enumerate(headers):
            if idx < len(r):
                rec[h] = r[idx]
        registry_records.append(rec)

    # Retrieve owner email address to restrict editors in protected range
    try:
        owner_email = drive_service.about().get(fields="user(emailAddress)").execute()['user']['emailAddress']
    except Exception:
        owner_email = "me"  # fallback

    log_message(f"Loaded {len(registry_records)} FMs from Master Registry.")

    # Target month settings for Boss summary update
    final_target_month_str = None
    final_num_days = 30

    for r in registry_records:
        fm_name = r.get('FM Name')
        zone = r.get('Zone')
        sheet_id = r.get('Sheet ID')
        fm_email = r.get('FM Email')
        boss_email = r.get('Boss Email')
        sh_email = r.get('SH Email')

        if zone not in selected_zones:
            continue

        log_message(f"\n--- Rollover process for FM: {fm_name} (Zone: {zone}) | Sheet ID: {sheet_id} ---")

        try:
            ss = gc.open_by_key(sheet_id)
            worksheets = ss.worksheets()

            # Rule-based Target Month selection
            latest_tab_ws = None
            latest_tab_date = None
            
            # Scan existing tabs
            valid_tabs = []
            for ws in worksheets:
                p = parse_month_str(ws.title)
                if p:
                    valid_tabs.append((p, ws))

            if current_month_override:
                target_month_str = current_month_override.upper()
                # Parse days in override target month
                override_parsed = parse_month_str(target_month_str)
                if override_parsed:
                    _, num_days = calendar.monthrange(override_parsed[0], override_parsed[1])
                else:
                    num_days = 30 # default
                prev_month_str = None
                log_message(f"Mode: User defined override month. Target = {target_month_str} ({num_days} days)")
            elif not valid_tabs:
                # Rule 1: No sheets exist. Create for current calendar month.
                target_month_str, num_days = get_current_month_info(datetime.date.today())
                prev_month_str = None
                log_message(f"Rule 1 triggered (No tabs found). Creating tab for current calendar month: {target_month_str} ({num_days} days)")
            else:
                # Rule 2: Find latest tab, then create next month.
                valid_tabs.sort(key=lambda x: x[0])
                latest_tab_date, latest_tab_ws = valid_tabs[-1]
                prev_month_str = latest_tab_ws.title
                
                # Calculate next month from latest tab date
                latest_dt = datetime.date(latest_tab_date[0], latest_tab_date[1], 1)
                target_month_str, num_days = get_next_month_info(latest_dt)
                log_message(f"Rule 2 triggered. Found latest month tab: {prev_month_str}. Creating next month: {target_month_str} ({num_days} days)")

            final_target_month_str = target_month_str
            final_num_days = num_days

            # Check if target tab already exists
            tab_already_exists = False
            for ws in worksheets:
                if ws.title == target_month_str:
                    tab_already_exists = True
                    break

            if tab_already_exists:
                log_message(f"Tab '{target_month_str}' already exists in sheet -> Enforcing LOCK & ARCHIVE on previous month tab '{prev_month_str}'...")
                if latest_tab_ws and not dry_run:
                    lock_sheet(sheets_service, sheet_id, latest_tab_ws.id, owner_email, message=f"Locked archive for {prev_month_str}")
                continue

            if dry_run:
                log_message(f"[DRY-RUN] Would duplicate latest worksheet to '{target_month_str}'")
                log_message(f"[DRY-RUN] Would lock old tab '{prev_month_str}'")
                continue

            # Duplicate the template (index 0)
            if latest_tab_ws:
                new_ws = ss.duplicate_sheet(source_sheet_id=latest_tab_ws.id, insert_sheet_index=0, new_sheet_name=target_month_str)
                log_message(f"Duplicated '{prev_month_str}' as '{target_month_str}'")
                
                # Lock old sheet (prev_month) for FMs (restricts edits to owner_email only)
                lock_sheet(sheets_service, sheet_id, latest_tab_ws.id, owner_email, message=f"Locked archive for {prev_month_str}")
            else:
                log_message("Warning: No template sheet found to duplicate. Skipping.")
                continue

            # Find the total cash in hand column dynamically to fix the shifting bug
            row_11_values = [str(x).strip().upper() for x in new_ws.row_values(11)]
            try:
                total_col_idx = row_11_values.index("TOTAL CASH IN HAND") + 1
            except ValueError:
                total_col_idx = len(row_11_values)
                
            total_col_letter = get_column_letter(total_col_idx)
            last_input_col_letter = get_column_letter(total_col_idx - 1)
            log_message(f"Total Column detected at index {total_col_idx} ({total_col_letter}), Last input column is {last_input_col_letter}")

            # Adjust row count for dates if target num_days is different from prev_days
            if latest_tab_date:
                _, prev_days = calendar.monthrange(latest_tab_date[0], latest_tab_date[1])
            else:
                prev_days = 30
            diff = num_days - prev_days

            if diff > 0:
                insert_row_idx = 18 + prev_days
                body = {
                    "requests": [{
                        "insertDimension": {
                            "range": {
                                "sheetId": new_ws.id,
                                "dimension": "ROWS",
                                "startIndex": insert_row_idx - 1,
                                "endIndex": insert_row_idx - 1 + diff
                            },
                            "inheritFromBefore": True
                        }
                    }]
                }
                sheets_service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body=body).execute()
                log_message(f"Inserted {diff} rows in sheet.")
            elif diff < 0:
                delete_row_idx = 18 + num_days
                body = {
                    "requests": [{
                        "deleteDimension": {
                            "range": {
                                "sheetId": new_ws.id,
                                "dimension": "ROWS",
                                "startIndex": delete_row_idx - 1,
                                "endIndex": delete_row_idx - 1 - diff
                            }
                        }
                    }]
                }
                sheets_service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body=body).execute()
                log_message(f"Deleted {abs(diff)} rows from sheet.")

            # Batch updates for dates and formulas
            date_vals = [f"{d} {target_month_str}" for d in range(num_days, 0, -1)]
            cell_updates = []
            
            for idx, d_val in enumerate(date_vals):
                r_num = 18 + idx
                cell_updates.append({
                    'range': f'B{r_num}',
                    'values': [[d_val]]
                })
                
            # Clear input cells range C18 to [last_input_col][18+num_days-1]
            clear_range = f"C18:{last_input_col_letter}{18 + num_days - 1}"
            new_ws.batch_clear([clear_range])
            
            # Re-write SUM formulas
            for idx in range(num_days):
                r_num = 18 + idx
                sum_formula = f"=SUM(C{r_num}:{last_input_col_letter}{r_num})"
                cell_updates.append({
                    'range': f'{total_col_letter}{r_num}',
                    'values': [[sum_formula]]
                })
                
            new_ws.batch_update(cell_updates, value_input_option='USER_ENTERED')
            
            # Force Number format on input & formula cells so SUM formula always evaluates as a number and never shows as text
            format_body = {
                "requests": [
                    {
                        "repeatCell": {
                            "range": {
                                "sheetId": new_ws.id,
                                "startRowIndex": 17,
                                "endRowIndex": 17 + num_days,
                                "startColumnIndex": 2,
                                "endColumnIndex": total_col_idx
                            },
                            "cell": {
                                "userEnteredFormat": {
                                    "numberFormat": {
                                        "type": "NUMBER",
                                        "pattern": "#,##0"
                                    }
                                }
                            },
                            "fields": "userEnteredFormat.numberFormat"
                        }
                    }
                ]
            }
            sheets_service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body=format_body).execute()
            log_message("Cleared input data, updated dates, and enforced NUMBER formatting on formula cells successfully.")

        except Exception as e:
            log_message(f"Error during rollover for {fm_name}: {e}")

    # Update Boss Summary Sheets at the end of the rollover process
    if final_target_month_str:
        # Filter registry records to only include those in selected zones
        filtered_records = [rec for rec in registry_records if rec.get('Zone') in selected_zones]
        update_boss_summary_sheets(drive_service, gc, sheets_service, filtered_records, final_target_month_str, final_num_days, dry_run=dry_run)

    log_message("Monthly Rollover process completed!")

# Interactive Cloud Deletion & Cleanup Engine
def run_deletion_process(target, scope, month_name, selected_zones, selected_persons, dry_run):
    log_message(f"\n--- Starting Cloud Deletion / Cleanup Process [{target.upper()}] ---")
    try:
        creds = get_oauth_credentials()
        drive_service = build('drive', 'v3', credentials=creds)
        gc = gspread.authorize(creds)
        
        registry_name = "Master_Registry_Cash_In_Hand"
        query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
        res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
        files = res.get('files', [])
        if not files:
            log_message("Error: Master Registry not found in Drive.")
            return
            
        reg_sheet = gc.open_by_key(files[0]['id'])
        reg_ws = reg_sheet.get_worksheet(0)
        reg_rows = reg_ws.get_all_values()
        if len(reg_rows) < 2:
            log_message("Master Registry is empty. Nothing to delete.")
            return
            
        reg_cols = get_reg_col_indices(reg_rows[0]) if reg_rows else {}
        fm_col, z_col, sid_col = reg_cols.get('fm', 0), reg_cols.get('zone', 1), reg_cols.get('sheet_id', 2)
        
        deleted_count = 0
        rows_to_delete_from_registry = []
        
        for idx, r in enumerate(reg_rows[1:], start=2):
            if len(r) <= max(fm_col, z_col, sid_col):
                continue
            fm_name = r[fm_col].strip().upper()
            zone = r[z_col].strip()
            sheet_id = r[sid_col].strip()
            
            # Check scope match
            if scope == "zones" and zone not in selected_zones:
                continue
            if scope == "persons" and fm_name not in selected_persons:
                continue
                
            if target == "workbook":
                log_message(f"Deleting entire workbook for {fm_name} ({zone}) -> ID: {sheet_id}")
                if not dry_run:
                    try:
                        create_version_snapshot(drive_service, gc, sheet_id, f"CASH IN HAND - {fm_name}", "DELETE_WORKBOOK", "ALL_TABS")
                        drive_service.files().update(fileId=sheet_id, body={'trashed': True}).execute()
                        rows_to_delete_from_registry.append(idx)
                        deleted_count += 1
                    except Exception as ex:
                        log_message(f"  Error trashing workbook for {fm_name}: {ex}")
                else:
                    log_message(f"  [DRY-RUN] Would move workbook {sheet_id} to Trash.")
                    deleted_count += 1
            elif target == "tab":
                log_message(f"Checking workbook of {fm_name} ({zone}) for tab '{month_name}'...")
                if not dry_run:
                    try:
                        ss = gc.open_by_key(sheet_id)
                        for ws in ss.worksheets():
                            if ws.title.upper() == month_name.upper():
                                create_version_snapshot(drive_service, gc, sheet_id, f"CASH IN HAND - {fm_name}", "DELETE_TAB", month_name)
                                ss.del_worksheet(ws)
                                log_message(f"  Successfully deleted tab '{month_name}' from {fm_name}'s workbook.")
                                deleted_count += 1
                                break
                    except Exception as ex:
                        log_message(f"  Error deleting tab for {fm_name}: {ex}")
                else:
                    log_message(f"  [DRY-RUN] Would delete tab '{month_name}' from workbook {sheet_id}.")
                    deleted_count += 1
                    
        if target == "workbook" and rows_to_delete_from_registry and not dry_run:
            log_message("Cleaning up deleted workbooks from Master Registry...")
            for idx in sorted(rows_to_delete_from_registry, reverse=True):
                reg_ws.delete_rows(idx)
            log_message("Master Registry updated successfully.")
            
        log_message(f"\n✅ Deletion / Cleanup complete! Processed {deleted_count} items.")
    except Exception as e:
        log_message(f"Fatal error during deletion: {e}")

class DeleteWizardDialog(tk.Toplevel):
    def __init__(self, parent, selected_zones, dry_run=False):
        super().__init__(parent)
        self.title("🗑️ Interactive Cloud Deletion & Cleanup Wizard")
        self.geometry("720x620")
        self.configure(bg="#0F172A")
        self.transient(parent)
        self.grab_set()
        
        self.selected_zones = selected_zones
        self.dry_run = dry_run
        
        # Header
        header = tk.Frame(self, bg="#060816", bd=1, relief="ridge")
        header.pack(fill="x", padx=15, pady=(15, 10))
        tk.Label(header, text="🗑️ INTERACTIVE CLOUD DELETION & CLEANUP WIZARD", fg="#F87171", bg="#060816", font=("Segoe UI", 13, "bold")).pack(pady=8)
        
        main_frame = tk.Frame(self, bg="#0F172A")
        main_frame.pack(fill="both", expand=True, padx=20, pady=5)
        
        # Question 1: What do you want to delete?
        q1_frame = ttk.LabelFrame(main_frame, text=" 1. What do you want to delete? ")
        q1_frame.pack(fill="x", pady=6)
        
        self.target_type = tk.StringVar(value="tab")
        tk.Radiobutton(q1_frame, text="📄 Specific Month Tab(s) only (e.g. Delete JUL'26 tab inside workbooks without deleting files)", variable=self.target_type, value="tab", fg="#38BDF8", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold"), command=self.on_target_change).pack(anchor="w", padx=15, pady=4)
        tk.Radiobutton(q1_frame, text="🗑️ Entire Workbooks / Files (Move Google Drive Spreadsheet files to Trash permanently)", variable=self.target_type, value="workbook", fg="#F87171", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold"), command=self.on_target_change).pack(anchor="w", padx=15, pady=4)

        # Month input for tab deletion
        self.month_frame = tk.Frame(q1_frame, bg="#060816")
        self.month_frame.pack(fill="x", padx=35, pady=4)
        tk.Label(self, text="Target Month Tab Name to Delete (e.g. JUL'26, AUG'26):", fg="#CBD5E1", bg="#060816", font=("Segoe UI", 9)).pack(in_=self.month_frame, side="left", padx=5)
        self.month_entry = ttk.Entry(self.month_frame, width=12)
        self.month_entry.insert(0, "JUL'26")
        self.month_entry.pack(in_=self.month_frame, side="left", padx=5)
        
        # Question 2: Who do you want to delete for?
        q2_frame = ttk.LabelFrame(main_frame, text=" 2. Who do you want to delete for? (Scope) ")
        q2_frame.pack(fill="x", pady=6)
        
        self.scope_type = tk.StringVar(value="zones")
        tk.Radiobutton(q2_frame, text=f"🌐 All Persons in Currently Selected Zones ({', '.join(selected_zones[:5])}{'...' if len(selected_zones)>5 else ''})", variable=self.scope_type, value="zones", fg="#34D399", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold"), command=self.on_scope_change).pack(anchor="w", padx=15, pady=4)
        tk.Radiobutton(q2_frame, text="👤 Specific Field Person(s) only (Select names from list below)", variable=self.scope_type, value="persons", fg="#FBBF24", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold"), command=self.on_scope_change).pack(anchor="w", padx=15, pady=4)
        
        # Person selection listbox
        self.person_frame = tk.Frame(q2_frame, bg="#060816")
        self.person_frame.pack(fill="both", expand=True, padx=35, pady=5)
        tk.Label(self.person_frame, text="Click 'Load Persons' to fetch person names from Google Sheets:", fg="#CBD5E1", bg="#060816", font=("Segoe UI", 9)).pack(anchor="w")
        
        load_btn = tk.Button(self.person_frame, text="🔄 Load Field Person List", bg="#0F172A", fg="#00F2FE", font=("Segoe UI", 8, "bold"), command=self.load_persons)
        load_btn.pack(anchor="w", pady=3)
        
        self.person_listbox = tk.Listbox(self.person_frame, selectmode="multiple", bg="#0F172A", fg="#CBD5E1", selectbackground="#F87171", selectforeground="#FFFFFF", height=5, exportselection=False)
        self.person_listbox.pack(fill="both", expand=True, pady=3)
        self.person_frame.pack_forget()
        
        # Buttons
        btn_frame = tk.Frame(self, bg="#0F172A")
        btn_frame.pack(fill="x", pady=15)
        
        del_btn = tk.Button(btn_frame, text="🗑️ EXECUTE DELETION NOW", bg="#EF4444", fg="#FFFFFF", font=("Segoe UI", 10, "bold"), padx=15, pady=6, cursor="hand2", command=self.on_delete)
        del_btn.pack(side="left", padx=25)
        
        canc_btn = tk.Button(btn_frame, text="✖ CANCEL / CLOSE", bg="#64748B", fg="#FFFFFF", font=("Segoe UI", 10, "bold"), padx=15, pady=6, cursor="hand2", command=self.destroy)
        canc_btn.pack(side="right", padx=25)

    def on_target_change(self):
        if self.target_type.get() == "tab":
            self.month_frame.pack(fill="x", padx=35, pady=4)
        else:
            self.month_frame.pack_forget()

    def on_scope_change(self):
        if self.scope_type.get() == "persons":
            self.person_frame.pack(fill="both", expand=True, padx=35, pady=5)
        else:
            self.person_frame.pack_forget()
            
    def load_persons(self):
        self.person_listbox.delete(0, tk.END)
        self.person_listbox.insert(tk.END, "Loading from Cloud...")
        def fetch():
            try:
                creds = get_oauth_credentials()
                gc = gspread.authorize(creds)
                fms = fetch_master_data(gc)
                names = sorted([k.split(',')[0].strip().upper() for k, v in fms.items() if v['zone'] in self.selected_zones])
                self.person_listbox.delete(0, tk.END)
                for n in names:
                    self.person_listbox.insert(tk.END, n)
            except Exception as e:
                self.person_listbox.delete(0, tk.END)
                self.person_listbox.insert(tk.END, f"Error: {e}")
        t = threading.Thread(target=fetch)
        t.daemon = True
        t.start()
        
    def on_delete(self):
        target = self.target_type.get()
        scope = self.scope_type.get()
        month_name = self.month_entry.get().strip().upper()
        
        selected_persons = []
        if scope == "persons":
            indices = self.person_listbox.curselection()
            if not indices:
                messagebox.showerror("Error", "Please select at least one Field Person from the list.")
                return
            selected_persons = [self.person_listbox.get(i) for i in indices]
            
        if target == "tab" and not month_name:
            messagebox.showerror("Error", "Please enter the target month tab name to delete (e.g. JUL'26).")
            return
            
        warn_msg = f"Are you absolutely sure you want to delete {'ENTIRE WORKBOOKS' if target=='workbook' else f'Tab [{month_name}]'} for {'All Persons in Selected Zones' if scope=='zones' else f'{len(selected_persons)} selected persons'}?\n\nThis action cannot be undone!"
        if not messagebox.askyesno("CONFIRM DELETION", warn_msg, icon="warning"):
            return
            
        self.destroy()
        t = threading.Thread(target=run_deletion_process, args=(target, scope, month_name, self.selected_zones, selected_persons, self.dry_run))
        t.daemon = True
        t.start()

# Universal Delete Month Tab Engine (for dropdown action)
def execute_delete_month_tab(zones, month_str, dry_run=False):
    log_message(f"\n--- Starting Universal Delete Month Tab [{month_str}] for Zones: {zones} ---")
    try:
        creds = get_oauth_credentials()
        gc = gspread.authorize(creds)
        drive_service = build('drive', 'v3', credentials=creds)
        
        valid_fms = fetch_master_data(gc)
        deleted_count = 0
        ignored_count = 0
        
        for fm_name, fm_data in valid_fms.items():
            zone = fm_data['zone']
            if zone not in zones:
                continue
                
            sheet_name = f"CASH IN HAND - {fm_name}"
            q = f"name = '{sheet_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
            res = drive_service.files().list(q=q, spaces='drive', fields='files(id)').execute()
            files = res.get('files', [])
            if not files:
                log_message(f"Workbook not found for {fm_name} -> Ignoring.")
                ignored_count += 1
                continue
                
            sheet_id = files[0]['id']
            ss = gc.open_by_key(sheet_id)
            target_ws = None
            for ws in ss.worksheets():
                if ws.title == month_str:
                    target_ws = ws
                    break
                    
            if target_ws:
                if len(ss.worksheets()) <= 1:
                    log_message(f"Warning: Cannot delete tab '{month_str}' from {fm_name} because it is the only tab in the workbook. Ignored.")
                    ignored_count += 1
                else:
                    if not dry_run:
                        create_version_snapshot(drive_service, gc, sheet_id, sheet_name, "DELETE_TAB", month_str)
                        ss.del_worksheet(target_ws)
                    log_message(f"✔ DELETED month tab '{month_str}' from {fm_name}'s workbook.")
                    deleted_count += 1
            else:
                log_message(f"Tab '{month_str}' does not exist in {fm_name}'s workbook -> Ignored/Skipped.")
                ignored_count += 1
                
        # Also check Zonal Summary sheets
        for z in zones:
            zs_name = f"ZONAL SUMMARY - CASH IN HAND ({z})"
            q = f"name = '{zs_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
            res = drive_service.files().list(q=q, spaces='drive', fields='files(id)').execute()
            files = res.get('files', [])
            if files:
                ss = gc.open_by_key(files[0]['id'])
                for ws in ss.worksheets():
                    if ws.title == month_str and len(ss.worksheets()) > 1:
                        if not dry_run:
                            create_version_snapshot(drive_service, gc, files[0]['id'], zs_name, "DELETE_TAB", month_str)
                            ss.del_worksheet(ws)
                        log_message(f"✔ DELETED tab '{month_str}' from Zonal Summary ({z}).")
                        
        log_message(f"\n--- Universal Delete Completed: Deleted {deleted_count} tabs, Ignored {ignored_count} ---\n")
    except Exception as e:
        log_message(f"Error during Universal Delete: {e}\n")

# Unified Delta Confirmation Modal Dialog
class UnifiedDeltaDialog(tk.Toplevel):
    def __init__(self, parent, mode, target_month, existing_found, missing_fms, changed_emails, rollover_info=None):
        super().__init__(parent)
        self.title("⚡ Unified Delta & Action Confirmation Modal")
        self.geometry("680x580")
        self.configure(bg="#0F172A")
        self.transient(parent)
        self.grab_set() # Modal blocking
        self.existing_found = existing_found
        self.target_month = target_month
        
        self.result = {"confirmed": False, "existing_action": "overwrite", "do_missing": True, "do_emails": True, "do_rollover": True}
        
        # Title Header
        header_frame = tk.Frame(self, bg="#060816", bd=1, relief="ridge")
        header_frame.pack(fill="x", padx=15, pady=(15, 10))
        tk.Label(header_frame, text="⚡ SYSTEM ACTION & DELTA CONFIRMATION", fg="#00F2FE", bg="#060816", font=("Segoe UI", 13, "bold")).pack(pady=8)
        
        # Scrollable container / Main Frame
        main_frame = tk.Frame(self, bg="#0F172A")
        main_frame.pack(fill="both", expand=True, padx=20, pady=5)
        
        # 1. Rollover Check (if mode == rollover)
        if mode == "rollover" and rollover_info:
            rf = ttk.LabelFrame(main_frame, text=" 🔄 Monthly Rollover Purpose & Target ")
            rf.pack(fill="x", pady=6)
            self.roll_var = tk.BooleanVar(value=True)
            lbl_text = f"Last month detected: {rollover_info.get('latest_tab', 'None')}  |  Target next month: {rollover_info.get('next_month', target_month)}"
            tk.Label(rf, text=lbl_text, fg="#CBD5E1", bg="#060816", justify="left", font=("Segoe UI", 9)).pack(anchor="w", padx=10, pady=4)
            tk.Checkbutton(rf, text="Confirm generating rollover tabs for next month & locking previous month", variable=self.roll_var, fg="#00F2FE", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10, pady=4)
        
        # 2. Existing Sheets Action (Separate selections for Overwrite vs Delete vs Skip)
        if existing_found:
            ef = ttk.LabelFrame(main_frame, text=f" 📂 Existing Sheets Found ({target_month}) ")
            ef.pack(fill="x", pady=6)
            tk.Label(ef, text="Sheets/Tabs for this month already exist in Google Drive. Choose your exact action:", fg="#FBBF24", bg="#060816", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10, pady=4)
            
            self.exist_var = tk.StringVar(value="overwrite")
            tk.Radiobutton(ef, text="🔄 Overwrite existing contents (Preserves Drive files, URLs & IMPORTRANGEs)", variable=self.exist_var, value="overwrite", fg="#34D399", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=15, pady=2)
            tk.Radiobutton(ef, text="🗑️ Delete old sheets & Re-create fresh files (Warning: Changes Drive URLs)", variable=self.exist_var, value="delete", fg="#F87171", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=15, pady=2)
            tk.Radiobutton(ef, text="⏭️ Skip / Ignore existing sheets", variable=self.exist_var, value="skip", fg="#94A3B8", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9)).pack(anchor="w", padx=15, pady=2)
        else:
            self.exist_var = tk.StringVar(value="none")

        # 3. Missing Markets Action
        if missing_fms:
            mf = ttk.LabelFrame(main_frame, text=f" 🆕 Missing Markets Detected ({len(missing_fms)}) ")
            mf.pack(fill="x", pady=6)
            names_str = ", ".join([f[0] for f in missing_fms[:6]])
            if len(missing_fms) > 6: names_str += f" (+{len(missing_fms)-6} more)"
            tk.Label(mf, text=f"New markets in Google Sheet not found in Drive Registry:\n{names_str}", fg="#CBD5E1", bg="#060816", justify="left", font=("Segoe UI", 9)).pack(anchor="w", padx=10, pady=4)
            self.miss_var = tk.BooleanVar(value=True)
            tk.Checkbutton(mf, text="Generate missing sheets & Auto-expand Boss Summary columns via IMPORTRANGE", variable=self.miss_var, fg="#00F2FE", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10, pady=4)
        else:
            self.miss_var = tk.BooleanVar(value=False)

        # 4. Email Changes Action
        if changed_emails:
            emf = ttk.LabelFrame(main_frame, text=f" 📧 Email ID Modifications ({len(changed_emails)}) ")
            emf.pack(fill="x", pady=6)
            tk.Label(emf, text="Email IDs changed in Google Sheets compared to Master Registry.", fg="#CBD5E1", bg="#060816", font=("Segoe UI", 9)).pack(anchor="w", padx=10, pady=4)
            self.em_var = tk.BooleanVar(value=True)
            tk.Checkbutton(emf, text="Update Drive sharing permissions & update Master Registry", variable=self.em_var, fg="#00F2FE", bg="#060816", selectcolor="#0F172A", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=10, pady=4)
        else:
            self.em_var = tk.BooleanVar(value=False)
            
        # If nothing special detected, show friendly info
        if not existing_found and not missing_fms and not changed_emails and mode != "rollover":
            inf_frame = tk.Frame(main_frame, bg="#060816", bd=1, relief="ridge")
            inf_frame.pack(fill="both", expand=True, pady=15)
            tk.Label(inf_frame, text="✅ All pre-checks passed! No existing conflicts or missing deltas found.\n\nClick Proceed to start generating sheets.", fg="#34D399", bg="#060816", font=("Segoe UI", 11, "bold")).pack(pady=25)

        # Buttons
        btn_frame = tk.Frame(self, bg="#0F172A")
        btn_frame.pack(fill="x", pady=15)
        
        proc_btn = tk.Button(btn_frame, text="✔ PROCEED WITH SELECTED ACTIONS", bg="#00F2FE", fg="#000000", font=("Segoe UI", 10, "bold"), padx=15, pady=6, cursor="hand2", command=self.on_proceed)
        proc_btn.pack(side="left", padx=25)
        
        canc_btn = tk.Button(btn_frame, text="✖ CANCEL", bg="#EF4444", fg="#FFFFFF", font=("Segoe UI", 10, "bold"), padx=15, pady=6, cursor="hand2", command=self.on_cancel)
        canc_btn.pack(side="right", padx=25)
        
    def on_proceed(self):
        selected_act = self.exist_var.get()
        if self.existing_found and selected_act in ["overwrite", "delete"]:
            ans = messagebox.askyesno(
                "WARNING: OVERRIDE / REPLACE EXISTING DATA",
                f"ARE YOU SURE?\n\nYou selected to '{selected_act.upper()}' existing sheets for {self.target_month}.\n\nWARNING: OLD DATA WILL BE DELETED AND REPLACED!\n\nClick YES to proceed with replacement.\nClick NO to cancel replacement and IGNORE existing sheets.",
                icon='warning'
            )
            if not ans:
                selected_act = "skip"
                messagebox.showinfo("Action Ignored", f"Existing sheets for {self.target_month} will be IGNORED (Skipped). No data will be replaced.")
        
        self.result["confirmed"] = True
        self.result["existing_action"] = selected_act
        self.result["do_missing"] = self.miss_var.get()
        self.result["do_emails"] = self.em_var.get()
        if hasattr(self, "roll_var"):
            self.result["do_rollover"] = self.roll_var.get()
        self.destroy()
        
    def on_cancel(self):
        self.result["confirmed"] = False
        self.destroy()

# Tkinter GUI Setup
class CashInHandApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cash In Hand Unified Manager")
        self.geometry("800x600")
        self.configure(bg="#060816")
        
        # Style
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure(".", background="#060816", foreground="#CBD5E1")
        self.style.configure("TLabel", background="#060816", foreground="#CBD5E1", font=("Segoe UI", 10))
        self.style.configure("TButton", background="#0F172A", foreground="#00F2FE", font=("Segoe UI", 10, "bold"))
        self.style.configure("TFrame", background="#060816")
        self.style.configure("TCombobox", fieldbackground="#0F172A", background="#0F172A", foreground="#00F2FE", font=("Segoe UI", 10, "bold"))
        self.style.map("TCombobox", fieldbackground=[("readonly", "#0F172A")], foreground=[("readonly", "#00F2FE")], selectbackground=[("readonly", "#00F2FE")], selectforeground=[("readonly", "#060816")])
        
        # Set global log callback
        global log_callback
        log_callback = self.gui_log
        
        self.create_widgets()

    def gui_log(self, text):
        self.log_text.insert(tk.END, text)
        self.log_text.see(tk.END)

    def create_widgets(self):
        # Title Label
        title_lbl = tk.Label(self, text="CASH IN HAND ORCHESTRATION SYSTEM", fg="#00F2FE", bg="#060816", font=("Segoe UI", 16, "bold"))
        title_lbl.pack(pady=10)

        # Status Bar Frame
        status_frame = tk.Frame(self, bg="#0F172A", bd=1, relief="ridge")
        status_frame.pack(fill="x", padx=20, pady=(0, 6))
        
        today_dhaka = get_dhaka_today()
        status_lbl = tk.Label(status_frame, text=f" 🌐 System Status: ONLINE 🟢 | 🕒 Timezone: Asia/Dhaka (BST) | 📅 Today's Date: {today_dhaka.strftime('%d %b %Y')} ", fg="#00F2FE", bg="#0F172A", font=("Segoe UI", 9, "bold"))
        status_lbl.pack(pady=4)

        # Quick Links Bar
        links_frame = ttk.LabelFrame(self, text=" 🔗 Quick Access Cloud Links & Resources (Click to Open in Browser) ")
        links_frame.pack(fill="x", padx=20, pady=4)
        
        l_btn_frame = tk.Frame(links_frame, bg="#060816")
        l_btn_frame.pack(fill="x", padx=5, pady=4)
        
        tk.Button(l_btn_frame, text="📊 Field Force Master", bg="#0F172A", fg="#38BDF8", font=("Segoe UI", 8, "bold"), cursor="hand2", command=lambda: webbrowser.open("https://docs.google.com/spreadsheets/d/1ywTyruBLxNXz6pjsGgufNstb0hOsrM9P-ER65iVvqN8")).pack(side="left", padx=5, pady=2)
        tk.Button(l_btn_frame, text="🗄️ Field Force Backup", bg="#0F172A", fg="#34D399", font=("Segoe UI", 8, "bold"), cursor="hand2", command=lambda: self.open_cloud_backup("mpo")).pack(side="left", padx=5, pady=2)
        tk.Button(l_btn_frame, text="📧 Mail Master Sheet", bg="#0F172A", fg="#FBBF24", font=("Segoe UI", 8, "bold"), cursor="hand2", command=lambda: webbrowser.open("https://docs.google.com/spreadsheets/d/1f5SFvhH8Bjb3OUlpof68teBktHuYyVELioxLv_KWXJo")).pack(side="left", padx=5, pady=2)
        tk.Button(l_btn_frame, text="📨 Mail Backup Sheet", bg="#0F172A", fg="#A78BFA", font=("Segoe UI", 8, "bold"), cursor="hand2", command=lambda: self.open_cloud_backup("mail")).pack(side="left", padx=5, pady=2)
        tk.Button(l_btn_frame, text="📂 Master Drive Folder", bg="#0F172A", fg="#F472B6", font=("Segoe UI", 8, "bold"), cursor="hand2", command=lambda: webbrowser.open("https://drive.google.com/drive/folders/1iOFeqywnIZ_yVclg_Em2U1npPtsokfGk")).pack(side="left", padx=5, pady=2)
        tk.Button(l_btn_frame, text="📋 Master Registry", bg="#0F172A", fg="#E2E8F0", font=("Segoe UI", 8, "bold"), cursor="hand2", command=lambda: webbrowser.open("https://drive.google.com/drive/folders/1iOFeqywnIZ_yVclg_Em2U1npPtsokfGk")).pack(side="left", padx=5, pady=2)

        # Action & Month Selection Frame (Universal Operation Engine with Dropdown Options)
        action_frame = ttk.LabelFrame(self, text=" ⚙️ Select Action & Target Month (Universal Dropdown Control Engine) ")
        action_frame.pack(fill="x", padx=20, pady=5)
        
        ttk.Label(action_frame, text="Select Operation Action:").grid(row=0, column=0, padx=10, pady=8, sticky="e")
        self.action_var = tk.StringVar(value="🟢 INSERT / GENERATE (Create if missing, Ignore if exists)")
        action_options = [
            "🟢 INSERT / GENERATE (Create if missing, Ignore if exists)",
            "🟠 OVERRIDE / REPLACE (Replace if exists, Create if missing)",
            "🔴 DELETE MONTH TAB (Delete selected month from workbooks)",
            "🔵 ARCHIVE & LOCK PREV MONTH / ROLLOVER (Freeze old data & start new)"
        ]
        self.action_cb = ttk.Combobox(action_frame, textvariable=self.action_var, values=action_options, width=58, state="readonly", exportselection=False)
        self.action_cb.grid(row=0, column=1, padx=10, pady=8, sticky="w")
        if action_options:
            self.action_cb.current(0)
        
        ttk.Label(action_frame, text="Select Target Month:").grid(row=1, column=0, padx=10, pady=8, sticky="e")
        
        # Populate month names dynamically starting from current year up to next 3 years (no old unnecessary years)
        today_dhaka = get_dhaka_today()
        current_yr_full = today_dhaka.year
        month_list = []
        for yr_full in range(current_yr_full, current_yr_full + 4):
            yr_short = str(yr_full)[2:]
            for m_idx, m_name in enumerate(['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'], start=1):
                if yr_full == current_yr_full and m_idx < today_dhaka.month:
                    continue  # Skip passed months of current year as requested ("old months dorkar e nei")
                month_list.append(f"{m_name}'{yr_short}")
        
        def_month_str, _ = get_current_month_info(today_dhaka)
        self.month_var = tk.StringVar(value=def_month_str)
        self.month_cb = ttk.Combobox(action_frame, textvariable=self.month_var, values=month_list, width=15, state="readonly", exportselection=False)
        self.month_cb.grid(row=1, column=1, padx=10, pady=8, sticky="w")
        if def_month_str in month_list:
            self.month_cb.current(month_list.index(def_month_str))
        elif month_list:
            self.month_cb.current(0)

        # Settings Frame (Zone Selection)
        settings_frame = ttk.LabelFrame(self, text="Configuration Settings")
        settings_frame.pack(fill="x", padx=20, pady=5)
        
        # Zone selection list (Multi-select) with Checkbox container
        zone_lbl = ttk.Label(settings_frame, text="Select Zones to Process (Hold Ctrl for multiple):")
        zone_lbl.grid(row=0, column=0, padx=10, pady=5, sticky="ne")
        
        zone_container = tk.Frame(settings_frame, bg="#0F172A")
        zone_container.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        
        self.all_zones_var = tk.BooleanVar(value=True)
        self.all_zones_chk = tk.Checkbutton(zone_container, text="☑ ALL ZONES (Select / Deselect All)", variable=self.all_zones_var, fg="#00F2FE", bg="#0F172A", selectcolor="#060816", font=("Segoe UI", 9, "bold"), command=self.toggle_all_zones)
        self.all_zones_chk.pack(anchor="w", pady=2)
        
        self.zone_listbox = tk.Listbox(zone_container, selectmode="multiple", bg="#0F172A", fg="#CBD5E1", selectbackground="#00F2FE", selectforeground="#060816", height=5, exportselection=False)
        self.zone_listbox.pack(fill="both", expand=True, pady=2)
        self.zone_listbox.bind("<<ListboxSelect>>", self.on_zone_select)
        
        # Pre-populate all 24 zones
        all_zones = ['DK.A', 'FENI', 'RAJ', 'HATIB.FM', 'MAIZ', 'JSR+KHL', 'JSR.B', 'NSD', 'TANG', 'DK.B', 'BARI', 'MYM.A', 'JPUR', 'RNG.A1', 'HOBI', 'COM', 'JSR.C', 'GAIB', 'MYM1.AM', 'CTG.B', 'THAK+DNJ', 'SLT', 'FRD', 'CTG.A']
        all_zones.sort()
        for z in all_zones:
            self.zone_listbox.insert(tk.END, z)
            
        # Select ALL zones by default
        self.zone_listbox.select_set(0, tk.END)

        # Action Buttons
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", padx=20, pady=8)
        
        self.run_btn = ttk.Button(btn_frame, text="▶ RUN SYSTEM PROCESS", command=self.start_process_thread)
        self.run_btn.pack(side="left", padx=8)

        del_wiz_btn = tk.Button(btn_frame, text="🗑️ OPEN DELETE / CLEANUP WIZARD", bg="#EF4444", fg="#FFFFFF", font=("Segoe UI", 9, "bold"), padx=10, pady=4, cursor="hand2", command=self.open_delete_wizard)
        del_wiz_btn.pack(side="left", padx=8)

        fix_form_btn = tk.Button(btn_frame, text="🔧 FIX FORMULAS / FORMATTING", bg="#3B82F6", fg="#FFFFFF", font=("Segoe UI", 9, "bold"), padx=10, pady=4, cursor="hand2", command=self.fix_formulas_in_selected_zones)
        fix_form_btn.pack(side="left", padx=8)

        quick_share_btn = tk.Button(btn_frame, text="🔒 QUICK LOCK & SHARE EXISTING SHEETS", bg="#10B981", fg="#FFFFFF", font=("Segoe UI", 9, "bold"), padx=10, pady=4, cursor="hand2", command=self.start_quick_share_thread)
        quick_share_btn.pack(side="left", padx=8)

        self.dryrun_var = tk.BooleanVar(value=False)
        self.dryrun_cb = ttk.Checkbutton(btn_frame, text="Dry Run (Simulate Only)", variable=self.dryrun_var)
        self.dryrun_cb.pack(side="left", padx=8)

        # Log Terminal Area
        log_lbl = ttk.Label(self, text="Real-time Execution Output Logs:")
        log_lbl.pack(anchor="w", padx=20, pady=(6, 2))
        
        self.log_text = scrolledtext.ScrolledText(self, height=11, bg="#000000", fg="#00F2FE", insertbackground="#00F2FE", font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True, padx=20, pady=5)
        
        self.gui_log("Welcome to Cash In Hand Orchestration System. Select options and click Run.\n")

    def toggle_all_zones(self):
        if self.all_zones_var.get():
            self.zone_listbox.select_set(0, tk.END)
        else:
            self.zone_listbox.selection_clear(0, tk.END)

    def on_zone_select(self, event=None):
        total = self.zone_listbox.size()
        selected = len(self.zone_listbox.curselection())
        if selected == total and total > 0:
            self.all_zones_var.set(True)
        else:
            self.all_zones_var.set(False)

    def open_cloud_backup(self, sheet_type):
        self.gui_log(f"Opening {sheet_type.upper()} Master Backup Copy from Google Drive...\n")
        def fetch_and_open():
            try:
                creds = get_oauth_credentials()
                drive_service = build('drive', 'v3', credentials=creds)
                backup_name = "Backup_Copy_DreamApps_MPO_FM_Codes" if sheet_type == "mpo" else "Backup_Copy_Mail_Address_Master"
                target_master_id = DATA_SHEET_ID if sheet_type == "mpo" else EMAIL_SHEET_ID
                
                query = f"name = '{backup_name}' and '{PARENT_FOLDER_ID}' in parents and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
                res = drive_service.files().list(q=query, spaces='drive', fields='files(id, webViewLink)').execute()
                files = res.get('files', [])
                
                if files:
                    link = files[0].get('webViewLink', f"https://docs.google.com/spreadsheets/d/{files[0]['id']}")
                    self.gui_log(f"Found existing backup copy -> Opening ID: {files[0]['id']}\n")
                    webbrowser.open(link)
                else:
                    self.gui_log(f"Backup copy '{backup_name}' does not exist yet. Generating ONE Master Backup Copy now...\n")
                    copy_meta = {'name': backup_name, 'parents': [PARENT_FOLDER_ID]}
                    new_file = drive_service.files().copy(fileId=target_master_id, body=copy_meta, fields='id, webViewLink').execute()
                    link = new_file.get('webViewLink', f"https://docs.google.com/spreadsheets/d/{new_file['id']}")
                    self.gui_log(f"Generated ONE Master Backup Copy ID: {new_file['id']} -> Opening in browser...\n")
                    webbrowser.open(link)
            except Exception as e:
                self.gui_log(f"Error opening/creating backup: {e}\n")
                messagebox.showerror("Backup Error", f"Could not access or generate backup sheet:\n{e}")
        t = threading.Thread(target=fetch_and_open)
        t.daemon = True
        t.start()

    def fix_formulas_in_selected_zones(self):
        selected_indices = self.zone_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Error", "Please select at least one Zone from the list.")
            return
        zones = [self.zone_listbox.get(i) for i in selected_indices]
        month_override = self.month_entry.get().strip()
        if not month_override:
            month_str, num_days = get_current_month_info(get_dhaka_today())
        else:
            month_str = month_override.upper()
            p = parse_month_str(month_str)
            num_days = calendar.monthrange(p[0], p[1])[1] if p else 31

        self.gui_log(f"\n--- Starting Formula & Cell Formatting Fix for {month_str} in zones: {zones} ---\n")
        
        def run_fix():
            try:
                creds = get_oauth_credentials()
                gc = gspread.authorize(creds)
                drive_service = build('drive', 'v3', credentials=creds)
                sheets_service = build('sheets', 'v4', credentials=creds)
                
                valid_fms = fetch_master_data(gc)
                for fm_name, fm_data in valid_fms.items():
                    zone = fm_data['zone']
                    if zone not in zones:
                        continue
                        
                    sheet_name = f"CASH IN HAND - {fm_name}"
                    q = f"name = '{sheet_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
                    res = drive_service.files().list(q=q, spaces='drive', fields='files(id)').execute()
                    files = res.get('files', [])
                    if not files:
                        continue
                        
                    sheet_id = files[0]['id']
                    ss = gc.open_by_key(sheet_id)
                    target_ws = None
                    for ws in ss.worksheets():
                        if ws.title == month_str:
                            target_ws = ws
                            break
                    if not target_ws:
                        continue
                        
                    self.gui_log(f"Fixing formulas & number formatting for: {fm_name} ({zone})...\n")
                    row_11_values = [str(x).strip().upper() for x in target_ws.row_values(11)]
                    try:
                        total_col_idx = row_11_values.index("TOTAL CASH IN HAND") + 1
                    except ValueError:
                        total_col_idx = len(row_11_values) if row_11_values else 10
                        
                    total_col_letter = get_column_letter(total_col_idx)
                    last_input_col_letter = get_column_letter(total_col_idx - 1)
                    
                    # 1. Re-write SUM formulas with USER_ENTERED
                    cell_updates = []
                    for idx in range(num_days):
                        r_num = 18 + idx
                        sum_formula = f"=SUM(C{r_num}:{last_input_col_letter}{r_num})"
                        cell_updates.append({
                            'range': f'{total_col_letter}{r_num}',
                            'values': [[sum_formula]]
                        })
                    target_ws.batch_update(cell_updates, value_input_option='USER_ENTERED')
                    
                    # 2. Force NUMBER format via Google Sheets API repeatCell
                    format_body = {
                        "requests": [
                            {
                                "repeatCell": {
                                    "range": {
                                        "sheetId": target_ws.id,
                                        "startRowIndex": 17,
                                        "endRowIndex": 17 + num_days,
                                        "startColumnIndex": 2,
                                        "endColumnIndex": total_col_idx
                                    },
                                    "cell": {
                                        "userEnteredFormat": {
                                            "numberFormat": {
                                                "type": "NUMBER",
                                                "pattern": "#,##0"
                                            }
                                        }
                                    },
                                    "fields": "userEnteredFormat.numberFormat"
                                }
                            }
                        ]
                    }
                    sheets_service.spreadsheets().batchUpdate(spreadsheetId=sheet_id, body=format_body).execute()
                    self.gui_log(f"  -> Successfully enforced NUMBER format & evaluated formulas for {fm_name}.\n")
                self.gui_log("Formula & Cell Formatting Fix completed successfully across selected zones!\n")
            except Exception as e:
                self.gui_log(f"Error during formula fix: {e}\n")
                
        t = threading.Thread(target=run_fix)
        t.daemon = True
        t.start()

    def start_quick_share_thread(self):
        selected_indices = self.zone_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("No Zone Selected", "Please select at least one zone from Configuration Settings.")
            return
        selected_zones = [self.zone_listbox.get(i) for i in selected_indices]
        self.gui_log(f"Starting Quick Share for Existing Sheets in zones: {selected_zones}...\n")
        threading.Thread(target=self.run_quick_share, args=(selected_zones,), daemon=True).start()

    def run_quick_share(self, selected_zones):
        try:
            creds = get_oauth_credentials()
            gc = gspread.authorize(creds)
            drive_service = build('drive', 'v3', credentials=creds)
            
            valid_fms = fetch_master_data(gc)
            email_mappings, sh_by_zone = get_email_mappings(gc)
            
            shared_count = 0
            for fm_name, fm_data in valid_fms.items():
                zone = fm_data['zone']
                if zone not in selected_zones:
                    continue
                    
                fm_clean = clean_person_name(fm_name, zone)
                mapping = lookup_email_mapping(email_mappings, fm_name, zone, sh_by_zone)
                fm_email = mapping['email']
                boss_email = mapping['boss_email']
                sh_email = mapping.get('sh_email', '') or sh_by_zone.get(zone, '')
                
                sheet_name = f"CASH IN HAND - {fm_name}"
                q = f"name = '{sheet_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
                res = drive_service.files().list(q=q, spaces='drive', fields='files(id)').execute()
                files = res.get('files', [])
                
                if files:
                    sheet_id = files[0]['id']
                    self.gui_log(f"Sharing & locking existing sheet for {fm_name} (ID: {sheet_id})...\n")
                    if fm_email:
                        share_file(drive_service, sheet_id, fm_email, role='writer')
                    if boss_email:
                        share_file(drive_service, sheet_id, boss_email, role='writer')
                    if sh_email and sh_email != fm_email:
                        share_file(drive_service, sheet_id, sh_email, role='reader')
                        
                    # Apply cell locking and tab sorting
                    try:
                        editors_list = [e for e in [boss_email, sh_email] if e]
                        sheets_service = build('sheets', 'v4', credentials=creds)
                        sort_and_protect_spreadsheet_tabs(gc, sheets_service, sheet_id, editors_list)
                    except Exception as l_ex:
                        self.gui_log(f"  Note during locking for {fm_name}: {l_ex}\n")
                        
                    shared_count += 1
                else:
                    self.gui_log(f"  Warning: Sheet not found in Drive for {fm_name}\n")
                    
            self.gui_log(f"\n✔ Quick Share & Lock completed! Successfully updated {shared_count} existing FM sheets without regenerating files.\n")
            messagebox.showinfo("Quick Share Complete", f"Successfully updated sharing permissions for {shared_count} existing FM sheets!")
        except Exception as e:
            self.gui_log(f"Error during Quick Share: {e}\n")

    def open_delete_wizard(self):
        selected_indices = self.zone_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Error", "Please select at least one Zone first from the list.")
            return
        zones = [self.zone_listbox.get(i) for i in selected_indices]
        dry_run = self.dryrun_var.get()
        DeleteWizardDialog(self, zones, dry_run=dry_run)

    def on_mode_change(self):
        if self.op_mode.get() == "rollover":
            self.month_lbl.configure(text="Target Month Override (e.g. AUG'26 or leave blank for auto):")
        else:
            self.month_lbl.configure(text="Target Month (e.g. JUL'26):")

    def start_process_thread(self):
        # 1. Check Online/Offline Status first
        if not check_online_status():
            self.gui_log("System is offline! Closing application auto...\n")
            self.after(2000, self.destroy)
            return

        # Fetch configurations
        selected_indices = self.zone_listbox.curselection()
        if not selected_indices:
            messagebox.showerror("Error", "Please select at least one Zone to process.")
            return

        zones = [self.zone_listbox.get(i) for i in selected_indices]
        action_sel = self.action_var.get()
        if "INSERT" in action_sel:
            mode = "provision"
            existing_action = "skip"
        elif "OVERRIDE" in action_sel or "REPLACE" in action_sel:
            mode = "provision"
            existing_action = "overwrite"
        elif "DELETE" in action_sel:
            mode = "delete"
            existing_action = "delete"
        elif "ROLLOVER" in action_sel:
            mode = "rollover"
            existing_action = "skip"
        else:
            mode = "provision"
            existing_action = "overwrite"
            
        month_override = self.month_var.get().strip()
        dry_run = self.dryrun_var.get()

        self.run_btn.configure(state="disabled")
        
        if mode == "delete":
            warn_msg = f"Are you sure you want to DELETE month tab [{month_override}] across {len(zones)} selected zones?\n\nThis will remove tab '{month_override}' from all individual and zonal summary workbooks where it exists."
            if not messagebox.askyesno("CONFIRM MONTH TAB DELETION", warn_msg, icon="warning"):
                self.run_btn.configure(state="normal")
                return
            self.gui_log(f"\nStarting Universal Month Tab Deletion [{month_override}]...\n")
            t_del = threading.Thread(target=self.execute_process, args=(mode, zones, month_override, dry_run, existing_action))
            t_del.daemon = True
            t_del.start()
            return

        self.gui_log("Analyzing Google Drive, checking existing sheets, and detecting system deltas...\n")

        def run_checks_and_confirm():
            try:
                creds = get_oauth_credentials()
                drive_service = build('drive', 'v3', credentials=creds)
                gc = gspread.authorize(creds)
                sheets_service = build('sheets', 'v4', credentials=creds)

                # Execute immediate Company Hardcoded Master vs Operational Sheet sync verification!
                if not perform_company_master_sync_check(gc, zones):
                    self.gui_log("\n❌ Execution paused/cancelled by user to update Operational Field Force Sheet.\n")
                    self.run_btn.configure(state="normal")
                    return

                rollover_info = None
                if mode == "rollover":
                    registry_name = "Master_Registry_Cash_In_Hand"
                    query = f"name = '{registry_name}' and mimeType = 'application/vnd.google-apps.spreadsheet' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
                    res = drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()
                    files = res.get('files', [])
                    if not files:
                        self.gui_log("Error: Master Registry not found. Cannot determine latest month.\n")
                        self.run_btn.configure(state="normal")
                        return
                        
                    reg_sheet = gc.open_by_key(files[0]['id'])
                    reg_ws = reg_sheet.get_worksheet(0)
                    reg_rows = reg_ws.get_all_values()
                    if len(reg_rows) < 2:
                        self.gui_log("Error: Master Registry is empty.\n")
                        self.run_btn.configure(state="normal")
                        return
                        
                    reg_cols = get_reg_col_indices(reg_rows[0]) if reg_rows else {}
                    sid_col = reg_cols.get('sheet_id', 2)
                    first_sheet_id = reg_rows[1][sid_col] if len(reg_rows[1]) > sid_col else reg_rows[1][2]
                    ss = gc.open_by_key(first_sheet_id)
                    worksheets = ss.worksheets()
                    
                    valid_tabs = []
                    for ws in worksheets:
                        p = parse_month_str(ws.title)
                        if p:
                            valid_tabs.append((p, ws.title))
                    
                    if not valid_tabs:
                        latest_tab = "None"
                        target_month, num_days = get_current_month_info(get_dhaka_today())
                    else:
                        valid_tabs.sort(key=lambda x: x[0])
                        latest_tab = valid_tabs[-1][1]
                        latest_dt = datetime.date(valid_tabs[-1][0][0], valid_tabs[-1][0][1], 1)
                        target_month, num_days = get_next_month_info(latest_dt)
                    rollover_info = {'latest_tab': latest_tab, 'next_month': target_month}
                else:
                    if not month_override:
                        target_month, num_days = get_current_month_info(get_dhaka_today())
                    else:
                        target_month = month_override.upper()
                        p = parse_month_str(target_month)
                        num_days = calendar.monthrange(p[0], p[1])[1] if p else 30

                # 1. Check existing sheets
                existing_found = check_existing_sheets_flow(gc, drive_service, zones, target_month)
                # 2. Discover/Generate Master Backups
                backups_info = get_or_create_master_backups(drive_service, gc)
                # 3. Check missing markets
                missing_fms, email_mappings, sh_by_zone, reg_ws = check_for_missing_markets(gc, drive_service, zones, backups_info)
                # 4. Check changed emails
                changed_items, em_reg_ws, em_reg_rows = check_for_changed_emails(gc, drive_service, zones, backups_info)

                # Open Unified Modal on main GUI thread
                dialog_event = threading.Event()
                dialog_res = {}
                def show_modal():
                    dlg = UnifiedDeltaDialog(self, mode, target_month, existing_found, missing_fms, changed_items, rollover_info)
                    self.wait_window(dlg)
                    dialog_res.update(dlg.result)
                    dialog_event.set()

                self.after(0, show_modal)
                dialog_event.wait()

                # Handle Dialog Selections
                if not dialog_res.get("confirmed"):
                    self.gui_log("Process cancelled by user from Unified Confirmation Modal.\n")
                    self.run_btn.configure(state="normal")
                    return

                if mode == "rollover" and not dialog_res.get("do_rollover", True):
                    self.gui_log("Rollover tab generation cancelled by user.\n")
                    self.run_btn.configure(state="normal")
                    return

                if "existing_action" in dialog_res:
                    existing_action = dialog_res["existing_action"]

                if dialog_res.get("do_missing") and missing_fms:
                    execute_missing_markets_provisioning(gc, drive_service, sheets_service, missing_fms, email_mappings, sh_by_zone, reg_ws, target_month, num_days, dry_run=dry_run)

                if dialog_res.get("do_emails") and changed_items:
                    execute_email_permissions_update(drive_service, changed_items, em_reg_ws, em_reg_rows, dry_run=dry_run)

                # Finally start main execution
                self.gui_log(f"\nStarting main process ({mode}) with existing action: [{existing_action.upper()}]...\n")
                t = threading.Thread(target=self.execute_process, args=(mode, zones, month_override, dry_run, existing_action))
                t.daemon = True
                t.start()
            except Exception as ex:
                self.gui_log(f"Error checking pre-requisites: {ex}\n")
                self.run_btn.configure(state="normal")

        t_check = threading.Thread(target=run_checks_and_confirm)
        t_check.daemon = True
        t_check.start()

    def execute_process(self, mode, zones, month_override, dry_run, existing_action="overwrite"):
        try:
            if mode == "provision":
                if not month_override:
                    # Auto compute current month (JUL'26)
                    month_str, num_days = get_current_month_info(datetime.date.today())
                else:
                    month_str = month_override.upper()
                    parsed = parse_month_str(month_str)
                    if parsed:
                        _, num_days = calendar.monthrange(parsed[0], parsed[1])
                    else:
                        num_days = 31 # default fallback
                run_provisioning(zones, month_str, num_days, dry_run=dry_run, existing_action=existing_action)
            elif mode == "rollover":
                run_rollover(zones, current_month_override=month_override if month_override else None, dry_run=dry_run)
            elif mode == "delete":
                if not month_override:
                    month_str, _ = get_current_month_info(datetime.date.today())
                else:
                    month_str = month_override.upper()
                execute_delete_month_tab(zones, month_str, dry_run=dry_run)
        except Exception as e:
            self.gui_log(f"\nFATAL SYSTEM ERROR: {e}\n")
        finally:
            self.run_btn.configure(state="normal")

if __name__ == "__main__":
    app = CashInHandApp()
    app.mainloop()
