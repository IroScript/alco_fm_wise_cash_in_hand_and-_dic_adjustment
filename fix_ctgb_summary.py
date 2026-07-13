import gspread
from google.oauth2.credentials import Credentials
from openpyxl.utils import get_column_letter
import json

creds = Credentials.from_authorized_user_file(
    r'FieldEdit\token.json',
    ['https://www.googleapis.com/auth/spreadsheets','https://www.googleapis.com/auth/drive']
)
gc = gspread.authorize(creds)

with open('zone_to_summary_sheets.json','r') as f:
    zs = json.load(f)

ctgb_id = zs.get('CTG.B')
ss = gc.open_by_key(ctgb_id)
ws = ss.get_worksheet(0)

sheet_url = 'https://docs.google.com/spreadsheets/d/11saoHVovwWjtvaq88rROlPQ2JBn5MDJBjhERUrFEFCE/edit'
fm_name = 'DEWAN JAHANGIR ALAM'

# Step 1: Clear bad column A entries
ws.batch_clear(['A5:A48'])
print('Cleared bad column A entries.')

# Step 2: Find correct column - after last used column in row 5
row5 = ws.row_values(5)
target_col = len(row5) + 1

print(f'Adding {fm_name} at column {target_col} ({get_column_letter(target_col)})...')

ws.update_cell(5, target_col, f'FM WISE TOTAL CASH IN HAND, {fm_name}')

cell_updates = []
for idx in range(31):
    r = 18 + idx
    formula = '=IMPORTRANGE("' + sheet_url + '", "JUL' + "'26!I" + str(r) + '")'
    cell_updates.append({
        'range': f'{get_column_letter(target_col)}{r}',
        'values': [[formula]]
    })
ws.batch_update(cell_updates, value_input_option='USER_ENTERED')
print(f'[OK] Added IMPORTRANGE at column {get_column_letter(target_col)} for {fm_name}.')
print('Other zones/sheets: UNTOUCHED.')
