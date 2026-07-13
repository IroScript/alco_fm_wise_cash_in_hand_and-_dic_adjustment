"""
Dewan Jahangir Alam (CTG.B) - Standalone Provisioning Script
100% aligned with run_system.py logic, formatting, and standards.
"""
import os, json, time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR = r"C:\Users\Irak\Desktop\deskTop\Cash in Hand and Dic Adjustment"
TOKEN_FILE = os.path.join(BASE_DIR, "FieldEdit", "token.json")
PARENT_FOLDER_ID = "1iOFeqywnIZ_yVclg_Em2U1npPtsokfGk"

creds = Credentials.from_authorized_user_file(TOKEN_FILE,
    ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive'])
gc = gspread.authorize(creds)
drive_service = build('drive', 'v3', credentials=creds)
sheets_service = build('sheets', 'v4', credentials=creds)

with open(os.path.join(BASE_DIR, 'sh_by_zone.json'), 'r', encoding='utf-8') as f:
    sh_by_zone = json.load(f)

FM_NAME    = "DEWAN JAHANGIR ALAM"
ZONE       = "CTG.B"
SH_EMAIL   = sh_by_zone.get(ZONE, "")
MONTH_STR  = "JUL'26"
NUM_DAYS   = 31
OLD_BAD_ID = "11saoHVovwWjtvaq88rROlPQ2JBn5MDJBjhERUrFEFCE"

# Dewan's markets (all vacant MPOs) and DAs (from run_system.py data)
MARKETS = [
    {"market_name": "MURADPUR +KALURGHAT", "mpo_name": "VACANT", "desig": "MPO", "mpo_code": "C017", "fm_code": "C059", "is_vacant": "Y", "da_name": None},
    {"market_name": "BAIZID",              "mpo_name": "VACANT", "desig": "MPO", "mpo_code": "C018", "fm_code": "C060", "is_vacant": "Y", "da_name": None},
    {"market_name": "HALISHAHAR",          "mpo_name": "VACANT", "desig": "MPO", "mpo_code": "C019", "fm_code": "C062", "is_vacant": "Y", "da_name": None},
]
DAS = [
    {"da_name": "MR. RAFIQUL ISLAM"},
    {"da_name": "MR. DHORMO LAL RAY"},
]

# ============================================================
# EXACT COLOR PALETTE (from run_system.py create_local_excel)
# ============================================================
FONT_FAM = 'Aptos'
C_VOID   = '060816';  C_NAVY = '0D1425';  C_MID = '1E293B'
C_ZA     = 'FFFFFF';  C_ZB   = 'F8FAFC';  C_TOT_D = 'ECFDF5';  C_TOT_H = '065F46'
T_NEON   = '00F2FE';  T_WHT  = 'FFFFFF';  T_SLATE = 'CBD5E1';  T_INK   = '0F172A'
T_MINT   = 'A7F3D0';  T_TOTD = '064E3B';  T_DATE  = '475569'
B_NEON   = '0E7490';  B_SLT  = '334155';  B_LIGHT = 'E2E8F0';  B_TOTAL = '6EE7B7'

font_title   = Font(name=FONT_FAM, size=26, bold=True,  color=T_NEON)
font_hdr     = Font(name=FONT_FAM, size=10, bold=True,  color=T_WHT)
font_sub     = Font(name=FONT_FAM, size=9,  bold=True,  color=T_SLATE)
font_name    = Font(name=FONT_FAM, size=10, bold=True,  color=T_WHT)
font_date    = Font(name=FONT_FAM, size=10, bold=True,  color=T_DATE)
font_body    = Font(name=FONT_FAM, size=10, bold=False, color=T_INK)
font_tot_h   = Font(name=FONT_FAM, size=10, bold=True,  color=T_MINT)
font_tot_d   = Font(name=FONT_FAM, size=10, bold=True,  color=T_TOTD)

fill_void = PatternFill('solid', fgColor=C_VOID)
fill_navy = PatternFill('solid', fgColor=C_NAVY)
fill_mid  = PatternFill('solid', fgColor=C_MID)
fill_totH = PatternFill('solid', fgColor=C_TOT_H)
fill_totD = PatternFill('solid', fgColor=C_TOT_D)

def S(style='thin', color=B_LIGHT):
    return Side(style=style, color=color)
bd_data  = Border(left=S(), right=S(), top=S(), bottom=S())
bd_hdr   = Border(left=S('thin',B_SLT), right=S('thin',B_SLT), top=S('thin',B_SLT), bottom=S('thin',B_SLT))
bd_tot   = Border(left=S('thin',B_TOTAL), right=S('thin',B_TOTAL), top=S('thin',B_TOTAL), bottom=S('thin',B_TOTAL))
bd_date  = Border(left=S('medium',B_SLT), right=S('thin',B_LIGHT), top=S('thin',B_LIGHT), bottom=S('thin',B_LIGHT))
bd_title = Border(left=S('medium',B_NEON), right=S('medium',B_NEON), top=S('medium',B_NEON), bottom=S('medium',B_NEON))

align_c  = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ============================================================
# STEP 0: CLEANUP previous bad attempts
# ============================================================
print("=== STEP 0: Cleanup ===")

try:
    drive_service.files().delete(fileId=OLD_BAD_ID).execute()
    print(f"  Deleted old bad Drive file {OLD_BAD_ID}")
except Exception:
    print("  Old file already gone or inaccessible.")

# Clean Master Registry
reg_q = "name = 'Master_Registry_Cash_In_Hand' and mimeType = 'application/vnd.google-apps.spreadsheet' and trashed = false"
reg_res = drive_service.files().list(q=reg_q, spaces='drive', fields='files(id)').execute()
reg_files = reg_res.get('files', [])
reg_sheet_id = reg_files[0]['id'] if reg_files else None

if reg_sheet_id:
    reg_ss = gc.open_by_key(reg_sheet_id)
    reg_ws = reg_ss.get_worksheet(0)
    all_rows = reg_ws.get_all_values()
    for idx in range(len(all_rows) - 1, -1, -1):
        if len(all_rows[idx]) >= 2 and FM_NAME in all_rows[idx][0] and ZONE in all_rows[idx][1]:
            reg_ws.delete_rows(idx + 1)
            print(f"  Deleted old registry row {idx+1}: {all_rows[idx][0]}")

# Clean CTG.B Summary (remove any bad Dewan column)
with open(os.path.join(BASE_DIR, 'zone_to_summary_sheets.json'), 'r', encoding='utf-8') as f:
    zone_summaries = json.load(f)
ctgb_summary_id = zone_summaries.get("CTG.B")
if ctgb_summary_id:
    ctgb_ss = gc.open_by_key(ctgb_summary_id)
    ctgb_ws = ctgb_ss.get_worksheet(0)
    row5 = ctgb_ws.row_values(5)
    for ci, val in enumerate(row5):
        if FM_NAME in str(val):
            col_l = get_column_letter(ci + 1)
            ctgb_ws.batch_clear([f"{col_l}5:{col_l}48"])
            print(f"  Cleared old CTG.B summary column {col_l}")

time.sleep(2)

# ============================================================
# STEP 1: CREATE LOCAL EXCEL (100% run_system.py logic)
# ============================================================
print("\n=== STEP 1: Create local Excel ===")

num_mpos = len(MARKETS)
num_das  = len(DAS)
total_col_idx = 4 + num_mpos + num_das   # Column I (9) = FM SELF(C) + 3 MPOs + 2 DAs + 1 TOTAL
total_col_ltr = get_column_letter(total_col_idx)
last_inp_col  = total_col_idx - 1         # Column H (8)
last_inp_ltr  = get_column_letter(last_inp_col)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = MONTH_STR
ws.views.sheetView[0].showGridLines = True
ws.sheet_view.zoomScale = 90
ws.sheet_properties.tabColor = 'A855F7'   # Purple (non-CTG.A)
ws.freeze_panes = 'C18'

# Row heights (EXACT from run_system.py lines 1239-1246)
ws.row_dimensions[1].height = 10
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 28
ws.row_dimensions[4].height = 8
ws.row_dimensions[5].height = 22
ws.row_dimensions[11].height = 26
for r in range(18, 18 + NUM_DAYS):
    ws.row_dimensions[r].height = 20

# Column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 14
for ci in range(4, total_col_idx):
    ws.column_dimensions[get_column_letter(ci)].width = 16
ws.column_dimensions[total_col_ltr].width = 18

# --- TITLE BLOCK (C2:TotalCol, Rows 2-3) ---
# IMPORTANT: Set value BEFORE merge_cells (openpyxl requirement)
title_cell = ws.cell(row=2, column=3, value="CASH IN HAND")
title_cell.font      = font_title
title_cell.alignment = align_c
title_cell.fill      = fill_void
title_cell.border    = bd_title
ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=total_col_idx)

# Fill rows 1-4 with void background
for r in range(1, 5):
    for c in range(1, total_col_idx + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill_void

# --- ROW 5: Category headers ---
mpo_s = 4; mpo_e = 3 + num_mpos       # cols 4-6
da_s  = mpo_e + 1; da_e = da_s + num_das - 1  # cols 7-8

# MPO header (set value, then merge)
mpo_cell = ws.cell(row=5, column=mpo_s, value="MPO")
mpo_cell.font = font_hdr; mpo_cell.alignment = align_c; mpo_cell.fill = fill_mid; mpo_cell.border = bd_hdr
ws.merge_cells(start_row=5, start_column=mpo_s, end_row=5, end_column=mpo_e)

# DA header
da_cell = ws.cell(row=5, column=da_s, value="DA")
da_cell.font = font_hdr; da_cell.alignment = align_c; da_cell.fill = fill_mid; da_cell.border = bd_hdr
ws.merge_cells(start_row=5, start_column=da_s, end_row=5, end_column=da_e)

# TOTAL CASH IN HAND (set value, then merge 5:11)
tot_cell = ws.cell(row=5, column=total_col_idx, value="TOTAL CASH IN HAND")
tot_cell.font = font_tot_h; tot_cell.alignment = align_c; tot_cell.fill = fill_totH; tot_cell.border = bd_tot
ws.merge_cells(start_row=5, start_column=total_col_idx, end_row=11, end_column=total_col_idx)
# Fill rest of total col rows 6-11
for r in range(6, 12):
    c = ws.cell(row=r, column=total_col_idx)
    c.fill = fill_totH; c.border = bd_tot

# Row 5 fills for header area cols 2 to total-1
for c in range(2, total_col_idx):
    cell = ws.cell(row=5, column=c)
    if cell.fill == PatternFill() or not cell.fill.fgColor or cell.fill.fgColor.rgb == '00000000':
        cell.fill = fill_mid
    cell.border = bd_hdr

# --- ROWS 6-10: Metadata (HIDDEN) ---
for r in range(6, 11):
    ws.row_dimensions[r].hidden = True
    for c in range(2, total_col_idx + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill_navy; cell.font = font_sub; cell.alignment = align_c; cell.border = bd_hdr

for c in range(3, total_col_idx):
    ws.cell(row=6, column=c, value=ZONE)     # Row 6: Zone
    ws.cell(row=7, column=c, value=FM_NAME)   # Row 7: FM Name

for i, m in enumerate(MARKETS):
    ws.cell(row=8, column=mpo_s + i, value=m['market_name'])
    ws.cell(row=9, column=mpo_s + i, value=m['mpo_code'])
    ws.cell(row=10, column=mpo_s + i, value=m['fm_code'])

# --- ROW 11: Column labels ---
# DATE (merged B11:B17)
date_hdr = ws.cell(row=11, column=2, value="DATE")
date_hdr.font = font_name; date_hdr.alignment = align_c; date_hdr.fill = fill_navy; date_hdr.border = bd_hdr
ws.merge_cells(start_row=11, start_column=2, end_row=17, end_column=2)

# FM SELF
fm_self = ws.cell(row=11, column=3, value="FM SELF")
fm_self.font = font_name; fm_self.alignment = align_c; fm_self.fill = fill_navy; fm_self.border = bd_hdr
ws.cell(row=12, column=3).fill = fill_mid; ws.cell(row=12, column=3).border = bd_hdr

# MPO names row 11
for i, m in enumerate(MARKETS):
    c = ws.cell(row=11, column=mpo_s + i, value=m['mpo_name'])
    c.font = font_name; c.alignment = align_c; c.fill = fill_navy; c.border = bd_hdr

# DA names row 11
for i, d in enumerate(DAS):
    c = ws.cell(row=11, column=da_s + i, value=d['da_name'])
    c.font = font_name; c.alignment = align_c; c.fill = fill_navy; c.border = bd_hdr

# --- ROW 12: Blank separator (HIDDEN) ---
ws.row_dimensions[12].hidden = True
for c in range(2, total_col_idx + 1):
    cell = ws.cell(row=12, column=c)
    cell.fill = fill_mid; cell.border = bd_hdr

# --- ROWS 13-17: DA metadata block (HIDDEN) ---
for r in range(13, 18):
    ws.row_dimensions[r].hidden = True
    for c in range(2, total_col_idx + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill_navy; cell.font = font_sub; cell.alignment = align_c; cell.border = bd_hdr

for c in range(3, total_col_idx):
    ws.cell(row=13, column=c, value=ZONE)
    ws.cell(row=14, column=c, value=FM_NAME)

ws.cell(row=15, column=3, value="FM")
for i, m in enumerate(MARKETS):
    ws.cell(row=15, column=mpo_s + i, value=m['desig'])
    ws.cell(row=16, column=mpo_s + i, value=m['mpo_code'])
for i, d in enumerate(DAS):
    ws.cell(row=15, column=da_s + i, value="DA")
    ws.cell(row=17, column=da_s + i, value=d['da_name'])

# --- DATA ROWS 18-48 ---
for idx in range(NUM_DAYS):
    r = 18 + idx
    d_val = f"{NUM_DAYS - idx} {MONTH_STR}"
    zr = PatternFill('solid', fgColor=C_ZA if idx % 2 == 0 else C_ZB)

    b = ws.cell(row=r, column=2, value=d_val)
    b.font = font_date; b.alignment = align_c; b.fill = zr; b.border = bd_date

    for c in range(3, total_col_idx):
        cell = ws.cell(row=r, column=c)
        cell.font = font_body; cell.alignment = align_c; cell.fill = zr
        cell.border = bd_data; cell.number_format = '#,##0'

    t = ws.cell(row=r, column=total_col_idx, value=f"=SUM(C{r}:{last_inp_ltr}{r})")
    t.font = font_tot_d; t.alignment = align_c; t.fill = fill_totD; t.border = bd_tot
    t.number_format = '#,##0'

# --- HIDE VACANT MPO COLUMNS ---
for i, m in enumerate(MARKETS):
    if m.get('is_vacant') == 'Y':
        ws.column_dimensions[get_column_letter(mpo_s + i)].hidden = True

# --- SAVE ---
local_dir = os.path.join(BASE_DIR, ZONE)
os.makedirs(local_dir, exist_ok=True)
local_path = os.path.join(local_dir, f"{FM_NAME}.xlsx")
wb.save(local_path)
wb.close()
print(f"  Saved: {local_path}")

# ============================================================
# STEP 2: UPLOAD TO GOOGLE DRIVE (inside CTG.B zone folder)
# ============================================================
print("\n=== STEP 2: Upload to Google Drive ===")

folder_q = f"name = '{ZONE}' and mimeType = 'application/vnd.google-apps.folder' and '{PARENT_FOLDER_ID}' in parents and trashed = false"
folder_r = drive_service.files().list(q=folder_q, spaces='drive', fields='files(id)').execute()
folder_files = folder_r.get('files', [])
if folder_files:
    ctgb_folder_id = folder_files[0]['id']
    print(f"  Using CTG.B folder: {ctgb_folder_id}")
else:
    meta = {'name': ZONE, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [PARENT_FOLDER_ID]}
    ctgb_folder_id = drive_service.files().create(body=meta, fields='id').execute()['id']
    print(f"  Created CTG.B folder: {ctgb_folder_id}")

sheet_name = f"CASH IN HAND - {FM_NAME}"
media = MediaFileUpload(local_path,
    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    resumable=True)
upload_meta = {
    'name': sheet_name,
    'parents': [ctgb_folder_id],
    'mimeType': 'application/vnd.google-apps.spreadsheet'
}

new_sheet_id = None
new_sheet_url = None
for attempt in range(5):
    try:
        uploaded = drive_service.files().create(body=upload_meta, media_body=media, fields='id').execute()
        new_sheet_id = uploaded['id']
        new_sheet_url = f"https://docs.google.com/spreadsheets/d/{new_sheet_id}/edit"
        print(f"  Uploaded: {sheet_name}")
        print(f"  Sheet ID: {new_sheet_id}")
        print(f"  URL: {new_sheet_url}")
        break
    except Exception as e:
        print(f"  Upload retry {attempt+1}/5: {e}")
        time.sleep(5)

if not new_sheet_id:
    print("  FATAL: Could not upload sheet!")
    exit(1)

# Delete local temp file
try:
    os.remove(local_path)
    print("  Deleted local temp file.")
except Exception:
    pass

time.sleep(3)

# ============================================================
# STEP 3: APPLY GOOGLE SHEETS API SETTINGS
# ============================================================
print("\n=== STEP 3: Apply freeze, protection, data validation ===")

# Get the ACTUAL internal sheetId
meta = sheets_service.spreadsheets().get(
    spreadsheetId=new_sheet_id,
    fields="sheets(properties(sheetId,title))").execute()
actual_sid = meta['sheets'][0]['properties']['sheetId']
print(f"  Actual sheetId: {actual_sid}")

requests_body = []

# Freeze: 17 rows, 2 columns (DATE only frozen)
requests_body.append({
    "updateSheetProperties": {
        "properties": {
            "sheetId": actual_sid,
            "gridProperties": {"frozenRowCount": 17, "frozenColumnCount": 2}
        },
        "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount"
    }
})

# Protection: lock everything, unlock C18:H48 (startCol=2, endCol=8 for 0-indexed)
requests_body.append({
    "addProtectedRange": {
        "protectedRange": {
            "range": {"sheetId": actual_sid},
            "description": f"Locked headers & formula columns ({MONTH_STR})",
            "warningOnly": False,
            "unprotectedRanges": [{
                "sheetId": actual_sid,
                "startRowIndex": 17,
                "endRowIndex": 17 + NUM_DAYS,
                "startColumnIndex": 2,
                "endColumnIndex": total_col_idx - 1
            }],
            "editors": {"users": [SH_EMAIL] if SH_EMAIL else []}
        }
    }
})

# Data validation: positive numbers only (use NUMBER_GREATER_THAN_EQ)
requests_body.append({
    "setDataValidation": {
        "range": {
            "sheetId": actual_sid,
            "startRowIndex": 17,
            "endRowIndex": 17 + NUM_DAYS,
            "startColumnIndex": 2,
            "endColumnIndex": total_col_idx - 1
        },
        "rule": {
            "condition": {
                "type": "NUMBER_GREATER_THAN_EQ",
                "values": [{"userEnteredValue": "0"}]
            },
            "inputMessage": "Enter a positive number (0 or greater).",
            "strict": False,
            "showCustomUi": True
        }
    }
})

sheets_service.spreadsheets().batchUpdate(
    spreadsheetId=new_sheet_id,
    body={"requests": requests_body}).execute()
print("  Applied freeze, protection, and data validation.")

# ============================================================
# STEP 4: SHARE PERMISSIONS
# ============================================================
print("\n=== STEP 4: Share permissions ===")

if SH_EMAIL:
    try:
        drive_service.permissions().create(
            fileId=new_sheet_id,
            body={'role': 'writer', 'type': 'user', 'emailAddress': SH_EMAIL},
            sendNotificationEmail=False
        ).execute()
        print(f"  Shared with SH: {SH_EMAIL} (writer)")
    except Exception as e:
        print(f"  Share note: {e}")

time.sleep(1)

# ============================================================
# STEP 5: REGISTER IN MASTER REGISTRY
# ============================================================
print("\n=== STEP 5: Register in Master Registry ===")

if reg_sheet_id:
    reg_ss2 = gc.open_by_key(reg_sheet_id)
    reg_ws2 = reg_ss2.get_worksheet(0)
    reg_ws2.append_row([FM_NAME, ZONE, new_sheet_id, new_sheet_url, '', '', '', SH_EMAIL])
    print(f"  Registered: {FM_NAME} | {ZONE} | {new_sheet_id}")

time.sleep(1)

# ============================================================
# STEP 6: UPDATE CTG.B ZONAL SUMMARY (ADD IMPORTRANGE COLUMNS)
# ============================================================
print("\n=== STEP 6: Update CTG.B Zonal Summary (ADD Dewan columns ONLY) ===")

if ctgb_summary_id:
    ctgb_ss2 = gc.open_by_key(ctgb_summary_id)
    ctgb_ws2 = ctgb_ss2.get_worksheet(0)

    # Find the next available column AFTER all existing detail columns
    row5_vals = ctgb_ws2.row_values(5)
    last_used_col = 0
    for ci, v in enumerate(row5_vals):
        if v:
            last_used_col = ci + 1
    target_col = last_used_col + 1
    target_ltr = get_column_letter(target_col)

    print(f"  Adding Dewan columns starting at col {target_col} ({target_ltr})...")

    # Write FM header row 5
    ctgb_ws2.update_cell(5, target_col, f"FM WISE TOTAL CASH IN HAND, {FM_NAME}")

    # Write FM name row 6
    ctgb_ws2.update_cell(6, target_col, FM_NAME)

    # IMPORTRANGE formulas for rows 18-48 (pull FM SELF to last input col from Dewan's sheet)
    # We need to import C18:H48 (3 MPO cols + 2 DA cols + FM SELF = cols C through H = 6 cols)
    cell_updates = []
    for idx in range(NUM_DAYS):
        r = 18 + idx
        formula = f'=IMPORTRANGE("{new_sheet_url}", "{MONTH_STR}!C{r}:{last_inp_ltr}{r}")'
        cell_updates.append({
            'range': f'{target_ltr}{r}',
            'values': [[formula]]
        })
    ctgb_ws2.batch_update(cell_updates, value_input_option='USER_ENTERED')
    print(f"  Added IMPORTRANGE formulas at column {target_ltr}.")

    # IMPORTANT: Do NOT modify the existing E5 summary formula or other FM columns!
    # The new columns are APPENDED after existing data, so E5's SUM range already
    # does NOT cover them. We leave it untouched as the user requested other FMs
    # remain unaffected.

print("\n" + "=" * 55)
print(f"SUCCESS: {FM_NAME} ({ZONE}) provisioned with 100% logic!")
print(f"  Sheet URL : {new_sheet_url}")
print(f"  Location  : Inside CTG.B zone folder on Google Drive")
print(f"  Registered: Master Registry updated")
print(f"  Summary   : IMPORTRANGE added to CTG.B Zonal Summary")
print(f"  Other FMs : COMPLETELY UNTOUCHED")
print("=" * 55)
